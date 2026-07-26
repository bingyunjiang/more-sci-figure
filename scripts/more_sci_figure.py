#!/usr/bin/env python3
"""more-sci-figure v0.3 统一中文本地命令行。"""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import math
import os
import secrets
import sys
from datetime import datetime, timezone
from collections import deque
from html import escape
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from typing import Any, Callable
from urllib.parse import parse_qs, urlparse

import numpy as np
from PIL import Image, ImageDraw


SCHEMA = "more-sci-figure.project.v1"
MANIFEST_SCHEMA = "more-sci-figure.manifest.v1"
REVIEW_SCHEMA = "more-sci-figure.review-decisions.v1"
ASSESSMENT_SCHEMA = "more-sci-figure.review-assessment.v1"
SPEC_REVIEW_SCHEMA = "more-sci-figure.spec-review.v1"
SPEC_CONFIRMATION_SCHEMA = "more-sci-figure.spec-confirmation.v1"
SUPPORTED_CHARTS = {"line", "polar_line", "scatter", "bar", "histogram"}
VERSION = "0.3.1"
ACCEPTANCE_PROFILES = {
    "exploratory": {
        "label": "趋势预览",
        "overall_threshold": 85.0,
        "minimum_dimension_score": 75.0,
    },
    "engineering": {
        "label": "工程分析",
        "overall_threshold": 90.0,
        "minimum_dimension_score": 85.0,
    },
    "publication": {
        "label": "论文定量数据",
        "overall_threshold": 95.0,
        "minimum_dimension_score": 90.0,
    },
}


class FigureError(RuntimeError):
    pass


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def write_json(path: Path, payload: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    temporary = path.with_suffix(path.suffix + ".tmp")
    temporary.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2, sort_keys=True) + "\n",
        encoding="utf-8",
    )
    temporary.replace(path)


def read_json(path: Path) -> dict[str, Any]:
    return json.loads(path.read_text(encoding="utf-8-sig"))


def resolve_from(base_file: Path, value: str) -> Path:
    path = Path(value).expanduser()
    return path.resolve() if path.is_absolute() else (base_file.parent / path).resolve()


def project_copy_with_rebased_paths(
    spec: dict[str, Any], source_spec_path: Path, destination: Path
) -> dict[str, Any]:
    """Keep source semantics stable when a project spec is copied to a new directory."""
    copied = json.loads(json.dumps(spec, ensure_ascii=False))
    source = copied.get("source")
    if not isinstance(source, dict):
        return copied
    for key in ("path", "measurement_raster"):
        value = source.get(key)
        if not value:
            continue
        resolved = resolve_from(source_spec_path, str(value))
        source[key] = os.path.relpath(resolved, destination.parent.resolve())
    return copied


def parse_hex_color(value: str) -> tuple[int, int, int]:
    text = value.strip().lstrip("#")
    if len(text) != 6:
        raise FigureError(f"颜色必须采用 #RRGGBB 格式：{value}")
    try:
        return tuple(int(text[index : index + 2], 16) for index in (0, 2, 4))
    except ValueError as exc:
        raise FigureError(f"颜色值无效：{value}") from exc


def source_metadata(path: Path) -> dict[str, Any]:
    if not path.is_file():
        raise FigureError(f"找不到来源文件：{path}")
    suffix = path.suffix.lower()
    result: dict[str, Any] = {
        "path": str(path),
        "sha256": sha256_file(path),
        "bytes": path.stat().st_size,
        "suffix": suffix,
    }
    if suffix in {".png", ".jpg", ".jpeg", ".tif", ".tiff", ".bmp", ".webp"}:
        with Image.open(path) as image:
            result.update(
                {
                    "kind": "raster",
                    "width": image.width,
                    "height": image.height,
                    "mode": image.mode,
                }
            )
    elif suffix == ".pdf":
        try:
            import fitz
        except ImportError as exc:
            raise FigureError("检查 PDF 需要安装 PyMuPDF") from exc
        with fitz.open(path) as document:
            result.update(
                {
                    "kind": "pdf",
                    "pages": document.page_count,
                    "page_sizes_points": [
                        [float(page.rect.width), float(page.rect.height)] for page in document
                    ],
                }
            )
    elif suffix in {".csv", ".tsv", ".json", ".xlsx", ".xlsm"}:
        result["kind"] = "data"
    else:
        result["kind"] = "unknown"
    return result


def render_pdf_page(path: Path, page_number: int, output: Path, dpi: int = 144) -> dict[str, Any]:
    try:
        import fitz
    except ImportError as exc:
        raise FigureError("渲染 PDF 页面需要安装 PyMuPDF") from exc
    with fitz.open(path) as document:
        if page_number < 1 or page_number > document.page_count:
            raise FigureError(f"PDF 页码必须位于 1 到 {document.page_count} 之间")
        page = document[page_number - 1]
        scale = dpi / 72.0
        pixmap = page.get_pixmap(matrix=fitz.Matrix(scale, scale), alpha=False)
        output.parent.mkdir(parents=True, exist_ok=True)
        pixmap.save(output)
    with Image.open(output) as image:
        return {
            "path": str(output),
            "sha256": sha256_file(output),
            "width": image.width,
            "height": image.height,
            "dpi": dpi,
            "page": page_number,
        }


def extract_pdf_embedded_image(
    path: Path, page_number: int, image_index: int, output_stem: Path
) -> dict[str, Any]:
    """Extract one declared PDF image object without rasterizing or rescaling the page."""
    try:
        import fitz
    except ImportError as exc:
        raise FigureError("提取 PDF 嵌入图像需要安装 PyMuPDF") from exc
    with fitz.open(path) as document:
        if page_number < 1 or page_number > document.page_count:
            raise FigureError(f"PDF 页码必须位于 1 到 {document.page_count} 之间")
        images = document[page_number - 1].get_images(full=True)
        if not images:
            raise FigureError(f"PDF 第 {page_number} 页没有可提取的嵌入栅格对象")
        if image_index < 0 or image_index >= len(images):
            raise FigureError(
                f"PDF 第 {page_number} 页的 image_index 必须位于 0 到 {max(0, len(images) - 1)} 之间"
            )
        xref = int(images[image_index][0])
        extracted = document.extract_image(xref)
    extension = str(extracted.get("ext") or "bin").lower()
    output = output_stem.with_suffix(f".{extension}")
    output.parent.mkdir(parents=True, exist_ok=True)
    output.write_bytes(extracted["image"])
    try:
        with Image.open(output) as image:
            width, height = image.size
            mode = image.mode
    except Exception as exc:
        raise FigureError(f"PDF 嵌入对象 {xref} 不是可测量栅格：{exc}") from exc
    return {
        "path": str(output),
        "sha256": sha256_file(output),
        "width": int(width),
        "height": int(height),
        "mode": mode,
        "page": page_number,
        "pdf_image_index": image_index,
        "pdf_xref": xref,
        "encoding": extension,
        "measurement_kind": "pdf_embedded_raster",
    }


def inspect_command(
    input_path: Path,
    chart_type: str | None,
    out_dir: Path,
    page: int,
    *,
    dpi: int = 144,
    pdf_image_index: int | None = None,
) -> None:
    input_path = input_path.expanduser().resolve()
    out_dir.mkdir(parents=True, exist_ok=True)
    metadata = source_metadata(input_path)
    if dpi <= 0:
        raise FigureError("PDF 页面渲染 DPI 必须为正整数")
    if pdf_image_index is not None and metadata["kind"] != "pdf":
        raise FigureError("--pdf-image-index 只适用于 PDF 输入")
    report: dict[str, Any] = {
        "schema": "more-sci-figure.source-report.v1",
        "tool_version": VERSION,
        "source": metadata,
        "chart_type_proposal": chart_type,
        "numeric_extraction_authorized": False,
        "required_next": ["确认图表类型", "确认绘图区", "确认坐标轴锚点"],
        "说明": "检查阶段不会授权数值提取；请先完成人工确认。",
    }
    measurement: dict[str, Any] | None = None
    if metadata["kind"] == "pdf":
        if pdf_image_index is None:
            measurement = render_pdf_page(input_path, page, out_dir / "source-page.png", dpi=dpi)
        else:
            measurement = extract_pdf_embedded_image(
                input_path, page, pdf_image_index, out_dir / "source-image"
            )
        report["measurement_raster"] = measurement
    elif metadata["kind"] == "raster":
        measurement = {
            "path": str(input_path),
            "sha256": metadata["sha256"],
            "width": metadata["width"],
            "height": metadata["height"],
        }
        report["measurement_raster"] = measurement
    write_json(out_dir / "source-report.json", report)

    relative_source = os.path.relpath(input_path, out_dir.resolve())
    measurement_path = Path(str(measurement["path"])) if measurement else input_path
    chart: dict[str, Any] = {
        "type": chart_type or "unknown",
        "plot_box": [0, 0, 0, 0],
        "x_axis": {
            "scale": "linear",
            "anchors": [{"pixel": 0, "value": 0}, {"pixel": 0, "value": 1}],
        },
        "y_axis": {
            "scale": "linear",
            "anchors": [{"pixel": 0, "value": 0}, {"pixel": 0, "value": 1}],
        },
        "series": [
            {
                "id": "series-a",
                "color": "#d62728",
                "color_space": "lab",
                "tolerance": 12,
                "min_area": 4,
                "max_area": 500,
                "min_fill_ratio": 0.25,
                "max_aspect_ratio": 4.0,
                "min_rectangularity": 0.8,
                "baseline_tolerance": 3.0,
            }
        ],
    }
    if chart_type == "polar_line":
        chart.pop("x_axis")
        chart.pop("y_axis")
        chart["polar"] = {
            "center_px": [0, 0],
            "angle_axis": {
                "unit": "degree",
                "zero_bearing_deg": 0,
                "direction": "clockwise",
                "anchors": [
                    {"pixel": [0, 0], "value": 0},
                    {"pixel": [0, 0], "value": 90},
                ],
            },
            "radius_axis": {
                "scale": "linear",
                "anchors": [{"pixel": 0, "value": 0}, {"pixel": 0, "value": 1}],
            },
            "inner_radius_px": 0,
            "outer_radius_px": 0,
            "angle_start_deg": 0,
            "angle_end_deg": 360,
            "angle_step_deg": 1,
        }
        chart["series"][0]["extraction_mode"] = "polar_radial"
        chart["series"][0]["angular_half_width_deg"] = 0.75
    project = {
        "schema": SCHEMA,
        "project_id": input_path.stem,
        "source": {
            "path": relative_source,
            "sha256": metadata["sha256"],
            "page": page if metadata["kind"] == "pdf" else None,
            "measurement_raster": os.path.relpath(measurement_path, out_dir.resolve()),
            "measurement_sha256": measurement["sha256"] if measurement else None,
            "pdf_image_index": pdf_image_index,
            "pdf_xref": measurement.get("pdf_xref") if measurement else None,
        },
        "chart": chart,
        "quality_gates": {
            "calibration": {
                "max_normalized_rmse": None,
                "require_three_anchors": False
            },
            "line": {
                "min_coverage": 0.5,
                "max_gap_fraction": None
            },
            "scatter": {
                "min_accepted_components": 1,
                "max_rejected_ratio": None
            },
            "bar": {
                "min_accepted_components": 1,
                "max_rejected_ratio": None
            }
        },
        "render": {
            "plot_type": chart_type or "line",
            "x": "x",
            "y": "y",
            "group": "series",
            "x_label": "",
            "y_label": "",
            "title": "",
        },
    }
    write_json(out_dir / "project.json", project)
    print(
        json.dumps(
            {
                "status": "pass",
                "说明": "已锁定来源并生成项目模板，尚未授权数值提取。",
                "report": str(out_dir / "source-report.json"),
                "spec": str(out_dir / "project.json"),
            },
            ensure_ascii=False,
        )
    )


def validate_spec(spec: dict[str, Any], spec_path: Path, *, extraction: bool = True) -> list[str]:
    errors: list[str] = []
    if spec.get("schema") != SCHEMA:
        errors.append(f"schema 必须是 {SCHEMA}")
    source = spec.get("source")
    if not isinstance(source, dict) or not source.get("path"):
        errors.append("必须填写 source.path")
    else:
        path = resolve_from(spec_path, str(source["path"]))
        if not path.is_file():
            errors.append(f"找不到来源文件：{path}")
        elif source.get("sha256") != sha256_file(path):
            errors.append("source.sha256 与当前来源文件不一致")
    chart = spec.get("chart")
    if not isinstance(chart, dict):
        errors.append("必须提供 chart 对象")
        return errors
    chart_type = chart.get("type")
    if extraction and chart_type not in SUPPORTED_CHARTS:
        errors.append(f"不支持的 chart.type：{chart_type}")
    box = chart.get("plot_box")
    if extraction and (
        not isinstance(box, list)
        or len(box) != 4
        or not all(
            isinstance(value, (int, float)) and math.isfinite(float(value))
            for value in box
        )
        or box[2] <= box[0]
        or box[3] <= box[1]
    ):
        errors.append("chart.plot_box 必须是具有正面积的 [left, top, right, bottom]")
    if extraction:
        axis_names = () if chart_type == "polar_line" else ("x_axis", "y_axis")
        for axis_name in axis_names:
            axis = chart.get(axis_name)
            anchors = axis.get("anchors") if isinstance(axis, dict) else None
            if not isinstance(anchors, list) or len(anchors) < 2:
                errors.append(f"chart.{axis_name} 至少需要两个锚点")
                continue
            pixels = [anchor.get("pixel") for anchor in anchors if isinstance(anchor, dict)]
            values = [anchor.get("value") for anchor in anchors if isinstance(anchor, dict)]
            numeric_pixels = [
                float(pixel)
                for pixel in pixels
                if isinstance(pixel, (int, float)) and math.isfinite(float(pixel))
            ]
            if len(numeric_pixels) != len(pixels):
                errors.append(f"chart.{axis_name} 的锚点像素必须是有限数值")
            elif len(set(numeric_pixels)) < 2:
                errors.append(f"chart.{axis_name} 的锚点像素位置必须不同")
            if any(not isinstance(value, (int, float)) for value in values):
                errors.append(f"chart.{axis_name} 的锚点值必须是有限数值")
            numeric_values = [
                float(value)
                for value in values
                if isinstance(value, (int, float)) and math.isfinite(float(value))
            ]
            if len(numeric_values) != len(values):
                errors.append(f"chart.{axis_name} 的锚点值不能包含 NaN 或无穷值")
            elif len(set(numeric_values)) < 2:
                errors.append(f"chart.{axis_name} 的锚点数值必须至少包含两个不同值")
            if axis.get("scale", "linear") == "log10" and any(value <= 0 for value in values if isinstance(value, (int, float))):
                errors.append(f"chart.{axis_name} 的 log10 锚点必须为正数")
            if isinstance(axis, dict) and axis.get("scale", "linear") not in {"linear", "log10"}:
                errors.append(f"chart.{axis_name} 只支持 linear 或 log10")
        series = chart.get("series")
        if not isinstance(series, list) or not series:
            errors.append("chart.series 至少需要一个系列")
        else:
            for entry in series:
                if not isinstance(entry, dict):
                    errors.append("chart.series 中每个系列都必须是对象")
                    continue
                try:
                    parse_hex_color(str(entry.get("color", "")))
                except FigureError as exc:
                    errors.append(str(exc))
                if str(entry.get("color_space", "rgb")).lower() not in {"rgb", "lab"}:
                    errors.append(f"系列 {entry.get('id', '')} 的 color_space 只支持 rgb 或 lab")
                tolerance = entry.get("tolerance", 45)
                if (
                    not isinstance(tolerance, (int, float))
                    or not math.isfinite(float(tolerance))
                    or float(tolerance) <= 0
                ):
                    errors.append(f"系列 {entry.get('id', '')} 的 tolerance 必须为正数")
                localization_half_width = entry.get(
                    "center_localization_half_width_px", 0.5
                )
                if (
                    not isinstance(localization_half_width, (int, float))
                    or not math.isfinite(float(localization_half_width))
                    or float(localization_half_width) <= 0
                    or float(localization_half_width) > 10
                ):
                    errors.append(
                        f"系列 {entry.get('id', '')} 的 "
                        "center_localization_half_width_px 必须位于 0 到 10 像素之间"
                    )
                extraction_mode = str(entry.get("extraction_mode", "color_column_median"))
                if extraction_mode not in {
                    "color_column_median",
                    "guided_path",
                    "guided_group_path",
                    "marker_centers",
                    "polar_radial",
                }:
                    errors.append(
                        f"系列 {entry.get('id', '')} 的 extraction_mode 不受支持"
                    )
                if chart_type == "polar_line" and extraction_mode != "polar_radial":
                    errors.append(
                        f"极坐标系列 {entry.get('id', '')} 的 extraction_mode 必须为 polar_radial"
                    )
                curve_topology = str(entry.get("curve_topology", "continuous"))
                if chart_type in {"line", "polar_line"} and curve_topology not in {
                    "continuous",
                    "segmented",
                }:
                    errors.append(
                        f"系列 {entry.get('id', '')} 的 curve_topology "
                        "只支持 continuous 或 segmented"
                    )
                curve_data_mode = str(entry.get("curve_data_mode", "observations"))
                if chart_type in {"line", "polar_line"} and curve_data_mode not in {
                    "observations",
                    "guide_constrained",
                }:
                    errors.append(
                        f"系列 {entry.get('id', '')} 的 curve_data_mode "
                        "只支持 observations 或 guide_constrained"
                    )
                if curve_data_mode == "guide_constrained":
                    if chart_type != "line":
                        errors.append(
                            f"系列 {entry.get('id', '')} 的 guide_constrained 正式数据"
                            "目前只支持笛卡尔曲线"
                        )
                    if curve_topology != "continuous":
                        errors.append(
                            f"系列 {entry.get('id', '')} 的 guide_constrained 正式数据"
                            "要求 curve_topology=continuous"
                        )
                    if extraction_mode not in {"guided_path", "guided_group_path"}:
                        errors.append(
                            f"系列 {entry.get('id', '')} 的 guide_constrained 正式数据"
                            "要求 guided_path 或 guided_group_path 提取"
                        )
                    for key, default in (
                        ("curve_data_max_residual_px", 5.0),
                        ("curve_data_residual_smoothing_window", 21),
                    ):
                        value = entry.get(key, default)
                        invalid = (
                            not isinstance(value, (int, float))
                            or not math.isfinite(float(value))
                            or float(value) <= 0
                        )
                        if key.endswith("window"):
                            invalid = invalid or not isinstance(value, int)
                        if invalid:
                            errors.append(
                                f"系列 {entry.get('id', '')} 的 {key} 必须为正数"
                            )
                angular_half_width = entry.get("angular_half_width_deg")
                if angular_half_width is not None and (
                    not isinstance(angular_half_width, (int, float))
                    or not math.isfinite(float(angular_half_width))
                    or float(angular_half_width) <= 0
                    or float(angular_half_width) > 10
                ):
                    errors.append(
                        f"系列 {entry.get('id', '')} 的 angular_half_width_deg 必须位于 0 到 10 度之间"
                    )
                if extraction_mode in {"guided_path", "guided_group_path", "marker_centers"}:
                    guide_points = parse_guide_points(entry)
                    if len(guide_points) < 2:
                        errors.append(f"系列 {entry.get('id', '')} 的 guided_path 至少需要两个 guide_points_px")
                    elif len({point[0] for point in guide_points}) != len(guide_points):
                        errors.append(f"系列 {entry.get('id', '')} 的 guide_points_px 横坐标必须唯一")
                    corridor = entry.get("guide_corridor_px", 7.0)
                    if (
                        not isinstance(corridor, (int, float))
                        or not math.isfinite(float(corridor))
                        or float(corridor) <= 0
                    ):
                        errors.append(f"系列 {entry.get('id', '')} 的 guide_corridor_px 必须为正数")
                    x_range = entry.get("x_pixel_range")
                    if x_range is not None and (
                        not isinstance(x_range, list)
                        or len(x_range) != 2
                        or not all(
                            isinstance(value, (int, float)) and math.isfinite(float(value))
                            for value in x_range
                        )
                        or float(x_range[0]) >= float(x_range[1])
                    ):
                        errors.append(
                            f"系列 {entry.get('id', '')} 的 x_pixel_range 必须是递增的两个有限数值"
                        )
                    if str(entry.get("guide_interpolation", "linear")) not in {
                        "linear",
                        "shape_preserving",
                    }:
                        errors.append(
                            f"系列 {entry.get('id', '')} 的 guide_interpolation 只支持 linear 或 shape_preserving"
                        )
                    if str(entry.get("line_semantics", "solid")) not in {"solid", "dashed"}:
                        errors.append(
                            f"系列 {entry.get('id', '')} 的 line_semantics 只支持 solid 或 dashed"
                        )
                    for key in (
                        "missing_penalty",
                        "continuity_weight",
                        "search_half_width_px",
                        "max_track_gap_px",
                    ):
                        value = entry.get(key)
                        if value is not None and (
                            not isinstance(value, (int, float))
                            or not math.isfinite(float(value))
                            or float(value) < 0
                        ):
                            errors.append(f"系列 {entry.get('id', '')} 的 {key} 必须为非负数")
                if extraction_mode == "marker_centers":
                    minimum_markers = entry.get("min_marker_count", 1)
                    if (
                        not isinstance(minimum_markers, int)
                        or isinstance(minimum_markers, bool)
                        or minimum_markers <= 0
                    ):
                        errors.append(
                            f"系列 {entry.get('id', '')} 的 min_marker_count 必须为正整数"
                        )
                    for key, default in (
                        ("marker_min_span_px", 6.0),
                        ("marker_min_width_px", 2.0),
                        ("marker_window_radius_px", 7.0),
                        ("marker_min_distance_px", 8.0),
                    ):
                        value = entry.get(key, default)
                        if (
                            not isinstance(value, (int, float))
                            or not math.isfinite(float(value))
                            or float(value) <= 0
                        ):
                            errors.append(f"系列 {entry.get('id', '')} 的 {key} 必须为正数")
                    marker_shape = str(entry.get("marker_shape", "unspecified"))
                    if marker_shape not in {"square", "triangle", "diamond", "circle", "x", "unspecified"}:
                        errors.append(
                            f"系列 {entry.get('id', '')} 的 marker_shape 不受支持"
                        )
                for box in entry.get("exclude_boxes_px", []):
                    if (
                        not isinstance(box, list)
                        or len(box) != 4
                        or not all(
                            isinstance(value, (int, float)) and math.isfinite(float(value))
                            for value in box
                        )
                        or float(box[0]) > float(box[2])
                        or float(box[1]) > float(box[3])
                    ):
                        errors.append(f"系列 {entry.get('id', '')} 的 exclude_boxes_px 必须是有效 [x0,y0,x1,y1]")
            grouped: dict[str, list[dict[str, Any]]] = {}
            for entry in series:
                if isinstance(entry, dict) and entry.get("extraction_mode") == "guided_group_path":
                    grouped.setdefault(
                        str(entry.get("shared_color_group", entry.get("color"))), []
                    ).append(entry)
            for group_id, group_entries in grouped.items():
                if len(group_entries) < 2:
                    errors.append(f"guided_group_path 颜色组 {group_id!r} 至少需要两个系列")
                if len({str(item.get("color")) for item in group_entries}) != 1:
                    errors.append(f"guided_group_path 颜色组 {group_id!r} 的系列必须使用相同颜色")

        if chart_type == "polar_line":
            polar = chart.get("polar")
            if not isinstance(polar, dict):
                errors.append("polar_line 必须提供 chart.polar")
            else:
                center = polar.get("center_px")
                if (
                    not isinstance(center, list)
                    or len(center) != 2
                    or not all(
                        isinstance(value, (int, float)) and math.isfinite(float(value))
                        for value in center
                    )
                ):
                    errors.append("chart.polar.center_px 必须是两个有限像素坐标")
                inner = polar.get("inner_radius_px")
                outer = polar.get("outer_radius_px")
                if (
                    not isinstance(inner, (int, float))
                    or not isinstance(outer, (int, float))
                    or not math.isfinite(float(inner))
                    or not math.isfinite(float(outer))
                    or float(inner) < 0
                    or float(outer) <= float(inner)
                ):
                    errors.append("chart.polar 的内外半径必须满足 0 <= inner_radius_px < outer_radius_px")
                angle_axis = polar.get("angle_axis")
                angle_anchors = angle_axis.get("anchors") if isinstance(angle_axis, dict) else None
                if not isinstance(angle_anchors, list) or len(angle_anchors) < 2:
                    errors.append("chart.polar.angle_axis 至少需要两个角度锚点")
                else:
                    valid_angle_values: list[float] = []
                    for anchor in angle_anchors:
                        pixel = anchor.get("pixel") if isinstance(anchor, dict) else None
                        value = anchor.get("value") if isinstance(anchor, dict) else None
                        if (
                            not isinstance(pixel, list)
                            or len(pixel) != 2
                            or not all(
                                isinstance(item, (int, float)) and math.isfinite(float(item))
                                for item in pixel
                            )
                            or not isinstance(value, (int, float))
                            or not math.isfinite(float(value))
                        ):
                            errors.append("chart.polar.angle_axis 锚点必须包含有限 pixel=[x,y] 和 value")
                        else:
                            valid_angle_values.append(float(value))
                    if len(set(valid_angle_values)) < 2:
                        errors.append("chart.polar.angle_axis 锚点值必须至少包含两个不同角度")
                if isinstance(angle_axis, dict):
                    if angle_axis.get("unit", "degree") != "degree":
                        errors.append("chart.polar.angle_axis.unit 只支持 degree")
                    if angle_axis.get("direction", "clockwise") not in {
                        "clockwise",
                        "counterclockwise",
                    }:
                        errors.append("chart.polar.angle_axis.direction 只支持 clockwise 或 counterclockwise")
                    zero = angle_axis.get("zero_bearing_deg", 0)
                    if not isinstance(zero, (int, float)) or not math.isfinite(float(zero)):
                        errors.append("chart.polar.angle_axis.zero_bearing_deg 必须是有限数值")
                radius_axis = polar.get("radius_axis")
                radius_anchors = radius_axis.get("anchors") if isinstance(radius_axis, dict) else None
                if not isinstance(radius_anchors, list) or len(radius_anchors) < 2:
                    errors.append("chart.polar.radius_axis 至少需要两个径向锚点")
                else:
                    try:
                        AxisMap(radius_axis)
                    except (FigureError, KeyError, TypeError, ValueError) as exc:
                        errors.append(f"chart.polar.radius_axis 无效：{exc}")
                start = polar.get("angle_start_deg", 0)
                end = polar.get("angle_end_deg", 360)
                step = polar.get("angle_step_deg", 1)
                if (
                    not all(isinstance(value, (int, float)) and math.isfinite(float(value)) for value in (start, end, step))
                    or float(end) <= float(start)
                    or float(end) - float(start) > 360
                    or float(step) <= 0
                    or float(step) > float(end) - float(start)
                ):
                    errors.append("chart.polar 的角度范围/步长无效，范围必须递增且不超过 360 度")

        quality = spec.get("quality_gates", {})
        if quality and not isinstance(quality, dict):
            errors.append("quality_gates 必须是对象")
        elif isinstance(quality, dict):
            calibration = quality.get("calibration", {})
            maximum_rmse = calibration.get("max_normalized_rmse") if isinstance(calibration, dict) else None
            if maximum_rmse is not None and (
                not isinstance(maximum_rmse, (int, float)) or float(maximum_rmse) < 0
            ):
                errors.append("quality_gates.calibration.max_normalized_rmse 必须为非负数或 null")
            line_gates = quality.get("polar_line" if chart_type == "polar_line" else "line", {})
            if isinstance(line_gates, dict):
                for key in (
                    "min_coverage",
                    "max_gap_fraction",
                    "min_coverage_solid",
                    "min_coverage_dashed",
                    "max_gap_fraction_solid",
                    "max_gap_fraction_dashed",
                ):
                    value = line_gates.get(key)
                    if value is not None and (
                        not isinstance(value, (int, float)) or not 0 <= float(value) <= 1
                    ):
                        errors.append(f"quality_gates.line.{key} 必须位于 0 到 1 之间或为 null")
            for section in ("scatter", "bar"):
                gates = quality.get(section, {})
                if not isinstance(gates, dict):
                    continue
                ratio = gates.get("max_rejected_ratio")
                if ratio is not None and (
                    not isinstance(ratio, (int, float)) or not 0 <= float(ratio) <= 1
                ):
                    errors.append(
                        f"quality_gates.{section}.max_rejected_ratio 必须位于 0 到 1 之间或为 null"
                    )
                minimum = gates.get("min_accepted_components")
                if minimum is not None and (
                    not isinstance(minimum, int) or isinstance(minimum, bool) or minimum < 1
                ):
                    errors.append(
                        f"quality_gates.{section}.min_accepted_components 必须是正整数"
                    )
        assessment = spec.get("assessment", {})
        if assessment and not isinstance(assessment, dict):
            errors.append("assessment 必须是对象")
        elif isinstance(assessment, dict):
            profile = assessment.get("acceptance_profile", "engineering")
            if profile not in ACCEPTANCE_PROFILES:
                errors.append(
                    "assessment.acceptance_profile 只支持 exploratory、engineering 或 publication"
                )
    render = spec.get("render", {})
    if render and not isinstance(render, dict):
        errors.append("render 必须是对象")
    elif isinstance(render, dict):
        display = render.get("display_geometry", {})
        if display and not isinstance(display, dict):
            errors.append("render.display_geometry 必须是对象")
        elif isinstance(display, dict):
            mode = display.get("mode", "none")
            if mode not in {"none", "shape_preserving"}:
                errors.append("render.display_geometry.mode 只支持 none 或 shape_preserving")
            for key in (
                "smoothing_window",
                "samples_per_interval",
                "outlier_window",
                "knot_stride",
            ):
                value = display.get(key)
                if value is not None and (not isinstance(value, int) or value < 1):
                    errors.append(f"render.display_geometry.{key} 必须为正整数")
            if "max_bridge_gap_px" in display:
                errors.append(
                    "render.display_geometry.max_bridge_gap_px 已弃用；"
                    "请在 chart.series 中声明 curve_topology，由 review-apply 构建连续 data.csv"
                )
            outlier_residual = display.get("max_outlier_pixel_residual")
            if outlier_residual is not None and (
                not isinstance(outlier_residual, (int, float))
                or not math.isfinite(float(outlier_residual))
                or float(outlier_residual) < 0
            ):
                errors.append(
                    "render.display_geometry.max_outlier_pixel_residual 必须为非负数"
                )
        styles = render.get("series_styles", {})
        if styles and not isinstance(styles, dict):
            errors.append("render.series_styles 必须是对象")
        elif isinstance(styles, dict):
            for series_id, style in styles.items():
                if not isinstance(style, dict):
                    errors.append(f"render.series_styles.{series_id} 必须是对象")
                    continue
                if "max_bridge_gap_px" in style:
                    errors.append(
                        f"render.series_styles.{series_id}.max_bridge_gap_px 已弃用；"
                        "正式曲线连续性必须在 data.csv 生成前确定"
                    )
                geometry_source = style.get("geometry_source")
                if geometry_source == "guide_constrained":
                    errors.append(
                        f"render.series_styles.{series_id}.geometry_source=guide_constrained "
                        "已迁移到 chart.series.curve_data_mode=guide_constrained"
                    )
                elif geometry_source is not None and geometry_source != "observations":
                    errors.append(
                        f"render.series_styles.{series_id}.geometry_source 只支持 observations"
                    )
        canvas = render.get("canvas_px")
        if canvas is not None and (
            not isinstance(canvas, list)
            or len(canvas) != 2
            or not all(isinstance(value, int) and value > 0 for value in canvas)
        ):
            errors.append("render.canvas_px 必须是两个正整数")
        axes_box = render.get("axes_box_px")
        if axes_box is not None:
            valid_box = (
                isinstance(axes_box, list)
                and len(axes_box) == 4
                and all(
                    isinstance(value, (int, float)) and math.isfinite(float(value))
                    for value in axes_box
                )
                and float(axes_box[0]) < float(axes_box[2])
                and float(axes_box[1]) < float(axes_box[3])
            )
            if not valid_box:
                errors.append("render.axes_box_px 必须是有效 [left,top,right,bottom]")
            elif isinstance(canvas, list) and len(canvas) == 2 and (
                float(axes_box[0]) < 0
                or float(axes_box[1]) < 0
                or float(axes_box[2]) > float(canvas[0])
                or float(axes_box[3]) > float(canvas[1])
            ):
                errors.append("render.axes_box_px 必须位于 canvas_px 内")
    return errors


class AxisMap:
    def __init__(self, axis: dict[str, Any]):
        self.scale = axis.get("scale", "linear")
        if self.scale not in {"linear", "log10"}:
            raise FigureError(f"不支持的坐标轴尺度：{self.scale}")
        anchors = axis["anchors"]
        self.pixels = np.asarray([float(item["pixel"]) for item in anchors], dtype=float)
        self.values = np.asarray([float(item["value"]) for item in anchors], dtype=float)
        transformed = np.log10(self.values) if self.scale == "log10" else self.values
        self.slope, self.intercept = np.polyfit(self.pixels, transformed, 1)
        if not math.isfinite(float(self.slope)) or abs(float(self.slope)) < 1e-15:
            raise FigureError("坐标轴标定斜率无效，请检查锚点")
        predicted = self.slope * self.pixels + self.intercept
        self.residuals = predicted - transformed
        self.rmse = float(np.sqrt(np.mean((predicted - transformed) ** 2)))
        transformed_span = float(np.ptp(transformed))
        self.normalized_rmse = self.rmse / transformed_span if transformed_span > 0 else math.inf
        self.rmse_evaluable = len(anchors) >= 3

    def value(self, pixel: float) -> float:
        transformed = self.slope * float(pixel) + self.intercept
        return float(10**transformed if self.scale == "log10" else transformed)

    def pixel(self, value: float) -> float:
        if self.scale == "log10" and value <= 0:
            raise FigureError("log10 坐标轴无法映射非正值")
        transformed = math.log10(value) if self.scale == "log10" else value
        return float((transformed - self.intercept) / self.slope)

    def uncertainty(self, pixel: float, pixel_half_width: float = 0.5) -> float:
        """给出像素量化与可评估标定残差共同形成的保守局部半宽。"""
        center = self.value(pixel)
        candidates = [
            abs(self.value(pixel - pixel_half_width) - center),
            abs(self.value(pixel + pixel_half_width) - center),
        ]
        if self.rmse_evaluable:
            transformed = self.slope * float(pixel) + self.intercept
            if self.scale == "log10":
                candidates.extend(
                    [
                        abs(10 ** (transformed - self.rmse) - center),
                        abs(10 ** (transformed + self.rmse) - center),
                    ]
                )
            else:
                candidates.append(self.rmse)
        return float(max(candidates, default=0.0))

    def report(self) -> dict[str, Any]:
        return {
            "scale": self.scale,
            "slope": float(self.slope),
            "intercept": float(self.intercept),
            "rmse_transformed_units": self.rmse,
            "normalized_rmse": self.normalized_rmse,
            "rmse_evaluable": self.rmse_evaluable,
            "anchor_count": int(len(self.pixels)),
            "anchor_residuals_transformed_units": [float(value) for value in self.residuals],
            "说明": (
                "至少三个锚点，可用残差辅助评估标定质量。"
                if self.rmse_evaluable
                else "仅有两个锚点，拟合残差必然接近零，不能独立证明标定准确。"
            ),
        }


def declared_axis_spans(project: dict[str, Any]) -> dict[str, float]:
    """Return confirmed full-axis value spans for uncertainty normalization."""
    chart = project.get("chart", {}) if isinstance(project, dict) else {}
    if not isinstance(chart, dict):
        return {}

    def anchor_span(axis: Any) -> float | None:
        if not isinstance(axis, dict):
            return None
        values: list[float] = []
        for anchor in axis.get("anchors", []):
            try:
                value = float(anchor["value"])
            except (KeyError, TypeError, ValueError):
                continue
            if math.isfinite(value):
                values.append(value)
        if len(values) < 2:
            return None
        span = max(values) - min(values)
        return span if span > 0 else None

    spans: dict[str, float] = {}
    if str(chart.get("type", "line")) == "polar_line":
        polar = chart.get("polar", {})
        if isinstance(polar, dict):
            try:
                angle_span = abs(
                    float(polar.get("angle_end_deg", 360.0))
                    - float(polar.get("angle_start_deg", 0.0))
                )
            except (TypeError, ValueError):
                angle_span = 0.0
            radius_span = anchor_span(polar.get("radius_axis"))
            if math.isfinite(angle_span) and angle_span > 0:
                spans["x"] = angle_span
            if radius_span is not None:
                spans["y"] = radius_span
                spans["value"] = radius_span
        return spans

    x_span = anchor_span(chart.get("x_axis"))
    y_span = anchor_span(chart.get("y_axis"))
    if x_span is not None:
        spans["x"] = x_span
    if y_span is not None:
        spans["y"] = y_span
        spans["value"] = y_span
    return spans


def anchor_jackknife_stability(
    axis: dict[str, Any], pixels: list[float]
) -> tuple[dict[str, Any], list[float]]:
    """Measure calibration sensitivity by refitting after leaving out each anchor."""
    anchors = axis.get("anchors", [])
    if not isinstance(anchors, list) or len(anchors) < 3 or not pixels:
        return (
            {
                "status": "not_evaluable",
                "method": "leave_one_anchor_out",
                "anchor_count": len(anchors) if isinstance(anchors, list) else 0,
                "reason": "至少需要三个锚点和一个候选像素。",
                "normalized_shift_p95": None,
                "normalized_shift_max": None,
                "score": None,
            },
            [],
        )
    baseline = AxisMap(axis)
    baseline_values = np.asarray([baseline.value(pixel) for pixel in pixels], dtype=float)
    span = float(np.ptp(baseline_values))
    if not math.isfinite(span) or span <= 0:
        return (
            {
                "status": "not_evaluable",
                "method": "leave_one_anchor_out",
                "anchor_count": len(anchors),
                "reason": "候选值跨度退化，无法归一化锚点扰动。",
                "normalized_shift_p95": None,
                "normalized_shift_max": None,
                "score": None,
            },
            [],
        )
    shifts: list[np.ndarray] = []
    for omitted_index in range(len(anchors)):
        reduced_axis = dict(axis)
        reduced_axis["anchors"] = [
            item for index, item in enumerate(anchors) if index != omitted_index
        ]
        reduced = AxisMap(reduced_axis)
        perturbed = np.asarray([reduced.value(pixel) for pixel in pixels], dtype=float)
        shifts.append(np.abs(perturbed - baseline_values) / span)
    per_candidate = np.max(np.vstack(shifts), axis=0)
    normalized_shift_p95 = float(np.percentile(per_candidate, 95))
    normalized_shift_max = float(np.max(per_candidate))
    score = round(max(0.0, min(100.0, 100.0 - normalized_shift_p95 * 1000.0)), 1)
    return (
        {
            "status": "measured",
            "method": "leave_one_anchor_out",
            "anchor_count": len(anchors),
            "trials": len(anchors),
            "candidate_count": len(pixels),
            "normalized_shift_p95": normalized_shift_p95,
            "normalized_shift_max": normalized_shift_max,
            "score": score,
            "说明": "逐次移除一个标定锚点并重新拟合；分数反映标定对锚点选择的敏感性。",
        },
        [float(value) for value in per_candidate],
    )


def circular_difference_degrees(left: float, right: float) -> float:
    return float((float(left) - float(right) + 180.0) % 360.0 - 180.0)


def polar_angle_from_pixel(
    pixel_x: float, pixel_y: float, polar: dict[str, Any]
) -> float:
    center_x, center_y = [float(value) for value in polar["center_px"]]
    bearing = math.degrees(math.atan2(float(pixel_y) - center_y, float(pixel_x) - center_x)) % 360.0
    angle_axis = polar["angle_axis"]
    zero = float(angle_axis.get("zero_bearing_deg", 0.0))
    direction = 1.0 if angle_axis.get("direction", "clockwise") == "clockwise" else -1.0
    return float((direction * (bearing - zero)) % 360.0)


def polar_angle_calibration_report(polar: dict[str, Any]) -> dict[str, Any]:
    anchors = polar["angle_axis"]["anchors"]
    residuals = [
        circular_difference_degrees(
            polar_angle_from_pixel(float(anchor["pixel"][0]), float(anchor["pixel"][1]), polar),
            float(anchor["value"]),
        )
        for anchor in anchors
    ]
    rmse = float(np.sqrt(np.mean(np.square(residuals))))
    return {
        "scale": "circular_degree",
        "zero_bearing_deg": float(polar["angle_axis"].get("zero_bearing_deg", 0.0)),
        "direction": str(polar["angle_axis"].get("direction", "clockwise")),
        "anchor_count": len(anchors),
        "anchor_residuals_degrees": residuals,
        "rmse_degrees": rmse,
        "normalized_rmse": rmse / 360.0,
        "rmse_evaluable": len(anchors) >= 3,
        "说明": "角度锚点以声明中心、零度方位和方向进行圆周残差核验。",
    }


def color_mask(array: np.ndarray, color: tuple[int, int, int], tolerance: float) -> np.ndarray:
    delta = array.astype(np.int32) - np.asarray(color, dtype=np.int32)
    return np.sqrt(np.sum(delta * delta, axis=2)) <= tolerance


def rgb_to_lab(array: np.ndarray) -> np.ndarray:
    """把 sRGB 数组转换为 CIE Lab，供抗压缩颜色距离使用。"""
    rgb = array.astype(np.float64) / 255.0
    linear = np.where(rgb <= 0.04045, rgb / 12.92, ((rgb + 0.055) / 1.055) ** 2.4)
    matrix = np.asarray(
        [
            [0.4124564, 0.3575761, 0.1804375],
            [0.2126729, 0.7151522, 0.0721750],
            [0.0193339, 0.1191920, 0.9503041],
        ]
    )
    xyz = linear @ matrix.T
    xyz /= np.asarray([0.95047, 1.0, 1.08883])
    delta = 6 / 29
    transformed = np.where(
        xyz > delta**3,
        np.cbrt(xyz),
        xyz / (3 * delta**2) + 4 / 29,
    )
    return np.stack(
        [
            116 * transformed[..., 1] - 16,
            500 * (transformed[..., 0] - transformed[..., 1]),
            200 * (transformed[..., 1] - transformed[..., 2]),
        ],
        axis=-1,
    )


def series_mask(array: np.ndarray, entry: dict[str, Any]) -> np.ndarray:
    color = parse_hex_color(entry["color"])
    tolerance = float(entry.get("tolerance", 45))
    if str(entry.get("color_space", "rgb")).lower() == "lab":
        target = rgb_to_lab(np.asarray(color, dtype=np.uint8).reshape(1, 1, 3))[0, 0]
        delta = rgb_to_lab(array) - target
        return np.sqrt(np.sum(delta * delta, axis=2)) <= tolerance
    return color_mask(array, color, tolerance)


def parse_guide_points(entry: dict[str, Any]) -> list[tuple[float, float]]:
    points = entry.get("guide_points_px")
    if not isinstance(points, list):
        return []
    parsed: list[tuple[float, float]] = []
    for point in points:
        if isinstance(point, dict):
            x, y = point.get("x"), point.get("y")
        elif isinstance(point, list) and len(point) == 2:
            x, y = point
        else:
            continue
        if (
            isinstance(x, (int, float))
            and isinstance(y, (int, float))
            and math.isfinite(float(x))
            and math.isfinite(float(y))
        ):
            parsed.append((float(x), float(y)))
    return sorted(parsed)


def guide_y_at(entry: dict[str, Any], pixel_x: float) -> float | None:
    points = parse_guide_points(entry)
    if len(points) < 2 or pixel_x < points[0][0] or pixel_x > points[-1][0]:
        return None
    x_values = np.asarray([point[0] for point in points], dtype=float)
    y_values = np.asarray([point[1] for point in points], dtype=float)
    if str(entry.get("guide_interpolation", "linear")) != "shape_preserving" or len(points) == 2:
        return float(np.interp(pixel_x, x_values, y_values))
    widths = np.diff(x_values)
    deltas = np.diff(y_values) / widths
    slopes = np.zeros_like(y_values)
    for index in range(1, len(y_values) - 1):
        if deltas[index - 1] * deltas[index] <= 0:
            slopes[index] = 0.0
        else:
            left_weight = 2.0 * widths[index] + widths[index - 1]
            right_weight = widths[index] + 2.0 * widths[index - 1]
            slopes[index] = (left_weight + right_weight) / (
                left_weight / deltas[index - 1] + right_weight / deltas[index]
            )
    slopes[0], slopes[-1] = deltas[0], deltas[-1]
    interval = min(len(widths) - 1, max(0, int(np.searchsorted(x_values, pixel_x) - 1)))
    t = (pixel_x - x_values[interval]) / widths[interval]
    t2, t3 = t * t, t * t * t
    return float(
        (2 * t3 - 3 * t2 + 1) * y_values[interval]
        + (t3 - 2 * t2 + t) * widths[interval] * slopes[interval]
        + (-2 * t3 + 3 * t2) * y_values[interval + 1]
        + (t3 - t2) * widths[interval] * slopes[interval + 1]
    )


def weighted_median(values: np.ndarray, weights: np.ndarray) -> float:
    order = np.argsort(values)
    sorted_values = values[order]
    cumulative = np.cumsum(weights[order])
    index = int(np.searchsorted(cumulative, cumulative[-1] * 0.5))
    return float(sorted_values[min(index, len(sorted_values) - 1)])


def apply_series_exclusions(
    mask: np.ndarray,
    entry: dict[str, Any],
    plot_box: tuple[int, int, int, int],
) -> tuple[np.ndarray, list[list[int]]]:
    result = mask.copy()
    left, top, right, bottom = plot_box
    applied: list[list[int]] = []
    for value in entry.get("exclude_boxes_px", []):
        if not isinstance(value, list) or len(value) != 4:
            continue
        x0, y0, x1, y1 = [int(round(float(item))) for item in value]
        x0, x1 = max(left, x0), min(right, x1)
        y0, y1 = max(top, y0), min(bottom, y1)
        if x0 > x1 or y0 > y1:
            continue
        result[y0 - top : y1 - top + 1, x0 - left : x1 - left + 1] = False
        applied.append([x0, y0, x1, y1])
    return result, applied


def guide_x_bounds(entry: dict[str, Any], plot_box: tuple[int, int, int, int]) -> tuple[int, int]:
    left, _, right, _ = plot_box
    points = parse_guide_points(entry)
    x_range = entry.get("x_pixel_range")
    if isinstance(x_range, list) and len(x_range) == 2:
        return (
            max(left, int(math.ceil(float(x_range[0])))),
            min(right, int(math.floor(float(x_range[1])))),
        )
    return (
        max(left, int(math.ceil(points[0][0]))),
        min(right, int(math.floor(points[-1][0]))),
    )


def cluster_pixel_rows(pixel_rows: np.ndarray) -> list[dict[str, float | int]]:
    if pixel_rows.size == 0:
        return []
    values = np.sort(pixel_rows.astype(float))
    split_points = np.flatnonzero(np.diff(values) > 1.01) + 1
    clusters: list[dict[str, float | int]] = []
    for cluster in np.split(values, split_points):
        clusters.append(
            {
                "center": float(np.median(cluster)),
                "minimum": float(np.min(cluster)),
                "maximum": float(np.max(cluster)),
                "support": int(cluster.size),
            }
        )
    return clusters


def extract_guided_group(
    crop: np.ndarray,
    box: tuple[int, int, int, int],
    entries: list[dict[str, Any]],
    x_map: AxisMap,
    y_map: AxisMap,
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    """对同色系列执行排他式联合分配；缺失分支不借用同一像素。"""
    left, top, right, _ = box
    entry_masks: dict[str, np.ndarray] = {}
    exclusions: dict[str, list[list[int]]] = {}
    bounds: dict[str, tuple[int, int]] = {}
    stats: dict[str, dict[str, Any]] = {}
    for entry in entries:
        series_id = str(entry["id"])
        entry_masks[series_id], exclusions[series_id] = apply_series_exclusions(
            series_mask(crop, entry), entry, box
        )
        bounds[series_id] = guide_x_bounds(entry, box)
        stats[series_id] = {
            "supported": 0,
            "previous_x": None,
            "previous_y": None,
            "current_gap": 0,
            "maximum_gap": 0,
            "residuals": [],
            "competitive": 0,
            "fused_unassigned": 0,
        }
    start_x = min(value[0] for value in bounds.values())
    end_x = max(value[1] for value in bounds.values())
    rows: list[dict[str, Any]] = []
    ordered_entries = sorted(
        entries,
        key=lambda item: 0 if str(item.get("line_semantics", "solid")) == "solid" else 1,
    )
    for pixel_x in range(start_x, end_x + 1):
        active = [
            entry
            for entry in ordered_entries
            if bounds[str(entry["id"])][0] <= pixel_x <= bounds[str(entry["id"])][1]
            and guide_y_at(entry, pixel_x) is not None
        ]
        if not active:
            continue
        search_half_width = max(int(entry.get("search_half_width_px", 0)) for entry in active)
        x0 = max(0, pixel_x - left - search_half_width)
        x1 = min(crop.shape[1] - 1, pixel_x - left + search_half_width)
        union_mask = np.zeros((crop.shape[0], x1 - x0 + 1), dtype=bool)
        for entry in active:
            union_mask |= entry_masks[str(entry["id"])][:, x0 : x1 + 1]
        local_rows = np.where(union_mask)[0].astype(float) + top
        clusters = cluster_pixel_rows(local_rows)
        expectations = {str(entry["id"]): float(guide_y_at(entry, pixel_x)) for entry in active}
        eligible: dict[str, list[int]] = {}
        for entry in active:
            series_id = str(entry["id"])
            corridor = float(entry.get("guide_corridor_px", 7.0))
            eligible[series_id] = [
                index
                for index, cluster in enumerate(clusters)
                if abs(float(cluster["center"]) - expectations[series_id]) <= corridor
            ]

        best_cost = math.inf
        best_assignment: dict[str, int | None] = {}

        def search_assignment(
            index: int,
            used: set[int],
            cost: float,
            assignment: dict[str, int | None],
        ) -> None:
            nonlocal best_cost, best_assignment
            if cost >= best_cost:
                return
            if index == len(active):
                best_cost = cost
                best_assignment = dict(assignment)
                return
            entry = active[index]
            series_id = str(entry["id"])
            corridor = float(entry.get("guide_corridor_px", 7.0))
            semantics = str(entry.get("line_semantics", "solid"))
            missing_factor = float(
                entry.get("missing_penalty", 2.5 if semantics == "solid" else 0.8)
            )
            assignment[series_id] = None
            search_assignment(index + 1, used, cost + missing_factor * corridor, assignment)
            previous_y = stats[series_id]["previous_y"]
            previous_x = stats[series_id]["previous_x"]
            continuity_weight = float(entry.get("continuity_weight", 0.35))
            track_gap = int(entry.get("max_track_gap_px", 40 if semantics == "dashed" else 4))
            for cluster_index in eligible[series_id]:
                if cluster_index in used:
                    continue
                center = float(clusters[cluster_index]["center"])
                local_cost = abs(center - expectations[series_id])
                if (
                    previous_y is not None
                    and previous_x is not None
                    and pixel_x - int(previous_x) <= track_gap
                ):
                    local_cost += continuity_weight * abs(center - float(previous_y))
                assignment[series_id] = cluster_index
                search_assignment(
                    index + 1,
                    used | {cluster_index},
                    cost + local_cost,
                    assignment,
                )
            assignment.pop(series_id, None)

        search_assignment(0, set(), 0.0, {})
        cluster_competitors = {
            index: [str(entry["id"]) for entry in active if index in eligible[str(entry["id"])]]
            for index in range(len(clusters))
        }
        for entry in active:
            series_id = str(entry["id"])
            state = stats[series_id]
            cluster_index = best_assignment.get(series_id)
            if cluster_index is None:
                state["previous_x"] = None
                state["previous_y"] = None
                state["current_gap"] += 1
                state["maximum_gap"] = max(state["maximum_gap"], state["current_gap"])
                if any(len(cluster_competitors[index]) > 1 for index in eligible[series_id]):
                    state["fused_unassigned"] += 1
                continue
            cluster = clusters[cluster_index]
            pixel_y = float(cluster["center"])
            residual = abs(pixel_y - expectations[series_id])
            competitors = cluster_competitors[cluster_index]
            evidence_status = (
                "model_assisted_exclusive_assignment"
                if len(competitors) > 1
                else "visible_pixel_support"
            )
            if len(competitors) > 1:
                state["competitive"] += 1
            rows.append(
                {
                    "series": series_id,
                    "x": x_map.value(pixel_x),
                    "y": y_map.value(pixel_y),
                    "x_uncertainty": x_map.uncertainty(
                        pixel_x,
                        float(entry.get("center_localization_half_width_px", 0.5)),
                    ),
                    "y_uncertainty": y_map.uncertainty(
                        pixel_y,
                        float(entry.get("center_localization_half_width_px", 0.5)),
                    ),
                    "pixel_x": pixel_x,
                    "pixel_y": pixel_y,
                    "pixel_half_height": max(
                        0.5, (float(cluster["maximum"]) - float(cluster["minimum"])) / 2.0
                    ),
                    "center_localization_half_width_px": float(
                        entry.get("center_localization_half_width_px", 0.5)
                    ),
                    "support_pixels": int(cluster["support"]),
                    "support_x_min": max(left, pixel_x - search_half_width),
                    "support_x_max": min(right, pixel_x + search_half_width),
                    "guide_residual_px": residual,
                    "group_assignment_cost": best_cost,
                    "competing_series": "|".join(competitors) if len(competitors) > 1 else "",
                    "line_semantics": str(entry.get("line_semantics", "solid")),
                    "evidence_status": evidence_status,
                    "segment_break": state["previous_x"] is None,
                    "status": "visible_candidate",
                }
            )
            state["supported"] += 1
            state["previous_x"] = pixel_x
            state["previous_y"] = pixel_y
            state["current_gap"] = 0
            state["residuals"].append(residual)
    diagnostics: list[dict[str, Any]] = []
    for entry in entries:
        series_id = str(entry["id"])
        state = stats[series_id]
        width = max(1, bounds[series_id][1] - bounds[series_id][0] + 1)
        diagnostics.append(
            {
                "id": series_id,
                "extraction_mode": "guided_group_path",
                "line_semantics": str(entry.get("line_semantics", "solid")),
                "supported_columns": state["supported"],
                "coverage": state["supported"] / width,
                "maximum_gap_columns": state["maximum_gap"],
                "maximum_gap_fraction": state["maximum_gap"] / width,
                "model_assisted_columns": state["competitive"],
                "fused_unassigned_columns": state["fused_unassigned"],
                "ambiguous_columns": 0,
                "mean_guide_residual_px": (
                    float(np.mean(state["residuals"])) if state["residuals"] else None
                ),
                "applied_exclusions_px": exclusions[series_id],
            }
        )
    return rows, diagnostics


def components(mask: np.ndarray) -> list[list[tuple[int, int]]]:
    height, width = mask.shape
    visited = np.zeros_like(mask, dtype=bool)
    found: list[list[tuple[int, int]]] = []
    for y in range(height):
        for x in range(width):
            if not mask[y, x] or visited[y, x]:
                continue
            queue = deque([(x, y)])
            visited[y, x] = True
            pixels: list[tuple[int, int]] = []
            while queue:
                px, py = queue.popleft()
                pixels.append((px, py))
                for ny in range(max(0, py - 1), min(height, py + 2)):
                    for nx in range(max(0, px - 1), min(width, px + 2)):
                        if mask[ny, nx] and not visited[ny, nx]:
                            visited[ny, nx] = True
                            queue.append((nx, ny))
            found.append(pixels)
    return found


def prepare_measurement_raster(spec: dict[str, Any], spec_path: Path, out_dir: Path) -> tuple[Image.Image, Path]:
    source = spec["source"]
    source_path = resolve_from(spec_path, source["path"])
    measurement_value = source.get("measurement_raster")
    if measurement_value:
        measurement_path = resolve_from(spec_path, str(measurement_value))
    elif source_path.suffix.lower() == ".pdf":
        measurement_path = out_dir / "source-page.png"
        render_pdf_page(source_path, int(source.get("page") or 1), measurement_path)
    else:
        measurement_path = source_path
    if not measurement_path.is_file():
        raise FigureError(f"找不到测量栅格：{measurement_path}")
    expected = source.get("measurement_sha256")
    if expected and sha256_file(measurement_path) != expected:
        raise FigureError("测量栅格哈希与项目声明不一致")
    return Image.open(measurement_path).convert("RGB"), measurement_path


def extract_marker_centers(
    crop: np.ndarray,
    box: tuple[int, int, int, int],
    entry: dict[str, Any],
    x_map: AxisMap,
    y_map: AxisMap,
) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    """Recover visible marker centres from a connected marker-and-line series."""
    left, top, right, _ = box
    mask, applied_exclusions = apply_series_exclusions(series_mask(crop, entry), entry, box)
    start_x, end_x = guide_x_bounds(entry, box)
    corridor = float(entry.get("guide_corridor_px", 9.0))
    minimum_span = float(entry.get("marker_min_span_px", 6.0))
    minimum_width = int(math.ceil(float(entry.get("marker_min_width_px", 2.0))))
    radius = int(math.ceil(float(entry.get("marker_window_radius_px", 7.0))))
    localization_half_width = float(
        entry.get("center_localization_half_width_px", 0.5)
    )
    qualifying: list[int] = []
    column_spans: dict[int, float] = {}
    for pixel_x in range(start_x, end_x + 1):
        expected_y = guide_y_at(entry, pixel_x)
        if expected_y is None:
            continue
        local_rows = np.flatnonzero(mask[:, pixel_x - left]).astype(float) + top
        accepted = local_rows[np.abs(local_rows - expected_y) <= corridor]
        span = float(np.ptp(accepted)) if accepted.size >= 2 else 0.0
        column_spans[pixel_x] = span
        if span >= minimum_span:
            qualifying.append(pixel_x)

    runs: list[list[int]] = []
    for pixel_x in qualifying:
        if not runs or pixel_x > runs[-1][-1] + 1:
            runs.append([pixel_x])
        else:
            runs[-1].append(pixel_x)

    rows: list[dict[str, Any]] = []
    rejected_runs = 0
    previous_x: float | None = None
    for run in runs:
        if len(run) < minimum_width:
            rejected_runs += 1
            continue
        peak_x = max(run, key=lambda value: column_spans.get(value, 0.0))
        expected_y = guide_y_at(entry, peak_x)
        if expected_y is None:
            rejected_runs += 1
            continue
        x0 = max(left, run[0] - 1)
        x1 = min(right, run[-1] + 1)
        y0 = max(top, int(math.floor(expected_y - radius)))
        y1 = min(top + crop.shape[0] - 1, int(math.ceil(expected_y + radius)))
        patch = mask[y0 - top : y1 - top + 1, x0 - left : x1 - left + 1]
        local_y, local_x = np.where(patch)
        if local_x.size == 0:
            rejected_runs += 1
            continue
        pixel_x = float(np.median(local_x + x0))
        pixel_y = float(np.median(local_y + y0))
        half_width = max(0.5, float(np.ptp(local_x)) / 2.0)
        half_height = max(0.5, float(np.ptp(local_y)) / 2.0)
        rows.append(
            {
                "series": entry["id"],
                "x": x_map.value(pixel_x),
                "y": y_map.value(pixel_y),
                "x_uncertainty": x_map.uncertainty(
                    pixel_x, localization_half_width
                ),
                "y_uncertainty": y_map.uncertainty(
                    pixel_y, localization_half_width
                ),
                "pixel_x": pixel_x,
                "pixel_y": pixel_y,
                "pixel_half_width": half_width,
                "pixel_half_height": half_height,
                "center_localization_half_width_px": localization_half_width,
                "support_pixels": int(local_x.size),
                "marker_shape": str(entry.get("marker_shape", "unspecified")),
                "evidence_status": "visible_marker_support",
                "segment_break": previous_x is None,
                "status": "visible_candidate",
            }
        )
        previous_x = pixel_x
    minimum_distance = float(entry.get("marker_min_distance_px", 8.0))
    merged_rows: list[dict[str, Any]] = []
    merged_detections = 0
    for row in sorted(rows, key=lambda item: float(item["pixel_x"])):
        if (
            merged_rows
            and float(row["pixel_x"]) - float(merged_rows[-1]["pixel_x"])
            < minimum_distance
        ):
            merged_detections += 1
            if int(row["support_pixels"]) > int(merged_rows[-1]["support_pixels"]):
                merged_rows[-1] = row
            continue
        merged_rows.append(row)
    rows = merged_rows
    for index, row in enumerate(rows):
        row["segment_break"] = index == 0

    minimum_markers = int(entry.get("min_marker_count", 1))
    diagnostic = {
        "id": entry["id"],
        "extraction_mode": "marker_centers",
        "marker_shape": str(entry.get("marker_shape", "unspecified")),
        "accepted_markers": len(rows),
        "rejected_runs": rejected_runs,
        "merged_detections": merged_detections,
        "marker_min_distance_px": minimum_distance,
        "minimum_declared_marker_count": minimum_markers,
        "coverage": min(1.0, len(rows) / max(1, minimum_markers)),
        "coverage_basis": "minimum_declared_marker_count_not_completeness",
        "maximum_gap_columns": 0,
        "maximum_gap_fraction": 0.0,
        "ambiguous_columns": 0,
        "applied_exclusions_px": applied_exclusions,
    }
    return rows, diagnostic


def extract_line(
    array: np.ndarray,
    box: tuple[int, int, int, int],
    series: list[dict[str, Any]],
    x_map: AxisMap,
    y_map: AxisMap,
) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    left, top, right, bottom = box
    crop = array[top : bottom + 1, left : right + 1]
    rows: list[dict[str, Any]] = []
    diagnostics: dict[str, Any] = {"series": []}
    processed_entries: set[int] = set()
    for entry in series:
        if id(entry) in processed_entries:
            continue
        extraction_mode = str(entry.get("extraction_mode", "color_column_median"))
        if extraction_mode == "guided_group_path":
            group_id = str(entry.get("shared_color_group", entry.get("color")))
            group_entries = [
                peer
                for peer in series
                if id(peer) not in processed_entries
                and str(peer.get("extraction_mode", "color_column_median"))
                == "guided_group_path"
                and str(peer.get("shared_color_group", peer.get("color"))) == group_id
            ]
            group_rows, group_diagnostics = extract_guided_group(
                crop, box, group_entries, x_map, y_map
            )
            rows.extend(group_rows)
            diagnostics["series"].extend(group_diagnostics)
            processed_entries.update(id(peer) for peer in group_entries)
            continue
        mask, applied_exclusions = apply_series_exclusions(series_mask(crop, entry), entry, box)
        if extraction_mode == "marker_centers":
            marker_rows, marker_diagnostics = extract_marker_centers(
                crop, box, entry, x_map, y_map
            )
            rows.extend(marker_rows)
            diagnostics["series"].append(marker_diagnostics)
            continue
        if extraction_mode == "guided_path":
            guide_points = parse_guide_points(entry)
            corridor = float(entry.get("guide_corridor_px", 7.0))
            x_range = entry.get("x_pixel_range")
            if isinstance(x_range, list) and len(x_range) == 2:
                start_x = max(left, int(math.ceil(float(x_range[0]))))
                end_x = min(right, int(math.floor(float(x_range[1]))))
            else:
                start_x = max(left, int(math.ceil(guide_points[0][0])))
                end_x = min(right, int(math.floor(guide_points[-1][0])))
            peer_entries = [
                peer
                for peer in series
                if peer is not entry
                and str(peer.get("shared_color_group", peer.get("color")))
                == str(entry.get("shared_color_group", entry.get("color")))
                and str(peer.get("extraction_mode", "color_column_median")) == "guided_path"
            ]
            supported = 0
            ambiguous = 0
            previous_x: int | None = None
            current_gap = 0
            maximum_gap = 0
            residuals: list[float] = []
            localization_half_width = float(
                entry.get("center_localization_half_width_px", 0.5)
            )
            for pixel_x in range(start_x, end_x + 1):
                expected_y = guide_y_at(entry, pixel_x)
                if expected_y is None:
                    continue
                local_ys = np.flatnonzero(mask[:, pixel_x - left]).astype(float) + top
                distances = np.abs(local_ys - expected_y)
                accepted = distances <= corridor
                if not np.any(accepted):
                    previous_x = None
                    current_gap += 1
                    maximum_gap = max(maximum_gap, current_gap)
                    continue
                current_gap = 0
                candidate_ys = local_ys[accepted]
                candidate_distances = distances[accepted]
                weights = np.exp(-0.5 * np.square(candidate_distances / max(1.0, corridor / 3.0)))
                pixel_y = weighted_median(candidate_ys, weights)
                residual = abs(pixel_y - expected_y)
                residuals.append(residual)
                evidence_status = "visible_pixel_support"
                for peer in peer_entries:
                    peer_y = guide_y_at(peer, pixel_x)
                    if peer_y is None:
                        continue
                    peer_corridor = float(peer.get("guide_corridor_px", corridor))
                    if abs(peer_y - expected_y) <= corridor + peer_corridor and abs(pixel_y - peer_y) <= peer_corridor:
                        evidence_status = "ambiguous_shared_colour"
                        ambiguous += 1
                        break
                pixel_half_height = max(0.5, float(np.ptp(candidate_ys)) / 2.0)
                rows.append(
                    {
                        "series": entry["id"],
                        "x": x_map.value(pixel_x),
                        "y": y_map.value(pixel_y),
                        "x_uncertainty": x_map.uncertainty(
                            pixel_x, localization_half_width
                        ),
                        "y_uncertainty": y_map.uncertainty(
                            pixel_y, localization_half_width
                        ),
                        "pixel_x": pixel_x,
                        "pixel_y": pixel_y,
                        "pixel_half_height": pixel_half_height,
                        "center_localization_half_width_px": localization_half_width,
                        "support_pixels": int(candidate_ys.size),
                        "guide_residual_px": residual,
                        "evidence_status": evidence_status,
                        "segment_break": previous_x is None,
                        "status": "visible_candidate",
                    }
                )
                supported += 1
                previous_x = pixel_x
            width = max(1, end_x - start_x + 1)
            diagnostics["series"].append(
                {
                    "id": entry["id"],
                    "extraction_mode": extraction_mode,
                    "line_semantics": str(entry.get("line_semantics", "solid")),
                    "supported_columns": supported,
                    "coverage": supported / width,
                    "maximum_gap_columns": maximum_gap,
                    "maximum_gap_fraction": maximum_gap / width,
                    "ambiguous_columns": ambiguous,
                    "mean_guide_residual_px": float(np.mean(residuals)) if residuals else None,
                    "applied_exclusions_px": applied_exclusions,
                }
            )
            continue
        supported = 0
        previous_x: int | None = None
        current_gap = 0
        maximum_gap = 0
        localization_half_width = float(
            entry.get("center_localization_half_width_px", 0.5)
        )
        for local_x in range(mask.shape[1]):
            local_ys = np.flatnonzero(mask[:, local_x])
            if local_ys.size == 0:
                previous_x = None
                current_gap += 1
                maximum_gap = max(maximum_gap, current_gap)
                continue
            current_gap = 0
            pixel_x = left + local_x
            pixel_y = top + float(np.median(local_ys))
            pixel_half_height = max(0.5, float(np.ptp(local_ys)) / 2.0)
            rows.append(
                {
                    "series": entry["id"],
                    "x": x_map.value(pixel_x),
                    "y": y_map.value(pixel_y),
                    "x_uncertainty": x_map.uncertainty(
                        pixel_x, localization_half_width
                    ),
                    "y_uncertainty": y_map.uncertainty(
                        pixel_y, localization_half_width
                    ),
                    "pixel_x": pixel_x,
                    "pixel_y": pixel_y,
                    "pixel_half_height": pixel_half_height,
                    "center_localization_half_width_px": localization_half_width,
                    "segment_break": previous_x is None,
                    "status": "visible",
                }
            )
            supported += 1
            previous_x = pixel_x
        coverage = supported / max(1, mask.shape[1])
        diagnostics["series"].append(
                {
                    "id": entry["id"],
                    "extraction_mode": extraction_mode,
                    "line_semantics": str(entry.get("line_semantics", "solid")),
                    "supported_columns": supported,
                    "coverage": coverage,
                    "maximum_gap_columns": maximum_gap,
                    "maximum_gap_fraction": maximum_gap / max(1, mask.shape[1]),
                    "ambiguous_columns": 0,
                    "applied_exclusions_px": applied_exclusions,
                }
            )
    return rows, diagnostics


def extract_polar_line(
    array: np.ndarray,
    box: tuple[int, int, int, int],
    series: list[dict[str, Any]],
    polar: dict[str, Any],
    radius_map: AxisMap,
) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    """Sample a single-valued radial trace by declared angle with real colour-pixel support."""
    left, top, right, bottom = box
    crop = array[top : bottom + 1, left : right + 1]
    center_x, center_y = [float(value) for value in polar["center_px"]]
    inner_radius = float(polar["inner_radius_px"])
    outer_radius = float(polar["outer_radius_px"])
    angle_start = float(polar.get("angle_start_deg", 0.0))
    angle_end = float(polar.get("angle_end_deg", 360.0))
    angle_step = float(polar.get("angle_step_deg", 1.0))
    sample_angles = np.arange(angle_start, angle_end, angle_step, dtype=float)
    pixel_y_grid, pixel_x_grid = np.indices(crop.shape[:2], dtype=float)
    pixel_x_grid += left
    pixel_y_grid += top
    radii = np.hypot(pixel_x_grid - center_x, pixel_y_grid - center_y)
    bearings = np.degrees(np.arctan2(pixel_y_grid - center_y, pixel_x_grid - center_x)) % 360.0
    angle_axis = polar["angle_axis"]
    zero = float(angle_axis.get("zero_bearing_deg", 0.0))
    direction = 1.0 if angle_axis.get("direction", "clockwise") == "clockwise" else -1.0
    angles = (direction * (bearings - zero)) % 360.0
    annulus = (radii >= inner_radius) & (radii <= outer_radius)
    rows: list[dict[str, Any]] = []
    diagnostics: dict[str, Any] = {
        "series": [],
        "angle_axis": polar_angle_calibration_report(polar),
        "radius_axis": radius_map.report(),
    }
    for entry in series:
        mask, applied_exclusions = apply_series_exclusions(series_mask(crop, entry), entry, box)
        mask &= annulus
        half_width = float(entry.get("angular_half_width_deg", max(0.75, angle_step * 0.75)))
        localization_half_width = float(
            entry.get("center_localization_half_width_px", 0.5)
        )
        supported = 0
        maximum_gap = 0
        current_gap = 0
        previous_supported = False
        for sample_angle in sample_angles:
            angle_distance = np.abs((angles - sample_angle + 180.0) % 360.0 - 180.0)
            accepted = mask & (angle_distance <= half_width)
            if not np.any(accepted):
                current_gap += 1
                maximum_gap = max(maximum_gap, current_gap)
                previous_supported = False
                continue
            accepted_radii = radii[accepted]
            accepted_angles = angles[accepted]
            accepted_x = pixel_x_grid[accepted]
            accepted_y = pixel_y_grid[accepted]
            weights = np.exp(-0.5 * np.square(
                np.asarray(
                    [circular_difference_degrees(value, sample_angle) for value in accepted_angles],
                    dtype=float,
                )
                / max(0.25, half_width / 2.0)
            ))
            radius = weighted_median(accepted_radii, weights)
            closest = int(np.argmin(np.abs(accepted_radii - radius)))
            pixel_x = float(accepted_x[closest])
            pixel_y = float(accepted_y[closest])
            radial_half_width = max(
                0.5,
                (float(np.max(accepted_radii)) - float(np.min(accepted_radii))) / 2.0,
            )
            rows.append(
                {
                    "series": str(entry["id"]),
                    "x": float(sample_angle),
                    "y": radius_map.value(radius),
                    "x_uncertainty": angle_step / 2.0,
                    "y_uncertainty": radius_map.uncertainty(
                        radius, localization_half_width
                    ),
                    "pixel_x": pixel_x,
                    "pixel_y": pixel_y,
                    "pixel_radius": radius,
                    "pixel_half_height": radial_half_width,
                    "center_localization_half_width_px": localization_half_width,
                    "support_pixels": int(np.count_nonzero(accepted)),
                    "angular_half_width_deg": half_width,
                    "evidence_status": "visible_pixel_support",
                    "segment_break": not previous_supported,
                    "status": "visible_candidate",
                }
            )
            supported += 1
            current_gap = 0
            previous_supported = True
        sample_count = max(1, len(sample_angles))
        diagnostics["series"].append(
            {
                "id": str(entry["id"]),
                "extraction_mode": "polar_radial",
                "line_semantics": "solid",
                "supported_columns": supported,
                "coverage": supported / sample_count,
                "maximum_gap_columns": maximum_gap,
                "maximum_gap_fraction": maximum_gap / sample_count,
                "angular_half_width_deg": half_width,
                "applied_exclusions_px": applied_exclusions,
            }
        )
    return rows, diagnostics


def extract_scatter(
    array: np.ndarray,
    box: tuple[int, int, int, int],
    series: list[dict[str, Any]],
    x_map: AxisMap,
    y_map: AxisMap,
) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    left, top, right, bottom = box
    crop = array[top : bottom + 1, left : right + 1]
    rows: list[dict[str, Any]] = []
    diagnostics: dict[str, Any] = {"series": []}
    for entry in series:
        mask = series_mask(crop, entry)
        min_area = int(entry.get("min_area", 4))
        max_area = int(entry.get("max_area", 500))
        minimum_fill_ratio = float(entry.get("min_fill_ratio", 0.25))
        maximum_aspect_ratio = float(entry.get("max_aspect_ratio", 4.0))
        accepted = 0
        rejected = 0
        for component in components(mask):
            area = len(component)
            if area < min_area or area > max_area:
                rejected += 1
                continue
            xs = np.asarray([point[0] for point in component], dtype=float)
            ys = np.asarray([point[1] for point in component], dtype=float)
            width = float(np.ptp(xs) + 1)
            height = float(np.ptp(ys) + 1)
            fill_ratio = area / max(1.0, width * height)
            aspect_ratio = max(width, height) / max(1.0, min(width, height))
            if fill_ratio < minimum_fill_ratio or aspect_ratio > maximum_aspect_ratio:
                rejected += 1
                continue
            pixel_x = left + float(xs.mean())
            pixel_y = top + float(ys.mean())
            rows.append(
                {
                    "series": entry["id"],
                    "x": x_map.value(pixel_x),
                    "y": y_map.value(pixel_y),
                    "x_uncertainty": x_map.uncertainty(
                        pixel_x, max(0.5, float(np.ptp(xs)) / 2.0)
                    ),
                    "y_uncertainty": y_map.uncertainty(
                        pixel_y, max(0.5, float(np.ptp(ys)) / 2.0)
                    ),
                    "pixel_x": pixel_x,
                    "pixel_y": pixel_y,
                    "component_area": area,
                    "component_width": width,
                    "component_height": height,
                    "component_fill_ratio": fill_ratio,
                    "component_aspect_ratio": aspect_ratio,
                    "status": "visible_component",
                }
            )
            accepted += 1
        diagnostics["series"].append(
            {"id": entry["id"], "accepted_components": accepted, "rejected_components": rejected}
        )
    rows.sort(key=lambda row: (str(row["series"]), float(row["x"])))
    return rows, diagnostics


def extract_bars(
    array: np.ndarray,
    box: tuple[int, int, int, int],
    series: list[dict[str, Any]],
    x_map: AxisMap,
    y_map: AxisMap,
    chart: dict[str, Any],
) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    left, top, right, bottom = box
    crop = array[top : bottom + 1, left : right + 1]
    if "baseline_pixel" in chart:
        baseline = float(chart["baseline_pixel"])
    elif y_map.scale == "log10":
        raise FigureError("对数纵轴的柱图必须显式提供正值基线对应的 baseline_pixel")
    else:
        baseline = float(y_map.pixel(0.0))
    rows: list[dict[str, Any]] = []
    diagnostics: dict[str, Any] = {"baseline_pixel": baseline, "series": []}
    candidates: list[dict[str, Any]] = []
    for entry in series:
        mask = series_mask(crop, entry)
        min_area = int(entry.get("min_area", 20))
        max_area = int(entry.get("max_area", 1000000))
        minimum_rectangularity = float(entry.get("min_rectangularity", 0.8))
        baseline_tolerance = float(entry.get("baseline_tolerance", 3.0))
        accepted = 0
        rejected = 0
        for component in components(mask):
            if len(component) < min_area or len(component) > max_area:
                rejected += 1
                continue
            xs = [point[0] for point in component]
            ys = [point[1] for point in component]
            x0, x1 = left + min(xs), left + max(xs)
            y0, y1 = top + min(ys), top + max(ys)
            if x1 - x0 < 2 or y1 - y0 < 2:
                rejected += 1
                continue
            bounding_area = (x1 - x0 + 1) * (y1 - y0 + 1)
            rectangularity = len(component) / max(1, bounding_area)
            baseline_distance = min(abs(y0 - baseline), abs(y1 - baseline))
            if rectangularity < minimum_rectangularity or baseline_distance > baseline_tolerance:
                rejected += 1
                continue
            endpoint = y0 if abs(y0 - baseline) >= abs(y1 - baseline) else y1
            candidates.append(
                {
                    "series": entry["id"],
                    "pixel_left": x0,
                    "pixel_top": y0,
                    "pixel_right": x1,
                    "pixel_bottom": y1,
                    "pixel_center": (x0 + x1) / 2.0,
                    "value": y_map.value(endpoint),
                    "value_uncertainty": y_map.uncertainty(endpoint),
                    "baseline_value": y_map.value(baseline),
                    "x_uncertainty": x_map.uncertainty(center := (x0 + x1) / 2.0, (x1 - x0 + 1) / 2.0),
                    "rectangularity": rectangularity,
                    "baseline_distance_pixels": baseline_distance,
                    "status": "visible_rectangle",
                }
            )
            accepted += 1
        diagnostics["series"].append(
            {"id": entry["id"], "accepted_components": accepted, "rejected_components": rejected}
        )
    candidates.sort(key=lambda row: float(row["pixel_center"]))
    centers: list[float] = []
    for candidate in candidates:
        center = float(candidate["pixel_center"])
        category = next((index for index, known in enumerate(centers) if abs(known - center) <= 3), None)
        if category is None:
            centers.append(center)
            category = len(centers) - 1
        candidate["category_index"] = category
        candidate["x_value"] = x_map.value(center)
        rows.append(candidate)
    return rows, diagnostics


def write_csv(path: Path, rows: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    if not rows:
        path.write_text("", encoding="utf-8")
        return
    fields: list[str] = []
    for row in rows:
        for key in row:
            if key not in fields:
                fields.append(key)
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields)
        writer.writeheader()
        writer.writerows(rows)


def draw_overlay(image: Image.Image, chart_type: str, rows: list[dict[str, Any]], output: Path) -> None:
    overlay = image.copy()
    draw = ImageDraw.Draw(overlay)
    if chart_type in {"line", "polar_line"}:
        grouped: dict[str, list[tuple[float, float]]] = {}
        for row in rows:
            grouped.setdefault(str(row["series"]), []).append((float(row["pixel_x"]), float(row["pixel_y"])))
        for points in grouped.values():
            for point in points:
                draw.ellipse((point[0] - 1.5, point[1] - 1.5, point[0] + 1.5, point[1] + 1.5), outline="#ff00ff")
    elif chart_type == "scatter":
        for row in rows:
            x, y = float(row["pixel_x"]), float(row["pixel_y"])
            draw.ellipse((x - 5, y - 5, x + 5, y + 5), outline="#ff00ff", width=2)
    else:
        for row in rows:
            draw.rectangle(
                (
                    float(row["pixel_left"]),
                    float(row["pixel_top"]),
                    float(row["pixel_right"]),
                    float(row["pixel_bottom"]),
                ),
                outline="#ff00ff",
                width=2,
            )
    output.parent.mkdir(parents=True, exist_ok=True)
    overlay.save(output)


def load_manifest(root: Path) -> dict[str, Any]:
    path = root / "manifest.json"
    if path.exists():
        return read_json(path)
    return {
        "schema": MANIFEST_SCHEMA,
        "tool_version": VERSION,
        "spec_review_status": "not_run",
        "spec_confirmation_status": "not_run",
        "extraction_status": "not_run",
        "review_status": "not_run",
        "render_status": "not_run",
        "delivery_status": "not_run",
        "artifacts": {},
    }


def artifact_entry(path: Path) -> dict[str, Any]:
    return {"path": str(path), "sha256": sha256_file(path), "bytes": path.stat().st_size}


def draw_spec_review_overlay(
    image: Image.Image, spec: dict[str, Any], output: Path
) -> None:
    """Render the declared plot geometry before any candidate extraction."""
    overlay = image.copy()
    draw = ImageDraw.Draw(overlay)
    chart = spec["chart"]
    left, top, right, bottom = [int(round(value)) for value in chart["plot_box"]]
    draw.rectangle((left, top, right, bottom), outline="#00bfff", width=3)
    draw.text((left + 6, top + 6), "PLOT BOX", fill="#006d99")

    if chart["type"] == "polar_line":
        polar = chart["polar"]
        center_x, center_y = [float(value) for value in polar["center_px"]]
        draw.line((center_x - 8, center_y, center_x + 8, center_y), fill="#00bfff", width=2)
        draw.line((center_x, center_y - 8, center_x, center_y + 8), fill="#00bfff", width=2)
        for anchor in polar["radius_axis"]["anchors"]:
            radius = float(anchor["pixel"])
            draw.ellipse(
                (center_x - radius, center_y - radius, center_x + radius, center_y + radius),
                outline="#00bfff",
                width=1,
            )
        for anchor in polar["angle_axis"]["anchors"]:
            point_x, point_y = [float(value) for value in anchor["pixel"]]
            draw.line((center_x, center_y, point_x, point_y), fill="#00bfff", width=1)
            draw.ellipse((point_x - 4, point_y - 4, point_x + 4, point_y + 4), fill="#00bfff")
    else:
        for anchor in chart["x_axis"]["anchors"]:
            pixel = float(anchor["pixel"])
            draw.line((pixel, bottom - 7, pixel, bottom + 7), fill="#00bfff", width=2)
        for anchor in chart["y_axis"]["anchors"]:
            pixel = float(anchor["pixel"])
            draw.line((left - 7, pixel, left + 7, pixel), fill="#00bfff", width=2)

    swatch_y = top + 24
    for series in chart["series"]:
        color = str(series["color"])
        draw.rectangle((left + 7, swatch_y, left + 25, swatch_y + 8), fill=color, outline="#ffffff")
        label = str(series["id"])
        if str(series.get("extraction_mode")) == "marker_centers":
            label += f" [MARKERS:{series.get('marker_shape', 'unspecified')}]"
        draw.text((left + 30, swatch_y - 2), label, fill="#202020")
        swatch_y += 14
        guide_points = series.get("guide_points_px", [])
        if len(guide_points) >= 2:
            points = [tuple(float(value) for value in point) for point in guide_points]
            draw.line(points, fill=color, width=2)
            for point_x, point_y in points:
                draw.ellipse(
                    (point_x - 3, point_y - 3, point_x + 3, point_y + 3),
                    fill="#ffffff",
                    outline=color,
                    width=2,
                )
        for exclusion in series.get("exclude_boxes_px", []):
            draw.rectangle(tuple(float(value) for value in exclusion), outline="#ff8c00", width=2)
    notice = "SPEC GUIDE ONLY - NO DATA EXTRACTED"
    draw.text(
        (max(4, image.width - 270), max(4, image.height - 18)),
        notice,
        fill="#d000ff",
    )
    output.parent.mkdir(parents=True, exist_ok=True)
    overlay.save(output)


def spec_review_command(spec_path: Path, out_dir: Path) -> dict[str, Any]:
    """Prepare a hash-bound visual specification for informed user confirmation."""
    spec_path = spec_path.expanduser().resolve()
    out_dir = out_dir.expanduser().resolve()
    out_dir.mkdir(parents=True, exist_ok=True)
    spec = read_json(spec_path)
    errors = validate_spec(spec, spec_path, extraction=True)
    if errors:
        raise FigureError("；".join(errors))
    image, measurement_path = prepare_measurement_raster(spec, spec_path, out_dir)
    overlay_path = out_dir / "spec-review.png"
    draw_spec_review_overlay(image, spec, overlay_path)
    chart = spec["chart"]
    report = {
        "schema": SPEC_REVIEW_SCHEMA,
        "tool_version": VERSION,
        "status": "awaiting_user_confirmation",
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "project_id": spec.get("project_id"),
        "project_path": str(spec_path),
        "project_sha256": sha256_file(spec_path),
        "source_sha256": spec["source"]["sha256"],
        "measurement_raster": str(measurement_path),
        "measurement_sha256": sha256_file(measurement_path),
        "measurement_size": [image.width, image.height],
        "chart_type": chart["type"],
        "overlay_semantics": {
            "blue": "plot box and calibration anchors",
            "series_color_with_hollow_nodes": "sparse guide only, not extracted data",
            "orange": "declared exclusion boxes",
        },
        "plot_box": chart["plot_box"],
        "series": [
            {
                "id": series["id"],
                "color": series["color"],
                "extraction_mode": series.get("extraction_mode", "color_column_median"),
                "marker_shape": series.get("marker_shape"),
                "line_semantics": series.get("line_semantics"),
                "exclude_boxes_px": series.get("exclude_boxes_px", []),
            }
            for series in chart["series"]
        ],
        "confirmation_scope": [
            "原始图对象是否正确",
            "蓝色绘图区和坐标锚点是否正确",
            "系列颜色、线型与语义是否正确",
            "橙色排除框是否只覆盖图中文字、图例或坐标轴",
        ],
        "required_next": "用户查看原图与 spec-review.png 后给出明确确认；再运行 spec-confirm。",
    }
    report_path = out_dir / "spec-review.json"
    write_json(report_path, report)
    manifest = load_manifest(out_dir)
    manifest.update(
        {
            "project_id": spec.get("project_id"),
            "source_sha256": spec["source"]["sha256"],
            "spec_review_status": "awaiting_user_confirmation",
            "spec_confirmation_status": "not_run",
            "tool_version": VERSION,
        }
    )
    manifest["artifacts"].update(
        {
            "project_spec": artifact_entry(spec_path),
            "spec_review": artifact_entry(report_path),
            "spec_review_overlay": artifact_entry(overlay_path),
        }
    )
    write_json(out_dir / "manifest.json", manifest)
    return report


def spec_confirm_command(
    spec_path: Path,
    project_dir: Path,
    confirmed_by: str,
    confirmation: str,
) -> dict[str, Any]:
    """Bind an explicit user statement to the exact reviewed specification."""
    spec_path = spec_path.expanduser().resolve()
    project_dir = project_dir.expanduser().resolve()
    if not confirmed_by.strip() or not confirmation.strip():
        raise FigureError("confirmed-by 和 confirmation 都不能为空")
    review_path = project_dir / "spec-review.json"
    if not review_path.is_file():
        raise FigureError("缺少 spec-review.json，请先运行 spec-review 并让用户查看叠图")
    review = read_json(review_path)
    if review.get("schema") != SPEC_REVIEW_SCHEMA:
        raise FigureError("spec-review.json schema 不受支持")
    project_hash = sha256_file(spec_path)
    spec = read_json(spec_path)
    if review.get("project_sha256") != project_hash:
        raise FigureError("project.json 在规格叠图生成后已变化，请重新运行 spec-review")
    if review.get("source_sha256") != spec["source"]["sha256"]:
        raise FigureError("来源哈希与规格复核记录不一致，请重新运行 spec-review")
    record = {
        "schema": SPEC_CONFIRMATION_SCHEMA,
        "tool_version": VERSION,
        "status": "confirmed",
        "confirmed_at": datetime.now(timezone.utc).isoformat(),
        "confirmed_by": confirmed_by.strip(),
        "confirmation": confirmation.strip(),
        "project_id": spec.get("project_id"),
        "project_sha256": project_hash,
        "source_sha256": review["source_sha256"],
        "measurement_sha256": review["measurement_sha256"],
        "spec_review_sha256": sha256_file(review_path),
        "spec_review_overlay_sha256": sha256_file(project_dir / "spec-review.png"),
    }
    record_path = project_dir / "spec-confirmation.json"
    write_json(record_path, record)
    manifest = load_manifest(project_dir)
    manifest["spec_review_status"] = "confirmed"
    manifest["spec_confirmation_status"] = "confirmed"
    manifest["artifacts"]["spec_confirmation"] = artifact_entry(record_path)
    write_json(project_dir / "manifest.json", manifest)
    return record


def require_spec_confirmation(spec_path: Path, project_dir: Path) -> dict[str, Any]:
    record_path = project_dir / "spec-confirmation.json"
    if not record_path.is_file():
        raise FigureError(
            "提取前缺少用户规格确认：请先运行 spec-review，向用户展示原图与规格叠图，再用 spec-confirm 记录其明确确认"
        )
    record = read_json(record_path)
    if record.get("schema") != SPEC_CONFIRMATION_SCHEMA or record.get("status") != "confirmed":
        raise FigureError("spec-confirmation.json 无效，请重新完成规格确认")
    spec = read_json(spec_path)
    if record.get("project_sha256") != sha256_file(spec_path):
        raise FigureError("project.json 在用户确认后已变化，必须重新进行规格复核与确认")
    if record.get("source_sha256") != spec["source"]["sha256"]:
        raise FigureError("来源哈希在用户确认后已变化，必须重新进行规格复核与确认")
    if not str(record.get("confirmed_by", "")).strip() or not str(record.get("confirmation", "")).strip():
        raise FigureError("规格确认记录缺少确认人或原始确认语句")
    return record


def assign_candidate_ids(chart_type: str, rows: list[dict[str, Any]]) -> None:
    counters: dict[str, int] = {}
    for row in rows:
        series = str(row.get("series", "series"))
        counters[series] = counters.get(series, 0) + 1
        safe_series = "".join(character if character.isalnum() else "-" for character in series)
        row["candidate_id"] = f"{chart_type}-{safe_series}-{counters[series]:06d}"


def evaluate_quality_gates(
    spec: dict[str, Any],
    chart_type: str,
    rows: list[dict[str, Any]],
    diagnostics: dict[str, Any],
    x_map: AxisMap | None,
    y_map: AxisMap,
    calibration_reports: dict[str, dict[str, Any]] | None = None,
) -> dict[str, Any]:
    configured = spec.get("quality_gates", {})
    checks: list[dict[str, Any]] = []

    calibration = configured.get("calibration", {})
    require_three = bool(calibration.get("require_three_anchors", False))
    maximum_rmse = calibration.get("max_normalized_rmse")
    mappings = (
        (("radius_axis", y_map),)
        if x_map is None
        else (("x_axis", x_map), ("y_axis", y_map))
    )
    for axis_name, mapping in mappings:
        if require_three:
            checks.append(
                {
                    "name": f"{axis_name}.anchor_count",
                    "status": "pass" if mapping.rmse_evaluable else "failed",
                    "observed": len(mapping.pixels),
                    "threshold": ">=3",
                }
            )
        if maximum_rmse is not None:
            checks.append(
                {
                    "name": f"{axis_name}.normalized_rmse",
                    "status": (
                        "pass"
                        if mapping.rmse_evaluable
                        and mapping.normalized_rmse <= float(maximum_rmse)
                        else "failed"
                    ),
                    "observed": mapping.normalized_rmse,
                    "threshold": f"<={float(maximum_rmse)}",
                    "evaluable": mapping.rmse_evaluable,
                }
            )
    if calibration_reports:
        for axis_name, report in calibration_reports.items():
            anchor_count = int(report.get("anchor_count", 0))
            evaluable = bool(report.get("rmse_evaluable", anchor_count >= 3))
            normalized_rmse = float(report.get("normalized_rmse", math.inf))
            if require_three:
                checks.append(
                    {
                        "name": f"{axis_name}.anchor_count",
                        "status": "pass" if evaluable else "failed",
                        "observed": anchor_count,
                        "threshold": ">=3",
                    }
                )
            if maximum_rmse is not None:
                checks.append(
                    {
                        "name": f"{axis_name}.normalized_rmse",
                        "status": (
                            "pass"
                            if evaluable and normalized_rmse <= float(maximum_rmse)
                            else "failed"
                        ),
                        "observed": normalized_rmse,
                        "threshold": f"<={float(maximum_rmse)}",
                        "evaluable": evaluable,
                    }
                )

    chart_gates = configured.get("bar" if chart_type == "histogram" else chart_type, {})
    if chart_type in {"line", "polar_line"}:
        for series in diagnostics["series"]:
            if str(series.get("extraction_mode")) == "marker_centers":
                minimum_markers = int(series.get("minimum_declared_marker_count", 1))
                accepted_markers = int(series.get("accepted_markers", 0))
                checks.append(
                    {
                        "name": f"{series['id']}.accepted_markers",
                        "status": "pass" if accepted_markers >= minimum_markers else "failed",
                        "observed": accepted_markers,
                        "threshold": f">={minimum_markers}",
                    }
                )
                continue
            semantics = str(series.get("line_semantics", "solid"))
            minimum_coverage = float(
                chart_gates.get(
                    f"min_coverage_{semantics}", chart_gates.get("min_coverage", 0.5)
                )
            )
            maximum_gap = chart_gates.get(
                f"max_gap_fraction_{semantics}", chart_gates.get("max_gap_fraction")
            )
            checks.append(
                {
                    "name": f"{series['id']}.coverage",
                    "status": "pass" if float(series["coverage"]) >= minimum_coverage else "failed",
                    "observed": float(series["coverage"]),
                    "threshold": f">={minimum_coverage}",
                }
            )
            if maximum_gap is not None:
                checks.append(
                    {
                        "name": f"{series['id']}.maximum_gap_fraction",
                        "status": (
                            "pass"
                            if float(series["maximum_gap_fraction"]) <= float(maximum_gap)
                            else "failed"
                        ),
                        "observed": float(series["maximum_gap_fraction"]),
                        "threshold": f"<={float(maximum_gap)}",
                    }
                )
    else:
        minimum_components = int(chart_gates.get("min_accepted_components", 1))
        maximum_rejected_ratio = chart_gates.get("max_rejected_ratio")
        for series in diagnostics["series"]:
            accepted = int(series["accepted_components"])
            rejected = int(series["rejected_components"])
            checks.append(
                {
                    "name": f"{series['id']}.accepted_components",
                    "status": "pass" if accepted >= minimum_components else "failed",
                    "observed": accepted,
                    "threshold": f">={minimum_components}",
                }
            )
            if maximum_rejected_ratio is not None:
                ratio = rejected / max(1, accepted + rejected)
                checks.append(
                    {
                        "name": f"{series['id']}.rejected_ratio",
                        "status": "pass" if ratio <= float(maximum_rejected_ratio) else "failed",
                        "observed": ratio,
                        "threshold": f"<={float(maximum_rejected_ratio)}",
                    }
                )

    failed = [check for check in checks if check["status"] == "failed"]
    if not rows:
        status = "failed"
    elif failed:
        status = "partial"
    else:
        status = "pass"
    return {
        "status": status,
        "checks": checks,
        "说明": "质量门槛只用于筛查风险，不能替代人工语义复核。",
    }


def extract_command(spec_path: Path, out_dir: Path) -> dict[str, Any]:
    spec_path = spec_path.expanduser().resolve()
    out_dir.mkdir(parents=True, exist_ok=True)
    spec = read_json(spec_path)
    errors = validate_spec(spec, spec_path, extraction=True)
    if errors:
        report = {
            "schema": "more-sci-figure.extraction-report.v1",
            "status": "failed",
            "numeric_output_authorized": False,
            "errors": errors,
        }
        write_json(out_dir / "extraction-report.json", report)
        raise FigureError("；".join(errors))

    specification_confirmation = require_spec_confirmation(spec_path, out_dir)

    image, measurement_path = prepare_measurement_raster(spec, spec_path, out_dir)
    array = np.asarray(image, dtype=np.uint8)
    box = tuple(int(round(value)) for value in spec["chart"]["plot_box"])
    left, top, right, bottom = box
    if left < 0 or top < 0 or right >= image.width or bottom >= image.height:
        raise FigureError("chart.plot_box 超出测量栅格范围")
    chart_type = spec["chart"]["type"]
    calibration_reports: dict[str, dict[str, Any]] = {}
    if chart_type == "polar_line":
        polar = spec["chart"]["polar"]
        x_map = None
        y_map = AxisMap(polar["radius_axis"])
        rows, diagnostics = extract_polar_line(
            array, box, spec["chart"]["series"], polar, y_map
        )
        calibration_reports = {
            "angle_axis": diagnostics["angle_axis"],
            "radius_axis": diagnostics["radius_axis"],
        }
    else:
        x_map = AxisMap(spec["chart"]["x_axis"])
        y_map = AxisMap(spec["chart"]["y_axis"])
    extractors: dict[str, Callable[..., tuple[list[dict[str, Any]], dict[str, Any]]]] = {
        "line": extract_line,
        "scatter": extract_scatter,
    }
    if chart_type == "polar_line":
        pass
    elif chart_type in extractors:
        rows, diagnostics = extractors[chart_type](
            array, box, spec["chart"]["series"], x_map, y_map
        )
    else:
        rows, diagnostics = extract_bars(
            array, box, spec["chart"]["series"], x_map, y_map, spec["chart"]
        )
    assign_candidate_ids(chart_type, rows)
    quality = evaluate_quality_gates(
        spec,
        chart_type,
        rows,
        diagnostics,
        x_map,
        y_map,
        {"angle_axis": calibration_reports["angle_axis"]}
        if chart_type == "polar_line"
        else None,
    )
    status = quality["status"]
    candidates_path = out_dir / "candidates.csv"
    overlay_path = out_dir / "overlay.png"
    write_csv(candidates_path, rows)
    draw_overlay(image, chart_type, rows, overlay_path)
    report = {
        "schema": "more-sci-figure.extraction-report.v1",
        "status": status,
        "numeric_output_authorized": False,
        "source": {
            "path": str(resolve_from(spec_path, spec["source"]["path"])),
            "sha256": spec["source"]["sha256"],
            "measurement_raster": str(measurement_path),
            "measurement_sha256": sha256_file(measurement_path),
            "width": image.width,
            "height": image.height,
        },
        "chart_type": chart_type,
        "specification_confirmation": {
            "status": "confirmed",
            "confirmed_by": specification_confirmation["confirmed_by"],
            "confirmed_at": specification_confirmation["confirmed_at"],
            "confirmation": specification_confirmation["confirmation"],
            "record_sha256": sha256_file(out_dir / "spec-confirmation.json"),
        },
        "plot_box": list(box),
        "calibration": (
            calibration_reports
            if chart_type == "polar_line"
            else {"x_axis": x_map.report(), "y_axis": y_map.report()}
        ),
        "rows": len(rows),
        "diagnostics": diagnostics,
        "quality_gates": quality,
        "review_status": "not_run",
        "limitations": "仅包含有颜色像素证据的可见标记；不会推断隐藏、粘连或遮挡数据。",
        "required_next": "查看 AI 综合评估与异常组；风险允许时由用户通过对话批量确认，之后才能创建正式 data.csv。",
    }
    report_path = out_dir / "extraction-report.json"
    write_json(report_path, report)
    project_copy = out_dir / "project.json"
    if project_copy.resolve() != spec_path:
        write_json(project_copy, project_copy_with_rebased_paths(spec, spec_path, project_copy))
    manifest = load_manifest(out_dir)
    manifest.update(
        {
            "project_id": spec.get("project_id"),
            "project_spec": artifact_entry(project_copy if project_copy.exists() else spec_path),
            "source_sha256": spec["source"]["sha256"],
            "spec_review_status": "confirmed",
            "spec_confirmation_status": "confirmed",
            "extraction_status": status,
            "review_status": "not_run",
            "tool_version": VERSION,
        }
    )
    manifest["artifacts"].update(
        {
            "project_spec": artifact_entry(project_copy if project_copy.exists() else spec_path),
            "candidates": artifact_entry(candidates_path),
            "overlay": artifact_entry(overlay_path),
            "extraction_report": artifact_entry(report_path),
            "spec_confirmation": artifact_entry(out_dir / "spec-confirmation.json"),
        }
    )
    source_report = out_dir / "source-report.json"
    if source_report.is_file():
        manifest["artifacts"]["source_report"] = artifact_entry(source_report)
    write_json(out_dir / "manifest.json", manifest)
    if rows:
        review_assess_command(out_dir)
        review_command(out_dir)
    return report


def review_assess_command(project_dir: Path) -> dict[str, Any]:
    """Fuse extraction evidence into an auditable risk score and anomaly shortlist."""
    project_dir = project_dir.expanduser().resolve()
    candidates_path = project_dir / "candidates.csv"
    report_path = project_dir / "extraction-report.json"
    if not candidates_path.is_file() or not report_path.is_file():
        raise FigureError("缺少 candidates.csv 或 extraction-report.json，请先运行 extract")
    rows = read_tabular_rows(candidates_path)
    if not rows:
        raise FigureError("没有可评估的候选值")
    report = read_json(report_path)
    candidate_hash = sha256_file(candidates_path)
    manifest = load_manifest(project_dir)

    critical_issues: list[str] = []
    high_risk_issues: list[str] = []
    warnings: list[str] = []
    provenance_warnings: list[str] = []

    expected_project_id = str(manifest.get("project_id", ""))
    expected_source_hash = str(manifest.get("source_sha256", ""))
    candidate_project_paths = []
    for path in (project_dir / "project.json", project_dir.parent / "project.json"):
        resolved = path.resolve()
        if resolved not in candidate_project_paths and resolved.is_file():
            candidate_project_paths.append(resolved)
    valid_projects: list[tuple[Path, dict[str, Any], str]] = []
    invalid_project_reasons: list[str] = []
    for path in candidate_project_paths:
        try:
            candidate_project = read_json(path)
        except (OSError, json.JSONDecodeError) as exc:
            invalid_project_reasons.append(f"{path}: JSON 无效（{exc}）")
            continue
        if candidate_project.get("schema") != SCHEMA:
            invalid_project_reasons.append(f"{path}: schema 不匹配")
            continue
        if expected_project_id and str(candidate_project.get("project_id", "")) != expected_project_id:
            invalid_project_reasons.append(f"{path}: project_id 不匹配")
            continue
        source = candidate_project.get("source", {})
        if expected_source_hash and str(source.get("sha256", "")) != expected_source_hash:
            invalid_project_reasons.append(f"{path}: source_sha256 不匹配")
            continue
        resolved_source = resolve_from(path, str(source.get("path", "")))
        if not resolved_source.is_file():
            invalid_project_reasons.append(f"{path}: 来源路径无效")
            continue
        if source.get("sha256") and sha256_file(resolved_source) != source.get("sha256"):
            invalid_project_reasons.append(f"{path}: 来源文件哈希不一致")
            continue
        measurement_value = source.get("measurement_raster")
        resolved_measurement = (
            resolve_from(path, str(measurement_value)) if measurement_value else resolved_source
        )
        if not resolved_measurement.is_file():
            invalid_project_reasons.append(f"{path}: 测量栅格路径无效")
            continue
        if source.get("measurement_sha256") and sha256_file(resolved_measurement) != source.get(
            "measurement_sha256"
        ):
            invalid_project_reasons.append(f"{path}: 测量栅格哈希不一致")
            continue
        semantic_copy = json.loads(json.dumps(candidate_project, ensure_ascii=False))
        semantic_copy["source"]["path"] = str(resolved_source)
        if measurement_value:
            semantic_copy["source"]["measurement_raster"] = str(resolved_measurement)
        fingerprint = hashlib.sha256(
            json.dumps(semantic_copy, ensure_ascii=False, sort_keys=True).encode("utf-8")
        ).hexdigest()
        valid_projects.append((path, candidate_project, fingerprint))
    semantic_fingerprints = {item[2] for item in valid_projects}
    if not valid_projects:
        critical_issues.append("同级或相邻目录中没有来源与哈希均有效的唯一 project.json")
        project_path = project_dir / "project.json"
        project: dict[str, Any] = {}
    elif len(semantic_fingerprints) > 1:
        critical_issues.append("同级或相邻目录中存在多个语义不同且有效的 project.json")
        project_path, project, _ = valid_projects[0]
    else:
        preferred = next(
            (item for item in valid_projects if item[0] == (project_dir / "project.json").resolve()),
            valid_projects[0],
        )
        project_path, project, _ = preferred
        if project_path != (project_dir / "project.json").resolve():
            provenance_warnings.append(
                "同级项目规格无效；已按来源哈希唯一匹配相邻目录中的有效规格"
            )
    if invalid_project_reasons:
        provenance_warnings.append(
            "已排除无效项目规格：" + "；".join(invalid_project_reasons)
        )
    warnings.extend(provenance_warnings)

    manifest_candidate = manifest.get("artifacts", {}).get("candidates", {})
    expected_candidate_hash = str(manifest_candidate.get("sha256", ""))
    if expected_candidate_hash and expected_candidate_hash != candidate_hash:
        critical_issues.append("manifest 中的候选哈希与当前 candidates.csv 不一致")
    extraction_status = str(report.get("status", manifest.get("extraction_status", "not_run")))
    if extraction_status == "failed":
        critical_issues.append("自动提取质量门失败")
    elif extraction_status == "partial":
        high_risk_issues.append("自动提取质量门仅部分通过")

    quality_checks = report.get("quality_gates", {}).get("checks", [])
    failed_checks = [
        check for check in quality_checks if isinstance(check, dict) and check.get("status") == "failed"
    ]
    if failed_checks:
        high_risk_issues.append(f"{len(failed_checks)} 项自动质量检查未通过")

    candidate_ids = [str(row.get("candidate_id", "")) for row in rows]
    duplicate_ids = len(candidate_ids) - len(set(candidate_ids))
    if not all(candidate_ids):
        critical_issues.append("候选表包含空 candidate_id")
    if duplicate_ids:
        critical_issues.append(f"候选表包含 {duplicate_ids} 个重复 candidate_id")

    evidence_counts: dict[str, int] = {}
    series_rows: dict[str, list[dict[str, Any]]] = {}
    anomaly_rows: list[dict[str, Any]] = []
    anomaly_ids: set[str] = set()

    def add_anomaly(
        row: dict[str, Any], severity: str, anomaly_type: str, reason: str, value: Any = ""
    ) -> None:
        candidate_id = str(row.get("candidate_id", ""))
        key = f"{candidate_id}:{anomaly_type}"
        if key in anomaly_ids:
            return
        anomaly_ids.add(key)
        anomaly_rows.append(
            {
                "candidate_id": candidate_id,
                "series": str(row.get("series", "")),
                "severity": severity,
                "anomaly_type": anomaly_type,
                "observed": value,
                "reason": reason,
                "recommended_action": "targeted_review" if severity == "high" else "batch_review",
            }
        )

    numeric_fields = (
        ("x", "y")
        if report.get("chart_type") in {"line", "polar_line", "scatter"}
        else ()
    )
    for row in rows:
        series_rows.setdefault(str(row.get("series", "")), []).append(row)
        evidence = str(row.get("evidence_status", row.get("status", "unknown")))
        evidence_counts[evidence] = evidence_counts.get(evidence, 0) + 1
        for field in numeric_fields:
            try:
                value = float(row[field])
            except (KeyError, TypeError, ValueError):
                add_anomaly(row, "high", "non_finite_coordinate", f"{field} 不是有限数值")
                continue
            if not math.isfinite(value):
                add_anomaly(row, "high", "non_finite_coordinate", f"{field} 不是有限数值")
        if evidence == "ambiguous_shared_colour":
            add_anomaly(row, "high", "ambiguous_assignment", "同色证据无法唯一归属")
        support_value = row.get("support_pixels")
        if support_value not in {None, ""}:
            try:
                support = float(support_value)
            except (TypeError, ValueError):
                support = math.nan
            if not math.isfinite(support) or support <= 0:
                add_anomaly(row, "high", "invalid_pixel_support", "缺少有效像素支持", support_value)

    for series, grouped_rows in series_rows.items():
        residual_pairs: list[tuple[dict[str, Any], float]] = []
        for row in grouped_rows:
            value = row.get("guide_residual_px")
            if value in {None, ""}:
                continue
            try:
                residual = float(value)
            except (TypeError, ValueError):
                continue
            if math.isfinite(residual):
                residual_pairs.append((row, residual))
        if len(residual_pairs) < 8:
            continue
        values = np.asarray([value for _, value in residual_pairs], dtype=float)
        median = float(np.median(values))
        mad = float(np.median(np.abs(values - median)))
        threshold = max(5.0, median + 6.0 * max(mad, 0.25))
        for row, residual in residual_pairs:
            if residual > threshold:
                add_anomaly(
                    row,
                    "medium",
                    "guide_residual_outlier",
                    f"系列 {series} 的引导残差显著高于稳健阈值 {threshold:.2f}px",
                    round(residual, 4),
                )

    model_assisted = evidence_counts.get("model_assisted_exclusive_assignment", 0)
    ambiguous = evidence_counts.get("ambiguous_shared_colour", 0)
    medium_anomalies = sum(item["severity"] == "medium" for item in anomaly_rows)
    high_anomalies = sum(item["severity"] == "high" for item in anomaly_rows)
    model_ratio = model_assisted / len(rows)
    if model_assisted:
        warnings.append(
            f"{model_assisted} 个候选使用模型辅助排他分配，已按系列汇总，不要求逐点点击"
        )

    reviewed_anomaly_ids: set[str] = set()
    review_record_ready = False
    decisions_path = project_dir / "review-decisions.json"
    if decisions_path.is_file():
        try:
            decisions_payload = read_json(decisions_path)
            if decisions_payload.get("candidate_sha256") == candidate_hash:
                decision_items = [
                    item
                    for item in decisions_payload.get("decisions", [])
                    if isinstance(item, dict)
                ]
                decision_by_id = {
                    str(item.get("candidate_id", "")): item for item in decision_items
                }
                completed_decisions = {
                    "accepted",
                    "rejected",
                    "corrected",
                    "reassigned",
                }
                reviewed_anomaly_ids = {
                    candidate_id
                    for candidate_id, item in decision_by_id.items()
                    if item.get("decision") in completed_decisions
                }
                review_record_ready = (
                    bool(str(decisions_payload.get("reviewed_by", "")).strip())
                    and set(candidate_ids) == set(decision_by_id)
                    and all(
                        item.get("decision") in completed_decisions
                        for item in decision_by_id.values()
                    )
                )
            else:
                warnings.append("已有复核文件的候选哈希不匹配；不得作为已解决证据")
        except (OSError, json.JSONDecodeError):
            warnings.append("已有复核文件无法读取；不计为异常已解决")
    anomaly_candidate_ids = {str(item["candidate_id"]) for item in anomaly_rows}
    unresolved_anomaly_ids = sorted(anomaly_candidate_ids - reviewed_anomaly_ids)
    unresolved_high_anomalies = sum(
        item["severity"] == "high" and str(item["candidate_id"]) in unresolved_anomaly_ids
        for item in anomaly_rows
    )
    unresolved_medium_anomalies = sum(
        item["severity"] == "medium" and str(item["candidate_id"]) in unresolved_anomaly_ids
        for item in anomaly_rows
    )

    series_assessment: list[dict[str, Any]] = []
    diagnostics = {
        str(item.get("id", "")): item
        for item in report.get("diagnostics", {}).get("series", [])
        if isinstance(item, dict)
    }
    project_quality = project.get("quality_gates", {}) if isinstance(project, dict) else {}
    chart_type = str(report.get("chart_type", "line"))
    chart_gate_name = "bar" if chart_type == "histogram" else chart_type
    chart_gates = (
        project_quality.get(chart_gate_name, {})
        if isinstance(project_quality, dict)
        else {}
    )
    series_scores: list[float] = []
    for series, grouped_rows in sorted(series_rows.items()):
        diagnostic = diagnostics.get(series, {})
        coverage = diagnostic.get("coverage")
        maximum_gap = diagnostic.get("maximum_gap_fraction")
        semantics = str(diagnostic.get("line_semantics", "solid"))
        if chart_type in {"line", "polar_line"}:
            minimum_coverage = float(
                chart_gates.get(
                    f"min_coverage_{semantics}", chart_gates.get("min_coverage", 0.5)
                )
            )
            maximum_gap_threshold = chart_gates.get(
                f"max_gap_fraction_{semantics}", chart_gates.get("max_gap_fraction")
            )
            maximum_gap_threshold = (
                float(maximum_gap_threshold) if maximum_gap_threshold is not None else 0.25
            )
            coverage_score = (
                min(100.0, float(coverage) / max(minimum_coverage, 1e-9) * 100.0)
                if isinstance(coverage, (int, float))
                else 85.0
            )
            gap_score = (
                100.0
                if isinstance(maximum_gap, (int, float))
                and float(maximum_gap) <= maximum_gap_threshold
                else min(
                    100.0,
                    maximum_gap_threshold / max(float(maximum_gap), 1e-9) * 100.0,
                )
                if isinstance(maximum_gap, (int, float))
                else 85.0
            )
            current_series_score = round((coverage_score + gap_score) / 2.0, 1)
        else:
            accepted_components = int(diagnostic.get("accepted_components", len(grouped_rows)))
            rejected_components = int(diagnostic.get("rejected_components", 0))
            current_series_score = round(
                accepted_components
                / max(1, accepted_components + rejected_components)
                * 100.0,
                1,
            )
        series_scores.append(current_series_score)
        series_assessment.append(
            {
                "series": series,
                "candidates": len(grouped_rows),
                "coverage": coverage,
                "maximum_gap_fraction": maximum_gap,
                "mean_guide_residual_px": diagnostic.get("mean_guide_residual_px"),
                "model_assisted": sum(
                    str(row.get("evidence_status", ""))
                    == "model_assisted_exclusive_assignment"
                    for row in grouped_rows
                ),
                "flagged_anomalies": sum(item["series"] == series for item in anomaly_rows),
                "series_quality_score": current_series_score,
            }
        )

    def bounded_score(value: float) -> float:
        return round(max(0.0, min(100.0, float(value))), 1)

    calibration = report.get("calibration", {})
    calibration_tolerance = 0.01
    configured_calibration = (
        project_quality.get("calibration", {}) if isinstance(project_quality, dict) else {}
    )
    if isinstance(configured_calibration, dict) and isinstance(
        configured_calibration.get("max_normalized_rmse"), (int, float)
    ):
        calibration_tolerance = max(
            1e-9, float(configured_calibration["max_normalized_rmse"])
        )
    calibration_metrics: list[dict[str, Any]] = []
    calibration_scores: list[float] = []
    normalized_rmse_values: list[float] = []
    for axis_name, axis in calibration.items():
        if not isinstance(axis, dict):
            continue
        anchors = int(axis.get("anchor_count", 0))
        evaluable = bool(axis.get("rmse_evaluable", anchors >= 3))
        rmse_value = axis.get("normalized_rmse")
        rmse = (
            float(rmse_value)
            if isinstance(rmse_value, (int, float)) and math.isfinite(float(rmse_value))
            else None
        )
        if rmse is not None:
            normalized_rmse_values.append(rmse)
        axis_score = (
            bounded_score(100.0 * min(1.0, calibration_tolerance / max(rmse, 1e-12)))
            if evaluable and rmse is not None
            else 85.0
            if anchors >= 2
            else 0.0
        )
        calibration_scores.append(axis_score)
        calibration_metrics.append(
            {
                "axis": axis_name,
                "anchor_count": anchors,
                "rmse_evaluable": evaluable,
                "normalized_rmse": rmse,
                "operational_tolerance": calibration_tolerance,
                "score": axis_score,
            }
        )
    calibration_score = min(calibration_scores) if calibration_scores else 0.0

    support_values: list[float] = []
    for row in rows:
        value = row.get("support_pixels")
        if value in {None, ""}:
            continue
        try:
            numeric_value = float(value)
        except (TypeError, ValueError):
            continue
        if math.isfinite(numeric_value):
            support_values.append(numeric_value)
    valid_support_ratio = (
        sum(value > 0 for value in support_values) / len(support_values)
        if support_values
        else None
    )
    support_score = valid_support_ratio * 100.0 if valid_support_ratio is not None else 85.0
    assignment_score = 100.0 - model_ratio * 40.0 - ambiguous / len(rows) * 100.0
    pixel_evidence_score = bounded_score(support_score * 0.6 + assignment_score * 0.4)

    worst_series_score = min(series_scores) if series_scores else 0.0
    mean_series_score = float(np.mean(series_scores)) if series_scores else 0.0
    series_quality_score = bounded_score((worst_series_score + mean_series_score) / 2.0)

    profile_name = str(
        project.get("assessment", {}).get("acceptance_profile", "engineering")
        if isinstance(project.get("assessment", {}), dict)
        else "engineering"
    )
    profile = ACCEPTANCE_PROFILES.get(profile_name, ACCEPTANCE_PROFILES["engineering"])

    normalized_uncertainties: list[float] = []
    candidate_normalized_uncertainty: dict[str, float] = {}
    axis_spans = declared_axis_spans(project)
    for series, grouped_rows in series_rows.items():
        for value_key, uncertainty_key in (
            ("x", "x_uncertainty"),
            ("y", "y_uncertainty"),
            ("value", "value_uncertainty"),
        ):
            valid_rows: list[tuple[dict[str, Any], float, float]] = []
            for row in grouped_rows:
                try:
                    numeric_value = float(row[value_key])
                    uncertainty_value = abs(float(row[uncertainty_key]))
                except (KeyError, TypeError, ValueError):
                    continue
                if math.isfinite(numeric_value) and math.isfinite(uncertainty_value):
                    valid_rows.append((row, numeric_value, uncertainty_value))
            if not valid_rows:
                continue
            span = axis_spans.get(value_key)
            if span is None:
                span = max(item[1] for item in valid_rows) - min(item[1] for item in valid_rows)
            if span <= 0:
                continue
            for row, _, uncertainty_value in valid_rows:
                normalized = uncertainty_value / span
                normalized_uncertainties.append(normalized)
                candidate_id = str(row.get("candidate_id", ""))
                candidate_normalized_uncertainty[candidate_id] = max(
                    normalized,
                    candidate_normalized_uncertainty.get(candidate_id, 0.0),
                )
    uncertainty_p95 = (
        float(np.percentile(np.asarray(normalized_uncertainties), 95))
        if normalized_uncertainties
        else None
    )
    guide_residuals = []
    for row in rows:
        try:
            value = float(row.get("guide_residual_px", ""))
        except (TypeError, ValueError):
            continue
        if math.isfinite(value):
            guide_residuals.append(abs(value))
    plot_box = report.get("plot_box", [])
    plot_height = (
        abs(float(plot_box[3]) - float(plot_box[1]))
        if isinstance(plot_box, list) and len(plot_box) == 4
        else 0.0
    )
    guide_residual_p95_fraction = (
        float(np.percentile(np.asarray(guide_residuals), 95)) / plot_height
        if guide_residuals and plot_height > 0
        else None
    )
    uncertainty_proxy_components = [
        bounded_score(100.0 - value * 500.0)
        for value in (uncertainty_p95, guide_residual_p95_fraction)
        if value is not None
    ]
    uncertainty_proxy_score = (
        bounded_score(float(np.mean(uncertainty_proxy_components)))
        if uncertainty_proxy_components
        else 85.0
    )

    stability_axes: list[dict[str, Any]] = []
    stability_scores: list[float] = []
    chart = project.get("chart", {}) if isinstance(project, dict) else {}
    stability_inputs: list[tuple[str, dict[str, Any], str]] = []
    if chart_type == "polar_line":
        radius_axis = chart.get("polar", {}).get("radius_axis", {})
        if isinstance(radius_axis, dict):
            stability_inputs.append(("radius_axis", radius_axis, "pixel_radius"))
    else:
        for axis_name, pixel_key in (("x_axis", "pixel_x"), ("y_axis", "pixel_y")):
            axis = chart.get(axis_name, {})
            if isinstance(axis, dict):
                stability_inputs.append((axis_name, axis, pixel_key))
    for axis_name, axis, pixel_key in stability_inputs:
        axis_rows: list[dict[str, Any]] = []
        pixels: list[float] = []
        for row in rows:
            try:
                pixel = float(row[pixel_key])
            except (KeyError, TypeError, ValueError):
                continue
            if math.isfinite(pixel):
                axis_rows.append(row)
                pixels.append(pixel)
        stability_report, local_shifts = anchor_jackknife_stability(axis, pixels)
        stability_report["axis"] = axis_name
        stability_axes.append(stability_report)
        if isinstance(stability_report.get("score"), (int, float)):
            stability_scores.append(float(stability_report["score"]))
        for row, shift in zip(axis_rows, local_shifts):
            candidate_id = str(row.get("candidate_id", ""))
            candidate_normalized_uncertainty[candidate_id] = max(
                shift,
                candidate_normalized_uncertainty.get(candidate_id, 0.0),
            )
    perturbation_score = min(stability_scores) if stability_scores else None
    perturbation_stability = {
        "status": "measured" if stability_scores else "not_evaluable",
        "method": "leave_one_anchor_out",
        "axes": stability_axes,
        "score": bounded_score(perturbation_score) if perturbation_score is not None else None,
        "说明": (
            "已自动完成锚点留一扰动；该检查不改写候选数据。"
            if stability_scores
            else "锚点不足，无法自动完成锚点留一扰动；不得把缺失检查视为通过。"
        ),
    }
    uncertainty_score = (
        bounded_score(min(uncertainty_proxy_score, perturbation_score))
        if perturbation_score is not None
        else uncertainty_proxy_score
    )

    uncertainty_threshold = max(
        0.0, (100.0 - float(profile["minimum_dimension_score"])) / 500.0
    )
    uncertainty_candidates: list[dict[str, Any]] = []
    uncertainty_groups: list[dict[str, Any]] = []
    for series, grouped_rows in sorted(series_rows.items()):
        selected: list[tuple[int, dict[str, Any], float]] = []
        for index, row in enumerate(grouped_rows):
            candidate_id = str(row.get("candidate_id", ""))
            normalized = candidate_normalized_uncertainty.get(candidate_id)
            if normalized is not None and normalized > uncertainty_threshold:
                selected.append((index, row, normalized))
        clusters: list[list[tuple[int, dict[str, Any], float]]] = []
        merge_gap = max(2, int(math.ceil(len(grouped_rows) * 0.02)))
        for item in selected:
            if not clusters or item[0] - clusters[-1][-1][0] > merge_gap:
                clusters.append([item])
            else:
                clusters[-1].append(item)
        for cluster_index, cluster in enumerate(clusters, start=1):
            representative = max(cluster, key=lambda item: item[2])
            group_id = f"{series}-uncertainty-{cluster_index:03d}"
            group = {
                "group_id": group_id,
                "series": series,
                "candidate_count": len(cluster),
                "start_candidate_id": str(cluster[0][1].get("candidate_id", "")),
                "end_candidate_id": str(cluster[-1][1].get("candidate_id", "")),
                "representative_candidate_id": str(
                    representative[1].get("candidate_id", "")
                ),
                "x_start": cluster[0][1].get("x", cluster[0][1].get("category")),
                "x_end": cluster[-1][1].get("x", cluster[-1][1].get("category")),
                "peak_normalized_uncertainty": round(representative[2], 8),
                "mean_normalized_uncertainty": round(
                    float(np.mean([item[2] for item in cluster])), 8
                ),
                "threshold": round(uncertainty_threshold, 8),
                "agent_action": "先检查图源分辨率、标定锚点和自动稳定性结果。",
                "user_action": "仅在 Agent 无法消除该区间不确定性时复核代表性局部证据。",
            }
            uncertainty_groups.append(group)
            for _, row, normalized in cluster:
                uncertainty_candidates.append(
                    {
                        "group_id": group_id,
                        "candidate_id": str(row.get("candidate_id", "")),
                        "series": series,
                        "x": row.get("x", row.get("category")),
                        "y": row.get("y", row.get("value")),
                        "normalized_uncertainty": round(normalized, 8),
                        "threshold": round(uncertainty_threshold, 8),
                        "review_role": "agent_first_user_only_if_unresolved",
                    }
                )
    uncertainty_groups.sort(
        key=lambda item: float(item["peak_normalized_uncertainty"]), reverse=True
    )
    priority_uncertainty_groups = uncertainty_groups[:12]

    anomaly_rate = len(unresolved_anomaly_ids) / len(rows)
    anomaly_health_score = bounded_score(
        100.0
        - unresolved_high_anomalies / len(rows) * 500.0
        - unresolved_medium_anomalies / len(rows) * 100.0
    )
    quality_gate_score = (
        bounded_score(
            sum(check.get("status") == "pass" for check in quality_checks)
            / len(quality_checks)
            * 100.0
        )
        if quality_checks
        else 85.0
    )
    provenance_score = 0.0 if critical_issues else 95.0 if provenance_warnings else 100.0

    dimensions = {
        "provenance_integrity": {
            "label": "来源与证据完整性",
            "weight": 10,
            "score": bounded_score(provenance_score),
            "metrics": {
                "critical_issue_count": len(critical_issues),
                "provenance_warning_count": len(provenance_warnings),
                "candidate_hash_matches_manifest": expected_candidate_hash in {"", candidate_hash},
            },
        },
        "calibration_quality": {
            "label": "坐标标定质量",
            "weight": 15,
            "score": bounded_score(calibration_score),
            "metrics": {"axes": calibration_metrics},
        },
        "pixel_evidence": {
            "label": "像素证据质量",
            "weight": 20,
            "score": pixel_evidence_score,
            "metrics": {
                "support_evaluable": valid_support_ratio is not None,
                "valid_support_ratio": valid_support_ratio,
                "model_assisted_ratio": round(model_ratio, 6),
                "ambiguous_ratio": round(ambiguous / len(rows), 6),
            },
        },
        "series_separation_continuity": {
            "label": "系列分离与连续性",
            "weight": 20,
            "score": series_quality_score,
            "metrics": {
                "worst_series_score": bounded_score(worst_series_score),
                "mean_series_score": bounded_score(mean_series_score),
            },
        },
        "uncertainty_stability": {
            "label": "不确定度与稳定性代理",
            "weight": 15,
            "score": uncertainty_score,
            "metrics": {
                "normalized_uncertainty_p95": uncertainty_p95,
                "guide_residual_p95_fraction": guide_residual_p95_fraction,
                "uncertainty_proxy_score": uncertainty_proxy_score,
                "perturbation_stability": perturbation_stability,
                "high_uncertainty_candidate_count": len(uncertainty_candidates),
                "high_uncertainty_group_count": len(uncertainty_groups),
                "operational_threshold": uncertainty_threshold,
            },
        },
        "anomaly_health": {
            "label": "异常负担",
            "weight": 10,
            "score": anomaly_health_score,
            "metrics": {
                "anomaly_rate": round(anomaly_rate, 6),
                "total_high_anomalies": high_anomalies,
                "total_medium_anomalies": medium_anomalies,
                "unresolved_high_anomalies": unresolved_high_anomalies,
                "unresolved_medium_anomalies": unresolved_medium_anomalies,
            },
        },
        "quality_gate_compliance": {
            "label": "项目质量门符合度",
            "weight": 10,
            "score": quality_gate_score,
            "metrics": {
                "checks_total": len(quality_checks),
                "checks_failed": len(failed_checks),
            },
        },
    }
    score = round(
        sum(float(item["score"]) * float(item["weight"]) for item in dimensions.values())
        / sum(float(item["weight"]) for item in dimensions.values()),
        1,
    )
    minimum_dimension_score = min(float(item["score"]) for item in dimensions.values())

    hard_gates = {
        "source_and_hash_integrity": not critical_issues,
        "automatic_quality_gates": extraction_status == "pass" and not failed_checks,
        "no_unresolved_high_anomalies": unresolved_high_anomalies == 0,
        "candidate_ids_complete": bool(candidate_ids) and not duplicate_ids and all(candidate_ids),
    }
    hard_gates_pass = all(hard_gates.values())
    score_threshold_pass = score >= float(profile["overall_threshold"])
    dimension_floor_pass = minimum_dimension_score >= float(
        profile["minimum_dimension_score"]
    )
    technical_eligibility = hard_gates_pass and score_threshold_pass and dimension_floor_pass
    formal_review_status = str(manifest.get("review_status", "not_run"))
    render_status = str(manifest.get("render_status", "not_run"))
    worst_dimension_key = min(
        dimensions,
        key=lambda key: float(dimensions[key]["score"]),
    )
    if (
        perturbation_score is not None
        and perturbation_score >= float(profile["minimum_dimension_score"])
        and uncertainty_proxy_score < float(profile["minimum_dimension_score"])
    ):
        uncertainty_repair_instruction = (
            f"锚点留一稳定性已达标（{bounded_score(perturbation_score)} 分），"
            f"主要问题是像素/局部测量不确定度；Agent 先检查 {len(priority_uncertainty_groups)} "
            "个最高风险代表区间并寻找更高分辨率或矢量图源，不要求用户调整锚点。"
        )
    elif perturbation_score is not None:
        uncertainty_repair_instruction = (
            f"锚点留一稳定性仅 {bounded_score(perturbation_score)} 分；"
            "Agent 先核对刻度锚点并用独立刻度重拟合，再比较候选偏移。"
        )
    else:
        uncertainty_repair_instruction = (
            "当前锚点不足以执行留一稳定性；Agent 先从原图定位额外独立刻度，"
            "只有刻度含义无法唯一确定时才请用户确认。"
        )
    repair_instructions = {
        "provenance_integrity": "重新锁定来源、项目规格和候选哈希后再评估。",
        "calibration_quality": "补充或核对坐标锚点，优先使用至少三个独立刻度后重新提取。",
        "pixel_evidence": "调整系列颜色容差、排除框或像素分离设置后重新提取。",
        "series_separation_continuity": "检查最差曲线的引导走廊、系列归属、覆盖率和真实缺口后重新提取。",
        "uncertainty_stability": uncertainty_repair_instruction,
        "anomaly_health": "先处理独立异常组；若异常集中成片，再调整提取参数。",
        "quality_gate_compliance": "查看失败的项目质量门，修复对应覆盖率、缺口或组件条件后重新提取。",
    }
    if not hard_gates["source_and_hash_integrity"] or not hard_gates["candidate_ids_complete"]:
        risk_level = "critical"
        recommendation = "stop"
        decision = "blocked"
        next_instruction = "停止：修复来源、项目规格或候选哈希问题后重新评估。"
    elif not hard_gates["automatic_quality_gates"]:
        risk_level = "high" if score < 80 or high_risk_issues else "medium"
        recommendation = "re_extract"
        decision = "not_qualified"
        next_instruction = "暂不接受：修复未通过的自动质量门并重新提取，然后再次评估。"
    elif unresolved_anomaly_ids:
        risk_level = "high" if unresolved_high_anomalies else "medium"
        recommendation = "review_anomaly_groups"
        decision = "targeted_review_required"
        next_instruction = f"复核 {len(unresolved_anomaly_ids)} 个独立异常候选；普通候选保持批量区。"
    elif not score_threshold_pass or not dimension_floor_pass:
        risk_level = "high" if score < 80 or high_risk_issues else "medium"
        recommendation = "re_extract"
        decision = "not_qualified"
        next_instruction = "暂不接受：" + repair_instructions[worst_dimension_key]
    elif review_record_ready and formal_review_status not in {"accepted", "partial"}:
        risk_level = "low"
        recommendation = "apply_review"
        decision = "review_record_ready"
        next_instruction = "复核记录已完整保存；请 Agent 校验哈希与覆盖后应用，生成正式 data.csv。"
    elif formal_review_status not in {"accepted", "partial"}:
        risk_level = "medium" if model_assisted else "low"
        recommendation = "batch_confirm"
        decision = "eligible_for_user_confirmation"
        next_instruction = "指标达到所选用途阈值且无未决异常；请用户确认后生成正式复核记录。"
    elif render_status != "pass":
        risk_level = "low"
        recommendation = "render_validate"
        decision = "extraction_accepted"
        next_instruction = "提取与复核已满足要求；继续生成 PNG/SVG/PDF 并执行 validate。"
    else:
        risk_level = "low"
        recommendation = "accept"
        decision = "accepted"
        next_instruction = "提取、复核和重绘均已通过；执行最终交付验证或接受当前版本。"

    responsibility = {
        "workflow_order": "agent_first_then_targeted_user_judgment",
        "agent_completed": [
            "七维质量评估",
            (
                "锚点留一扰动稳定性检查"
                if perturbation_stability["status"] == "measured"
                else "记录锚点扰动不可评估及原因"
            ),
            f"定位 {len(uncertainty_candidates)} 个高不确定候选并合并为 {len(uncertainty_groups)} 个区间",
        ],
        "agent_next": [],
        "user_required_now": False,
        "user_trigger": "仅当 Agent 完成自动诊断与安全重提取后仍不达标。",
        "user_tasks": [],
        "user_not_required": [
            "运行命令",
            "选择保存路径",
            "调整算法参数",
            "逐点复核全部普通候选",
        ],
    }
    if decision == "blocked":
        responsibility["agent_next"] = ["在项目范围内核对来源、规格和哈希，明确唯一阻断项。"]
        responsibility["user_tasks"] = ["仅在来源或项目规格无法唯一确定时选择正确文件。"]
    elif decision == "not_qualified":
        responsibility["agent_next"] = [
            repair_instructions[worst_dimension_key],
            "保留当前候选作为基线，比较重提取前后的最低维度和局部不确定区间。",
        ]
        responsibility["user_tasks"] = [
            "如有更高分辨率原图、矢量 PDF 或作者原始数据则提供。",
            "若自动改进仍不达标，只判断高不确定区间是否影响科研结论，或降低用途等级/拒绝使用。",
        ]
    elif decision == "targeted_review_required":
        responsibility["agent_next"] = [
            "只展示异常候选及其原始分辨率局部证据，普通候选保持批量区。"
        ]
        responsibility["user_required_now"] = True
        responsibility["user_tasks"] = [
            "仅对独立异常项选择接受、拒绝、校正或重归属。"
        ]
    elif decision == "eligible_for_user_confirmation":
        responsibility["agent_next"] = ["汇报用途阈值、硬门、最低维度和局限。"]
        responsibility["user_required_now"] = True
        responsibility["user_tasks"] = ["确认是否按当前用途批量接受普通候选。"]
    elif decision == "review_record_ready":
        responsibility["agent_next"] = ["校验复核哈希与覆盖并应用，生成正式 data.csv。"]
    elif decision == "extraction_accepted":
        responsibility["agent_next"] = ["生成 PNG/SVG/PDF 并执行 validate。"]
    else:
        responsibility["agent_next"] = ["完成最终交付验证并汇报路径。"]

    acceptance = {
        "profile": profile_name,
        "profile_label": profile["label"],
        "thresholds": {
            "overall_score": profile["overall_threshold"],
            "minimum_dimension_score": profile["minimum_dimension_score"],
        },
        "hard_gates": hard_gates,
        "hard_gates_pass": hard_gates_pass,
        "score_threshold_pass": score_threshold_pass,
        "dimension_floor_pass": dimension_floor_pass,
        "technical_eligibility": technical_eligibility,
        "unresolved_anomaly_count": len(unresolved_anomaly_ids),
        "unresolved_anomaly_ids": unresolved_anomaly_ids,
        "decision": decision,
        "next_instruction": next_instruction,
        "responsibility": responsibility,
        "qualification_note": "分数是操作门槛，不等同于统计准确率；硬门失败时高分也不得接受。",
    }

    severity_order = {"high": 0, "medium": 1, "low": 2}
    anomaly_rows.sort(
        key=lambda item: (severity_order.get(str(item["severity"]), 9), str(item["series"]), str(item["candidate_id"]))
    )
    assessment = {
        "schema": ASSESSMENT_SCHEMA,
        "tool_version": VERSION,
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "candidate_sha256": candidate_hash,
        "project_id": project.get("project_id"),
        "project_spec_path": str(project_path),
        "source_sha256": project.get("source", {}).get("sha256"),
        "candidate_count": len(rows),
        "overall_score": score,
        "minimum_dimension_score": round(minimum_dimension_score, 1),
        "risk_level": risk_level,
        "recommended_action": recommendation,
        "scorecard": {
            "method": "weighted_dimensions_with_non_compensable_hard_gates",
            "dimensions": dimensions,
            "worst_dimension": min(
                dimensions,
                key=lambda key: float(dimensions[key]["score"]),
            ),
            "worst_series_score": bounded_score(worst_series_score),
        },
        "acceptance": acceptance,
        "uncertainty_review": {
            "candidate_count": len(uncertainty_candidates),
            "group_count": len(uncertainty_groups),
            "groups": uncertainty_groups,
            "priority_group_count": len(priority_uncertainty_groups),
            "priority_groups": priority_uncertainty_groups,
            "default_owner": "agent",
            "user_review_policy": (
                "Agent 先完成自动稳定性诊断和安全重提取；仅把仍未解决的代表性区间交给用户，"
                "不得要求逐点复核全部高不确定候选。"
            ),
        },
        "indicator_summary": {
            "extraction_status": extraction_status,
            "quality_checks_total": len(quality_checks),
            "quality_checks_failed": len(failed_checks),
            "model_assisted_candidates": model_assisted,
            "model_assisted_ratio": round(model_ratio, 6),
            "ambiguous_candidates": ambiguous,
            "high_anomalies": high_anomalies,
            "medium_anomalies": medium_anomalies,
            "calibration_normalized_rmse": normalized_rmse_values,
        },
        "critical_issues": critical_issues,
        "high_risk_issues": high_risk_issues,
        "warnings": warnings,
        "evidence_status_counts": evidence_counts,
        "series_assessment": series_assessment,
        "anomaly_groups": {
            "ambiguous_assignment": sum(
                item["anomaly_type"] == "ambiguous_assignment" for item in anomaly_rows
            ),
            "invalid_or_non_finite": sum(
                item["anomaly_type"] in {"invalid_pixel_support", "non_finite_coordinate"}
                for item in anomaly_rows
            ),
            "guide_residual_outlier": sum(
                item["anomaly_type"] == "guide_residual_outlier" for item in anomaly_rows
            ),
            "model_assisted_assignment": model_assisted,
        },
        "top_anomalies": anomaly_rows[:20],
        "review_policy": {
            "default": "先看用途阈值、硬门、最低维度和异常，再决定下一步；总分不能覆盖失败硬门。",
            "batch_confirmation_allowed": recommendation == "batch_confirm",
            "formal_data_requires_confirmation": True,
            "point_table_role": "仅用于异常深挖或用户主动抽查。",
        },
    }
    assessment_path = project_dir / "review-assessment.json"
    anomalies_path = project_dir / "review-anomalies.csv"
    uncertainty_path = project_dir / "review-uncertainty.csv"
    write_json(assessment_path, assessment)
    write_csv(anomalies_path, anomaly_rows)
    write_csv(uncertainty_path, uncertainty_candidates)
    manifest["tool_version"] = VERSION
    manifest["artifacts"]["review_assessment"] = artifact_entry(assessment_path)
    manifest["artifacts"]["review_anomalies"] = artifact_entry(anomalies_path)
    manifest["artifacts"]["review_uncertainty"] = artifact_entry(uncertainty_path)
    write_json(project_dir / "manifest.json", manifest)
    return {
        "status": (
            "pass"
            if recommendation in {"batch_confirm", "apply_review", "render_validate", "accept"}
            else "attention_required"
        ),
        "overall_score": score,
        "risk_level": risk_level,
        "recommended_action": recommendation,
        "candidate_count": len(rows),
        "anomaly_count": len(anomaly_rows),
        "uncertainty_candidate_count": len(uncertainty_candidates),
        "uncertainty_group_count": len(uncertainty_groups),
        "acceptance_decision": decision,
        "acceptance_profile": profile_name,
        "minimum_dimension_score": round(minimum_dimension_score, 1),
        "assessment": str(assessment_path),
        "anomalies": str(anomalies_path),
        "uncertainty": str(uncertainty_path),
        "user_required_now": responsibility["user_required_now"],
        "agent_next": responsibility["agent_next"],
        "user_tasks": responsibility["user_tasks"],
        "下一步": next_instruction,
    }


def review_confirm_command(
    project_dir: Path,
    reviewed_by: str,
    confirmation: str,
    *,
    accept_anomalies: bool = False,
) -> dict[str, Any]:
    """Turn an explicit conversational approval into complete hash-bound decisions.

    ``accept_anomalies`` is deliberately opt-in: it is only valid after the user has
    explicitly authorized the reviewer to inspect the anomaly evidence and accept
    every listed anomaly. The original user wording remains bound into the record.
    """
    project_dir = project_dir.expanduser().resolve()
    reviewed_by = reviewed_by.strip()
    confirmation = confirmation.strip()
    if not reviewed_by:
        raise FigureError("批量确认必须记录 reviewed_by")
    if not confirmation:
        raise FigureError("批量确认必须记录用户的对话确认语句")
    assessment_path = project_dir / "review-assessment.json"
    candidates_path = project_dir / "candidates.csv"
    if not assessment_path.is_file() or not candidates_path.is_file():
        raise FigureError("缺少 review-assessment.json 或 candidates.csv，请先运行 review-assess")
    assessment = read_json(assessment_path)
    if assessment.get("schema") != ASSESSMENT_SCHEMA:
        raise FigureError(f"综合评估 schema 必须是 {ASSESSMENT_SCHEMA}")
    candidate_hash = sha256_file(candidates_path)
    if assessment.get("candidate_sha256") != candidate_hash:
        raise FigureError("综合评估绑定的候选哈希与当前 candidates.csv 不一致")
    anomalies_path = project_dir / "review-anomalies.csv"
    anomaly_rows = read_tabular_rows(anomalies_path) if anomalies_path.is_file() else []
    recommended_action = str(assessment.get("recommended_action", ""))
    allowed_action = recommended_action == "batch_confirm" or (
        accept_anomalies
        and bool(anomaly_rows)
        and recommended_action == "review_anomaly_groups"
    )
    if not allowed_action:
        raise FigureError(
            f"当前综合风险为 {assessment.get('risk_level')}，建议动作为 "
            f"{assessment.get('recommended_action')}；不得一键批量确认"
        )
    if anomaly_rows and not accept_anomalies:
        raise FigureError(
            f"存在 {len(anomaly_rows)} 个异常候选，必须先在独立异常复核区逐项处理；"
            "不得与普通候选一起批量确认"
        )
    rows = read_tabular_rows(candidates_path)
    assessment_hash = sha256_file(assessment_path)
    anomaly_groups = assessment.get("anomaly_groups", {})
    payload = {
        "schema": REVIEW_SCHEMA,
        "candidate_sha256": candidate_hash,
        "project_id": assessment.get("project_id"),
        "source_sha256": assessment.get("source_sha256"),
        "reviewed_by": reviewed_by,
        "reviewed_at": datetime.now(timezone.utc).isoformat(),
        "review_method": (
            "explicit_visual_anomaly_acceptance"
            if anomaly_rows
            else "ai_assisted_batch_confirmation"
        ),
        "assessment_sha256": assessment_hash,
        "assessment_score": assessment.get("overall_score"),
        "assessment_risk_level": assessment.get("risk_level"),
        "conversation_confirmation": confirmation,
        "anomaly_acknowledgement": {
            "confirmed": True,
            "candidate_count": len(anomaly_rows),
            "groups": anomaly_groups,
            "confirmation": confirmation,
            "accept_all_anomalies_authorized": bool(
                anomaly_rows and accept_anomalies
            ),
            "candidate_ids": [
                str(item.get("candidate_id", "")) for item in anomaly_rows
            ],
        },
        "decisions": [
            {
                "candidate_id": str(row["candidate_id"]),
                "decision": "accepted",
                "reason": "",
            }
            for row in rows
        ],
    }
    result = save_review_decisions_command(project_dir, payload)
    result.update(
        {
            "review_method": payload["review_method"],
            "assessment_score": payload["assessment_score"],
            "assessment_risk_level": payload["assessment_risk_level"],
            "confirmation": confirmation,
            "下一步": "Agent 应直接应用复核，并在其余质量门通过后完成重绘与验证。",
        }
    )
    return result


def review_command(project_dir: Path) -> dict[str, Any]:
    project_dir = project_dir.expanduser().resolve()
    candidates_path = project_dir / "candidates.csv"
    overlay_path = project_dir / "overlay.png"
    if not candidates_path.is_file() or not overlay_path.is_file():
        raise FigureError("缺少 candidates.csv 或 overlay.png，请先运行 extract")
    rows = read_tabular_rows(candidates_path)
    if not rows:
        raise FigureError("没有可供复核的候选值")
    project_path = project_dir / "project.json"
    project = read_json(project_path) if project_path.is_file() else {}
    assessment_path = project_dir / "review-assessment.json"
    if not assessment_path.is_file():
        review_assess_command(project_dir)
    assessment = read_json(assessment_path)
    anomalies_path = project_dir / "review-anomalies.csv"
    anomaly_rows = read_tabular_rows(anomalies_path) if anomalies_path.is_file() else []
    anomaly_by_id = {
        str(item.get("candidate_id", "")): item
        for item in anomaly_rows
        if str(item.get("candidate_id", ""))
    }
    assessment_score = escape(str(assessment.get("overall_score", "—")))
    assessment_risk = escape(str(assessment.get("risk_level", "unknown")))
    assessment_action_raw = str(assessment.get("recommended_action", "unknown"))
    action_labels = {
        "batch_confirm": "请用户批量确认",
        "apply_review": "应用已保存复核",
        "review_anomaly_groups": "只复核异常组",
        "re_extract": "修复后重新提取",
        "render_validate": "继续重绘并验证",
        "accept": "接受当前版本",
        "stop": "停止",
    }
    assessment_action = escape(
        f"{action_labels.get(assessment_action_raw, '需要处理')}（{assessment_action_raw}）"
    )
    assessment_anomalies = len(anomaly_rows)
    scorecard = assessment.get("scorecard", {})
    acceptance = assessment.get("acceptance", {})
    unresolved_assessment_anomalies = int(
        acceptance.get("unresolved_anomaly_count", assessment_anomalies)
    )
    dimensions = scorecard.get("dimensions", {}) if isinstance(scorecard, dict) else {}
    thresholds = acceptance.get("thresholds", {}) if isinstance(acceptance, dict) else {}
    try:
        dimension_floor = float(thresholds.get("minimum_dimension_score", 0.0))
    except (TypeError, ValueError):
        dimension_floor = 0.0

    def dimension_passes(item: dict[str, Any]) -> bool:
        try:
            return float(item.get("score", 0.0)) >= dimension_floor
        except (TypeError, ValueError):
            return False

    dimension_rows = "".join(
        "<tr>"
        f"<td>{escape(str(item.get('label', key)))}</td>"
        f"<td><strong>{escape(str(item.get('score', '—')))}</strong></td>"
        f"<td>{escape(str(item.get('weight', '—')))}%</td>"
        f"<td class='{'status-good' if dimension_passes(item) else 'status-error'}'>"
        f"{'达标' if dimension_passes(item) else '不足'}</td>"
        "</tr>"
        for key, item in dimensions.items()
        if isinstance(item, dict)
    )
    acceptance_decision = str(acceptance.get("decision", "unknown"))
    acceptance_labels = {
        "blocked": "停止",
        "not_qualified": "暂不合格",
        "targeted_review_required": "需要异常复核",
        "eligible_for_user_confirmation": "可以请用户确认",
        "review_record_ready": "复核记录已就绪",
        "extraction_accepted": "提取已接受，等待重绘验证",
        "accepted": "可以接受",
    }
    decision_class = (
        "status-good"
        if acceptance_decision in {
            "eligible_for_user_confirmation",
            "review_record_ready",
            "extraction_accepted",
            "accepted",
        }
        else "status-error"
        if acceptance_decision in {"blocked", "not_qualified"}
        else "status-warn"
    )
    hard_gates_pass = bool(acceptance.get("hard_gates_pass", False))
    next_instruction = escape(str(acceptance.get("next_instruction", "请查看异常与质量门。")))
    worst_dimension_key = str(scorecard.get("worst_dimension", ""))
    worst_dimension = dimensions.get(worst_dimension_key, {})
    worst_dimension_label = escape(
        str(worst_dimension.get("label", worst_dimension_key or "—"))
    )
    responsibility = acceptance.get("responsibility", {})

    def list_html(items: Any) -> str:
        if not isinstance(items, list) or not items:
            return "<li>无</li>"
        return "".join(f"<li>{escape(str(item))}</li>" for item in items)

    responsibility_html = (
        "<div class='responsibility-grid'>"
        "<div><h3>Agent 负责</h3>"
        f"<p><strong>已完成</strong></p><ul>{list_html(responsibility.get('agent_completed'))}</ul>"
        f"<p><strong>接下来</strong></p><ul>{list_html(responsibility.get('agent_next'))}</ul></div>"
        "<div><h3>用户只需</h3>"
        f"<p class='{'status-warn' if responsibility.get('user_required_now') else 'status-good'}'>"
        f"{'现在需要参与' if responsibility.get('user_required_now') else '现在无需逐点参与'}</p>"
        f"<ul>{list_html(responsibility.get('user_tasks'))}</ul>"
        f"<p><strong>触发条件：</strong>{escape(str(responsibility.get('user_trigger', '仅在自动处理无法解决时。')))}</p>"
        "</div></div>"
    )
    uncertainty_review = assessment.get("uncertainty_review", {})
    uncertainty_groups = (
        uncertainty_review.get("priority_groups", uncertainty_review.get("groups", []))
        if isinstance(uncertainty_review, dict)
        else []
    )
    uncertainty_group_rows = "".join(
        "<tr>"
        f"<td>{escape(str(item.get('group_id', '—')))}</td>"
        f"<td>{escape(str(item.get('series', '—')))}</td>"
        f"<td>{escape(str(item.get('candidate_count', '—')))}</td>"
        f"<td>{escape(str(item.get('x_start', '—')))} ～ {escape(str(item.get('x_end', '—')))}</td>"
        f"<td>{float(item.get('peak_normalized_uncertainty', 0.0)) * 100:.2f}%</td>"
        "</tr>"
        for item in uncertainty_groups[:12]
        if isinstance(item, dict)
    )
    uncertainty_html = (
        "<div class='uncertainty-summary'>"
        "<h3>高不确定区间（Agent 先处理）</h3>"
        f"<p>已将 {escape(str(uncertainty_review.get('candidate_count', 0)))} 个候选合并为 "
        f"{escape(str(uncertainty_review.get('group_count', 0)))} 个连续区间，页面只展示其中 "
        f"{escape(str(uncertainty_review.get('priority_group_count', len(uncertainty_groups))))} "
        "个最高风险代表区间，不要求用户逐点核对；完整清单保存在 "
        "<code>review-uncertainty.csv</code>。</p>"
        + (
            "<table class='dimension-table'><thead><tr><th>区间</th><th>系列</th>"
            "<th>候选数</th><th>x 范围</th><th>峰值</th></tr></thead>"
            f"<tbody>{uncertainty_group_rows}</tbody></table>"
            if uncertainty_group_rows
            else "<p class='status-good'>当前用途阈值下没有需要定位的高不确定区间。</p>"
        )
        + "</div>"
    )
    scorecard_html = (
        "<div class='scorecard-grid'>"
        "<div class='acceptance-result'>"
        f"<span>用途等级：{escape(str(acceptance.get('profile_label', acceptance.get('profile', '工程分析'))))}</span>"
        f"<strong class='{decision_class}'>判定：{escape(acceptance_labels.get(acceptance_decision, acceptance_decision))}</strong>"
        f"<span>合格线：总分 ≥ {escape(str(thresholds.get('overall_score', '—')))}；"
        f"单项最低 ≥ {escape(str(thresholds.get('minimum_dimension_score', '—')))}</span>"
        f"<span>当前最低维度：{escape(str(assessment.get('minimum_dimension_score', '—')))}；"
        f"硬门：{'通过' if hard_gates_pass else '未通过'}</span>"
        f"<span>最弱项：{worst_dimension_label}；"
        f"最差曲线分：{escape(str(scorecard.get('worst_series_score', '—')))}</span>"
        f"<p><strong>下一步指令：</strong>{next_instruction}</p>"
        "</div>"
        "<div class='dimension-table-wrap'><table class='dimension-table'>"
        "<thead><tr><th>细化指标</th><th>得分</th><th>权重</th><th>判定</th></tr></thead>"
        f"<tbody>{dimension_rows}</tbody></table></div>"
        "</div>"
        f"{responsibility_html}"
        f"{uncertainty_html}"
    )
    assessment_hash = sha256_file(assessment_path)
    current_manifest = load_manifest(project_dir)
    formal_review_status = str(current_manifest.get("review_status", "not_run"))
    applied_review = {}
    applied_review_path = project_dir / "review-decisions.json"
    if applied_review_path.is_file():
        applied_review = read_json(applied_review_path)
    applied_decision_by_id = {
        str(item.get("candidate_id", "")): item
        for item in applied_review.get("decisions", [])
        if isinstance(item, dict) and str(item.get("candidate_id", ""))
    }
    reviewed_by_existing = escape(str(applied_review.get("reviewed_by", "")))
    anomaly_groups_text = "；".join(
        f"{escape(str(key))}={int(value)}"
        for key, value in assessment.get("anomaly_groups", {}).items()
        if isinstance(value, int) and value
    ) or "无异常组"
    if formal_review_status in {"accepted", "partial", "rejected"}:
        assessment_confirmation = (
            "<div class='assessment-confirm confirmed'>"
            "<h3>异常候选正式复核记录</h3>"
            f"<p><strong>异常分组：</strong>{anomaly_groups_text}</p>"
            f"<strong>已完成正式确认：{escape(formal_review_status)}</strong>"
            f"<span>下方已把 {assessment_anomalies} 个异常候选单独列出，并显示各自已应用的决定；"
            f"复核人：{reviewed_by_existing or '已记录'}。为保护证据链，现有正式记录只读。</span>"
            "</div>"
        )
    elif unresolved_assessment_anomalies:
        normal_count = len(rows) - assessment_anomalies
        assessment_confirmation = (
            "<div class='assessment-confirm attention'>"
            "<h3>先复核异常候选</h3>"
            f"<p><strong>{unresolved_assessment_anomalies} 个未决异常候选</strong>已在下方独立列出；"
            f"其余 {normal_count} 个普通候选位于单独的批量区。</p>"
            "<strong>异常项未全部接受、拒绝、校正或重归属前，不能生成正式复核文件。</strong>"
            "</div>"
        )
    elif assessment_action_raw == "apply_review":
        assessment_confirmation = (
            "<div class='assessment-confirm confirmed'>"
            "<h3>复核记录已保存</h3>"
            f"<p>复核人：<strong>{reviewed_by_existing or '已记录'}</strong>；"
            "候选哈希和决策覆盖将在正式应用时再次强制校验。</p>"
            "<strong>下一步请让 Agent 应用复核并生成正式 data.csv；无需再次批量确认。</strong>"
            "</div>"
        )
    elif assessment.get("recommended_action") == "batch_confirm":
        assessment_confirmation = (
            "<div class='assessment-confirm'>"
            "<h3>普通候选批量确认</h3>"
            "<p><strong>当前没有异常候选。</strong>可直接确认普通候选批次。</p>"
            "<label class='assessment-reviewer'>复核人或可追溯记录："
            "<input id='assessment-reviewed-by' required aria-required='true' placeholder='必填，例如：Dr.Jiang'></label>"
            "<label class='assessment-ack'><input type='checkbox' id='confirm-assessment-anomalies'>"
            "<span>我已查看综合评分并确认当前没有异常候选，同意批量接受普通候选；"
            "这不表示逐点人工测量。</span></label>"
            "<button type='button' id='confirm-assessment' disabled>确认普通候选批次并保存复核记录</button>"
            "<strong id='assessment-confirm-status' class='status-warn' role='status'>请填写复核人并勾选批量确认。</strong>"
            "</div>"
        )
    else:
        assessment_confirmation = (
            "<div class='assessment-confirm blocked'>"
            "<strong>当前风险不允许批量确认。</strong>请先处理异常组或修复阻断问题。"
            "</div>"
        )
    assessment_summary = (
        "<section class='assessment-card'>"
        "<h2>AI 综合评判</h2>"
        f"<div class='assessment-score'>{assessment_score}<small>/100</small></div>"
        f"<p><strong>风险等级：</strong>{assessment_risk}　"
        f"<strong>建议动作：</strong>{assessment_action}　"
        f"<strong>异常候选：</strong>{assessment_anomalies}</p>"
        f"{scorecard_html}"
        "<p>普通候选可批量确认；异常候选必须在独立区域查看局部证据并逐项决定，"
        "两者不会混在同一个批次中。只有没有异常项时，<code>batch_confirm</code> 才能直接完成。</p>"
        f"{assessment_confirmation}"
        "</section>"
    )
    series_ids = [
        str(item["id"])
        for item in project.get("chart", {}).get("series", [])
        if isinstance(item, dict) and item.get("id")
    ]
    if not series_ids:
        series_ids = sorted({str(row.get("series", "")) for row in rows if row.get("series")})
    candidate_hash = sha256_file(candidates_path)
    project_id = str(project.get("project_id", "")).strip()
    source_sha256 = str(project.get("source", {}).get("sha256", "")).strip()
    template = {
        "schema": REVIEW_SCHEMA,
        "candidate_sha256": candidate_hash,
        "review_method": "anomaly_first_split_review" if anomaly_rows else "manual_point_review",
        "anomaly_review": {
            "candidate_count": len(anomaly_rows),
            "candidate_ids": sorted(anomaly_by_id),
            "separated_from_normal_batch": True,
        },
        "reviewed_by": "",
        "reviewed_at": "",
        "decisions": [
            {
                "candidate_id": str(row["candidate_id"]),
                "decision": (
                    str(applied_decision_by_id[str(row["candidate_id"])].get("decision", "pending"))
                    if str(row["candidate_id"]) in applied_decision_by_id
                    else "pending"
                    if str(row["candidate_id"]) in anomaly_by_id
                    else "accepted"
                ),
                "reason": str(
                    applied_decision_by_id.get(str(row["candidate_id"]), {}).get("reason", "")
                ),
            }
            for row in rows
        ],
    }
    if project_id:
        template["project_id"] = project_id
    if source_sha256:
        template["source_sha256"] = source_sha256
    template_path = project_dir / "review-template.json"
    write_json(template_path, template)

    anomaly_table_rows: list[str] = []
    normal_table_rows: list[str] = []
    for row in rows:
        raw_candidate_id = str(row["candidate_id"])
        candidate_id = escape(raw_candidate_id)
        anomaly = anomaly_by_id.get(raw_candidate_id)
        applied_decision = applied_decision_by_id.get(raw_candidate_id, {})
        default_decision = str(
            applied_decision.get("decision")
            or ("pending" if anomaly is not None else "accepted")
        )
        selected_value = "" if default_decision == "pending" else default_decision
        readonly = formal_review_status in {"accepted", "partial", "rejected"}
        series = escape(str(row.get("series", "")))
        x_value = escape(str(row.get("x", row.get("x_value", row.get("category_index", "")))))
        y_value = escape(str(row.get("y", row.get("value", ""))))
        evidence_status = escape(str(row.get("evidence_status", row.get("status", ""))))
        target_series = str(applied_decision.get("target_series") or row.get("series", ""))
        target_options = "".join(
            f"<option value='{escape(value)}'{' selected' if value == target_series else ''}>{escape(value)}</option>"
            for value in series_ids
        )
        decision_options = "".join(
            f"<option value='{value}'{' selected' if value == selected_value else ''}>{label}</option>"
            for value, label in (
                ("", "待决策"),
                ("accepted", "接受"),
                ("rejected", "拒绝"),
                ("corrected", "校正坐标"),
                ("reassigned", "重新归属"),
            )
        )
        readonly_attr = " disabled data-readonly='true'" if readonly else ""
        reason_readonly_attr = " disabled" if readonly else ""
        corrected_x = escape(str(applied_decision.get("corrected_x", "")))
        corrected_y = escape(str(applied_decision.get("corrected_y", "")))
        reason = escape(str(applied_decision.get("reason", "")))
        common_cells = (
            f"<td><code>{candidate_id}</code></td>"
            f"<td>{series}</td><td>{x_value}</td><td>{y_value}</td><td>{evidence_status}</td>"
            f"<td><select data-decision='{candidate_id}' data-anomaly='{'true' if anomaly else 'false'}' "
            f"data-evidence-status='{evidence_status}' data-current-series='{series}'{readonly_attr}>"
            f"{decision_options}</select><span class='decision-state state-{escape(default_decision)}' "
            f"data-decision-state='{candidate_id}'>{escape('待决策' if default_decision == 'pending' else {'accepted': '接受', 'rejected': '拒绝', 'corrected': '校正坐标', 'reassigned': '重新归属'}.get(default_decision, default_decision))}</span></td>"
            f"<td><select data-target-series='{candidate_id}' disabled>{target_options}</select></td>"
            f"<td><input data-corrected-x='{candidate_id}' inputmode='decimal' value='{corrected_x}' placeholder='{x_value}' disabled></td>"
            f"<td><input data-corrected-y='{candidate_id}' inputmode='decimal' value='{corrected_y}' placeholder='{y_value}' disabled></td>"
            f"<td><input data-reason='{candidate_id}' value='{reason}' placeholder='必要时填写理由'{reason_readonly_attr}></td>"
        )
        if anomaly is not None:
            try:
                crop_left = 110 - float(row.get("pixel_x", 0))
                crop_top = 70 - float(row.get("pixel_y", 0))
            except (TypeError, ValueError):
                crop_left = 0.0
                crop_top = 0.0
            anomaly_detail = (
                f"<strong>{escape(str(anomaly.get('anomaly_type', '异常')))}</strong><br>"
                f"严重度：{escape(str(anomaly.get('severity', '')))}；"
                f"观测指标：{escape(str(anomaly.get('observed', '')))}<br>"
                f"{escape(str(anomaly.get('reason', '')))}"
            )
            crop = (
                "<div class='evidence-crop' title='红色十字为异常候选像素位置'>"
                f"<img src=\"overlay.png\" alt='异常候选 {candidate_id} 的局部证据' "
                f"style='left:{crop_left:.1f}px;top:{crop_top:.1f}px'>"
                "<span class='crop-crosshair' aria-hidden='true'></span></div>"
            )
            anomaly_table_rows.append(
                f"<tr class='decision-{escape(default_decision)} anomaly-row'>"
                f"<td>{crop}</td>{common_cells}<td class='anomaly-reason'>{anomaly_detail}</td></tr>"
            )
        else:
            normal_table_rows.append(
                f"<tr class='decision-{escape(default_decision)}'>{common_cells}</tr>"
            )
    payload_data = {
        "schema": REVIEW_SCHEMA,
        "candidate_sha256": candidate_hash,
        "review_method": "anomaly_first_split_review" if anomaly_rows else "manual_point_review",
        "anomaly_review": {
            "candidate_count": len(anomaly_rows),
            "candidate_ids": sorted(anomaly_by_id),
            "separated_from_normal_batch": True,
        },
    }
    if project_id:
        payload_data["project_id"] = project_id
    if source_sha256:
        payload_data["source_sha256"] = source_sha256
    payload = json.dumps(payload_data, ensure_ascii=False)
    review_apply_prompt = (
        "请使用 more-sci-figure skill 继续处理当前科研图表项目，并直接执行安全范围内的任务。"
        "不要假定固定文件名或固定路径：先在当前对话附件、当前工作区以及用户明确指定的位置中，"
        "定位 schema=more-sci-figure.review-decisions.v1 的人工复核 JSON；不得为寻找文件而扫描整个用户磁盘。"
        "读取其中的 candidate_sha256，并只接受 SHA-256 完全一致的 candidates.csv；"
        "再从该候选文件的同级或相邻项目目录定位 project.json。project_id 和 source_sha256 只能作为辅助线索。"
        "如果复核 JSON、候选文件或项目规格没有唯一匹配，立即停止并请用户选择或附加文件，不得猜测。"
        "本次只应用人工复核，不执行重绘或最终验证。核验候选覆盖完整、reviewed_by 非空后，"
        "再仅把 accepted、corrected 和 reassigned 项写入正式 data.csv，并保留原值、原系列、理由和动作类型。"
        "若任一复核门不满足，停止并明确报告，不得生成正式 data.csv 或提升 review_status。"
    )
    pipeline_prompt = ""
    if project_path.is_file():
        pipeline_prompt = (
            "请使用 more-sci-figure skill 继续当前科研图表项目，并直接完成安全范围内的完整后续流程。"
            "不要假定固定文件名或固定路径：先在当前对话附件、当前工作区以及用户明确指定的位置中，"
            "定位 schema=more-sci-figure.review-decisions.v1 的人工复核 JSON；不得为寻找文件而扫描整个用户磁盘。"
            "读取其中的 candidate_sha256，并只接受 SHA-256 完全一致的 candidates.csv；"
            "再从该候选文件的同级或相邻项目目录定位 project.json。project_id 和 source_sha256 只能作为辅助线索。"
            "如果复核 JSON、候选文件或项目规格没有唯一匹配，立即停止并请用户选择或附加文件，不得猜测。"
            "核验候选覆盖完整、reviewed_by 非空后，应用人工复核生成正式 data.csv，"
            "再按匹配到的项目规格完成论文级 PNG/SVG/PDF 重绘和 validate。"
            "保持 extraction_status、review_status、render_status、delivery_status 相互独立；"
            "若哈希、复核覆盖或任何质量门不通过，立即停止并明确报告，不得伪造正式数据或抬升状态。"
            "最后汇报接受、拒绝、校正、重归属数量，以及各状态和交付文件路径。"
        )
    html = """<!doctype html>
<html lang="zh-CN">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>more-sci-figure 候选数据复核</title>
<style>
:root { color-scheme: light; font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", sans-serif; }
body { margin: 0 auto; max-width: 1440px; padding: 24px; color: #17202a; background: #f6f8fa; }
h1 { margin-bottom: 8px; } .note { padding: 12px 16px; background: #fff8c5; border-left: 4px solid #d4a72c; }
.layout { display: grid; grid-template-columns: minmax(420px, 1fr) minmax(560px, 1.2fr); gap: 20px; margin-top: 20px; }
.panel { background: white; border: 1px solid #d0d7de; border-radius: 8px; padding: 16px; overflow: auto; }
img { max-width: none; image-rendering: auto; } table { width: 100%; border-collapse: collapse; font-size: 14px; }
th, td { border-bottom: 1px solid #d8dee4; padding: 8px; text-align: left; }
th { position: sticky; top: 0; background: #f6f8fa; } input, select { width: 100%; box-sizing: border-box; padding: 6px; }
tbody td { transition: background-color .18s ease; }
tr.decision-pending td { background: #ffffff; }
tr.decision-accepted td { background: #eaf8ef; }
tr.decision-rejected td { background: #fff0f0; }
tr.decision-corrected td { background: #fff8db; }
tr.decision-reassigned td { background: #edf5ff; }
tr.decision-accepted td:first-child { box-shadow: inset 4px 0 #16844b; }
tr.decision-rejected td:first-child { box-shadow: inset 4px 0 #cf222e; }
tr.decision-corrected td:first-child { box-shadow: inset 4px 0 #bf8700; }
tr.decision-reassigned td:first-child { box-shadow: inset 4px 0 #0969da; }
button { padding: 9px 16px; border: 0; border-radius: 6px; background: #0969da; color: white; cursor: pointer; }
button.secondary { background: #16844b; } button.danger { background: #cf222e; }
button:disabled { background: #8c959f; cursor: not-allowed; opacity: .72; }
.controls { position: sticky; top: 0; z-index: 3; display: flex; flex-wrap: wrap; align-items: center; gap: 8px; padding: 12px; background: white; border: 1px solid #d8dee4; border-radius: 8px; box-shadow: 0 3px 10px rgba(31, 35, 40, .08); }
.reviewer { display: flex; align-items: center; gap: 8px; min-width: 340px; font-weight: 600; }
.reviewer input { min-width: 230px; }
.summary { color: #57606a; font-size: 13px; }
.review-gates { display: grid; gap: 8px; margin: 10px 0; padding: 12px; border: 1px solid #d0d7de; border-radius: 8px; background: #fff; }
.review-gates label { display: flex; align-items: flex-start; gap: 8px; }
.review-gates input[type="checkbox"] { width: auto; margin-top: 3px; }
.readiness { flex-basis: 100%; font-size: 13px; font-weight: 700; }
.decision-state { display: inline-block; margin-top: 5px; padding: 2px 7px; border-radius: 999px; font-size: 12px; font-weight: 700; }
.state-pending { color: #57606a; background: #eaeef2; }
.state-accepted { color: #116329; background: #aceebb; }
.state-rejected { color: #a40e26; background: #ffcecb; }
.state-corrected { color: #7d4e00; background: #f8df8b; }
.state-reassigned { color: #0550ae; background: #b6d7ff; }
.toast { position: fixed; top: 18px; left: 50%; z-index: 20; transform: translate(-50%, -16px); min-width: 280px; max-width: min(720px, calc(100vw - 32px)); padding: 12px 18px; border-radius: 8px; color: white; background: #24292f; box-shadow: 0 8px 24px rgba(31, 35, 40, .25); text-align: center; opacity: 0; pointer-events: none; transition: opacity .18s ease, transform .18s ease; }
.toast.show { opacity: 1; transform: translate(-50%, 0); }
.toast.accepted { background: #16844b; } .toast.rejected { background: #cf222e; }
.handoff-note { margin: 12px 0 0; padding: 10px 12px; border-left: 4px solid #0969da; background: #ddf4ff; font-size: 14px; }
.export-result { margin: 12px 0; padding: 14px; border: 2px solid #0969da; border-radius: 8px; background: #f6fbff; }
.export-result[hidden] { display: none; }
.export-result h3 { margin: 0 0 10px; }
.status-grid { display: grid; grid-template-columns: max-content 1fr; gap: 6px 12px; margin-bottom: 10px; }
.status-good { color: #116329; } .status-warn { color: #9a6700; } .status-error { color: #cf222e; }
.export-actions { display: flex; flex-wrap: wrap; gap: 8px; margin: 10px 0; }
.compat-note { margin-bottom: 0; color: #57606a; font-size: 13px; }
.assessment-card { margin: 16px 0; padding: 16px 18px; border: 1px solid #54aeff; border-radius: 10px; background: #ddf4ff; }
.assessment-card h2 { margin: 0 0 8px; }
.assessment-score { float: right; margin: -42px 0 8px 18px; color: #0969da; font-size: 32px; font-weight: 800; }
.assessment-score small { font-size: 14px; font-weight: 600; }
.scorecard-grid { clear: both; display: grid; grid-template-columns: minmax(300px, .9fr) minmax(380px, 1.1fr); gap: 14px; margin: 14px 0; }
.acceptance-result, .dimension-table-wrap { padding: 12px; border: 1px solid #8c959f; border-radius: 8px; background: #fff; }
.acceptance-result { display: grid; gap: 7px; }
.acceptance-result p { margin: 4px 0 0; padding: 9px; border-left: 4px solid #0969da; background: #ddf4ff; }
.dimension-table { font-size: 13px; }
.dimension-table th { position: static; }
.responsibility-grid { clear: both; display: grid; grid-template-columns: 1fr 1fr; gap: 14px; margin: 14px 0; }
.responsibility-grid > div, .uncertainty-summary { padding: 14px; border: 1px solid #8c959f; border-radius: 8px; background: #fff; }
.responsibility-grid h3, .uncertainty-summary h3 { margin-top: 0; }
.responsibility-grid ul { margin: 6px 0 12px; padding-left: 20px; }
.uncertainty-summary { clear: both; overflow-x: auto; }
.assessment-confirm { clear: both; display: grid; gap: 10px; margin-top: 14px; padding: 14px; border: 1px solid #54aeff; border-radius: 8px; background: #fff; }
.assessment-confirm h3, .assessment-confirm p { margin: 0; }
.assessment-confirm.confirmed { border-color: #16844b; background: #dafbe1; }
.assessment-confirm.attention { border-color: #bf8700; background: #fff8c5; }
.assessment-confirm.blocked { border-color: #cf222e; background: #ffebe9; }
.assessment-reviewer { display: grid; grid-template-columns: max-content minmax(220px, 420px); align-items: center; gap: 10px; font-weight: 700; }
.assessment-ack { display: flex; align-items: flex-start; gap: 8px; }
.assessment-ack input[type="checkbox"] { width: auto; margin-top: 3px; }
.anomaly-review { margin-top: 18px; padding: 18px; border: 2px solid #bf8700; border-radius: 10px; background: #fff; }
.anomaly-review h2 { margin-top: 0; }
.anomaly-count { display: inline-block; margin-left: 6px; padding: 2px 9px; border-radius: 999px; color: #7d4e00; background: #f8df8b; font-size: 14px; }
.anomaly-table-wrap { max-height: 720px; overflow: auto; border: 1px solid #d0d7de; border-radius: 8px; }
.anomaly-table { min-width: 1540px; }
.anomaly-table th { top: 0; z-index: 2; }
.anomaly-row td { vertical-align: top; }
.anomaly-reason { min-width: 250px; max-width: 340px; }
.evidence-crop { position: relative; width: 220px; height: 140px; overflow: hidden; border: 1px solid #8c959f; border-radius: 6px; background: #fff; }
.evidence-crop img { position: absolute; max-width: none; }
.crop-crosshair { position: absolute; left: 105px; top: 65px; width: 10px; height: 10px; border: 2px solid #cf222e; border-radius: 50%; box-shadow: 0 0 0 1px white; }
.crop-crosshair::before, .crop-crosshair::after { content: ""; position: absolute; background: #cf222e; }
.crop-crosshair::before { left: 4px; top: -8px; width: 2px; height: 24px; }
.crop-crosshair::after { left: -8px; top: 4px; width: 24px; height: 2px; }
.normal-review, .full-evidence { margin-top: 14px; }
.normal-review > summary, .full-evidence > summary { padding: 12px 14px; border: 1px solid #d0d7de; border-radius: 8px; background: #f6f8fa; cursor: pointer; font-weight: 700; }
.normal-review table { min-width: 1180px; }
.normal-table-wrap { max-height: 620px; overflow: auto; margin-top: 10px; border: 1px solid #d0d7de; border-radius: 8px; }
.empty-state { padding: 18px; color: #57606a; text-align: center; background: #f6f8fa; border-radius: 8px; }
.command-card { margin: 14px 0; padding: 12px; border: 1px solid #d0d7de; border-radius: 8px; background: #f6f8fa; }
pre { margin: 8px 0; padding: 12px; overflow: auto; border-radius: 6px; background: #24292f; color: #f0f6fc; white-space: pre-wrap; overflow-wrap: anywhere; }
button.neutral { background: #57606a; }
@media (max-width: 980px) {
  .layout, .scorecard-grid, .responsibility-grid { grid-template-columns: 1fr; }
}
</style>
</head>
<body>
<h1>候选数据人工复核</h1>
<p class="note">默认采用 AI 综合评判与异常优先复核，不要求用户逐点点击。人工确认的是综合证据、风险分组和异常处置策略，而不是假装逐一测量了所有像素。</p>
__ASSESSMENT_SUMMARY__
<div id="decision-toast" class="toast" role="status" aria-live="polite"></div>
<section class="anomaly-review">
  <h2>异常候选独立复核 <span class="anomaly-count">__ANOMALY_COUNT__ 项</span></h2>
  <p>这里只显示异常候选。每项提供原始分辨率局部证据、异常原因和独立决定；普通候选不会混入本表。</p>
  <div class="controls">
    <label class="reviewer">复核人或复核记录：<input id="reviewed-by" required aria-required="true" autocomplete="name" value="__REVIEWED_BY__" placeholder="必填，例如：张三 / 项目组联合复核" __FORM_DISABLED__></label>
    <button type="button" id="accept-all-anomalies" class="secondary" __FORM_DISABLED__>接受全部异常项</button>
    <button type="button" id="reject-all-anomalies" class="danger" __FORM_DISABLED__>拒绝全部异常项</button>
    <button type="button" id="export" disabled>生成完整复核文件（下一步）</button>
    <span id="decision-summary" class="summary"></span>
    <span id="export-readiness" class="readiness status-warn" role="status">请先填写复核人并完成所有异常决策。</span>
  </div>
  <div class="anomaly-table-wrap">
    <table class="anomaly-table"><thead><tr><th>局部证据</th><th>候选编号</th><th>系列</th><th>X</th><th>Y/值</th><th>证据状态</th><th>决策</th><th>目标系列</th><th>校正 X</th><th>校正 Y</th><th>理由</th><th>异常说明</th></tr></thead>
    <tbody id="anomaly-decisions">__ANOMALY_TABLE_ROWS__</tbody></table>
  </div>
  __ANOMALY_EMPTY__
</section>

<details class="normal-review">
  <summary>普通候选批量区：__NORMAL_COUNT__ 项（与异常候选分开）</summary>
  <div class="panel">
    <p>普通候选默认批量接受。如需抽查，可在此单独修改；这里不包含任何已标记异常候选。</p>
    <div class="controls">
      <button type="button" id="accept-all-normal" class="secondary" __FORM_DISABLED__>普通候选全部接受</button>
      <button type="button" id="reject-all-normal" class="danger" __FORM_DISABLED__>普通候选全部拒绝</button>
    </div>
    <div class="normal-table-wrap"><table><thead><tr><th>候选编号</th><th>系列</th><th>X</th><th>Y/值</th><th>证据状态</th><th>决策</th><th>目标系列</th><th>校正 X</th><th>校正 Y</th><th>理由</th></tr></thead>
    <tbody id="normal-decisions">__NORMAL_TABLE_ROWS__</tbody></table></div>
  </div>
</details>

<details class="full-evidence">
  <summary>查看完整证据叠图</summary>
  <section class="panel"><img src="overlay.png" alt="候选值完整证据叠图"></section>
</details>

<section class="panel">
    <p class="handoff-note"><strong>后续流程：</strong>① 完成决策　② 在页面生成复核 JSON　③ 由本地 skill 固定保存到当前项目目录的 <code>review-decisions.json</code>　④ 把页面提供的任务语句交给 Codex、Claude Code、Hermes 等 Agent 继续处理。用户无需选择保存路径；只有 Agent 成功应用复核后才会生成正式 <code>data.csv</code>。</p>
    <div class="review-gates">
      <label id="uncertain-gate" hidden><input type="checkbox" id="confirm-uncertain"><span id="uncertain-confirm-label">我已人工核对所有模型辅助分配候选。</span></label>
      <label><input type="checkbox" id="confirm-export-scope"><span>我确认当前操作只生成复核 JSON，不会自动保存文件、生成 <code>data.csv</code>、重绘或推进交付状态。</span></label>
    </div>
    <section id="export-result" class="export-result" hidden aria-live="polite">
      <h3>复核文件状态</h3>
      <div class="status-grid">
        <span>内容生成：</span><strong id="generated-status" class="status-warn">未生成</strong>
        <span>本地保存：</span><strong id="save-status" class="status-warn">未保存</strong>
        <span>正式应用：</span><strong class="status-warn">未应用；尚未生成 data.csv</strong>
      </div>
      <p id="export-detail"></p>
      <div class="export-actions">
        <button type="button" id="save-review-file">保存到当前项目目录</button>
        <button type="button" id="copy-review-json" class="secondary">复制复核 JSON</button>
      </div>
      <p id="save-mode-note" class="compat-note">正式复核会话会把文件固定保存到当前项目目录并返回实际路径。直接打开 <code>file://review.html</code> 时为兼容模式，无法写入本地目录；请让 Agent 启动本地复核会话。复制成功不等于文件已保存。</p>
      <section id="pipeline-option" class="command-card">
        <h3>保存后推荐：交给 Agent 继续完整管线</h3>
        <p>把下面任务语句发送给 Codex、Claude Code、Hermes 等 Agent。提示词不锁定机器路径；Agent 会用复核文件中的候选哈希进行唯一匹配，无法唯一匹配时必须询问用户。</p>
        <pre><code id="pipeline-agent-prompt"></code></pre>
        <button type="button" data-copy-target="pipeline-agent-prompt">复制完整管线 Agent 任务语句</button>
      </section>
      <section class="command-card">
        <h3>保存后交给 Agent 仅应用复核</h3>
        <p>该任务语句只要求生成正式 <code>data.csv</code>，不执行重绘或最终验证。</p>
        <pre><code id="review-apply-agent-prompt"></code></pre>
        <button type="button" data-copy-target="review-apply-agent-prompt">复制仅应用复核 Agent 任务语句</button>
      </section>
    </section>
</section>
<script>
const base = __PAYLOAD__;
const reviewApplyAgentPrompt = __REVIEW_APPLY_PROMPT__;
const pipelineAgentPrompt = __PIPELINE_PROMPT__;
const assessmentSha256 = __ASSESSMENT_SHA256__;
const assessmentAnomalyCount = __ASSESSMENT_ANOMALY_COUNT__;
const formalReviewStatus = __FORMAL_REVIEW_STATUS__;
const reviewSessionToken = new URLSearchParams(window.location.search).get("token") || "";
const directProjectSave = window.location.protocol === "http:" && window.location.hostname === "127.0.0.1" && Boolean(reviewSessionToken);
const decisionSelects = [...document.querySelectorAll("[data-decision]")];
const anomalySelects = decisionSelects.filter(select => select.dataset.anomaly === "true");
const normalSelects = decisionSelects.filter(select => select.dataset.anomaly === "false");
const summary = document.getElementById("decision-summary");
const toast = document.getElementById("decision-toast");
const exportResult = document.getElementById("export-result");
const generatedStatus = document.getElementById("generated-status");
const saveStatus = document.getElementById("save-status");
const exportDetail = document.getElementById("export-detail");
const saveReviewButton = document.getElementById("save-review-file");
const copyReviewButton = document.getElementById("copy-review-json");
const saveModeNote = document.getElementById("save-mode-note");
const reviewerInput = document.getElementById("reviewed-by");
const exportButton = document.getElementById("export");
const exportReadiness = document.getElementById("export-readiness");
const uncertainGate = document.getElementById("uncertain-gate");
const uncertainConfirm = document.getElementById("confirm-uncertain");
const uncertainConfirmLabel = document.getElementById("uncertain-confirm-label");
const exportScopeConfirm = document.getElementById("confirm-export-scope");
const decisionLabels = {"": "待决策", accepted: "接受", rejected: "拒绝", corrected: "校正坐标", reassigned: "重新归属"};
let toastTimer;
let preparedReviewJson = "";
let reviewRevision = 0;
function showToast(message, kind = "") {
  toast.textContent = message;
  toast.className = `toast show ${kind}`.trim();
  clearTimeout(toastTimer);
  toastTimer = setTimeout(() => { toast.className = "toast"; }, 2600);
}
const assessmentConfirmButton = document.getElementById("confirm-assessment");
const assessmentReviewer = document.getElementById("assessment-reviewed-by");
const assessmentAck = document.getElementById("confirm-assessment-anomalies");
const assessmentConfirmStatus = document.getElementById("assessment-confirm-status");
function updateAssessmentConfirmation() {
  if (!assessmentConfirmButton) return;
  if (formalReviewStatus !== "not_run") {
    assessmentConfirmButton.disabled = true;
    return;
  }
  if (!directProjectSave) {
    assessmentConfirmButton.disabled = true;
    assessmentConfirmStatus.className = "status-warn";
    assessmentConfirmStatus.textContent = "请让 Agent 启动本地复核会话；无需选择保存路径。";
    return;
  }
  const ready = Boolean(assessmentReviewer.value.trim()) && assessmentAck.checked;
  assessmentConfirmButton.disabled = !ready;
  assessmentConfirmStatus.className = ready ? "status-good" : "status-warn";
  assessmentConfirmStatus.textContent = ready ? "可以确认普通候选批次。" : "请填写复核人并勾选批量确认。";
}
if (assessmentConfirmButton) {
  assessmentReviewer.addEventListener("input", updateAssessmentConfirmation);
  assessmentAck.addEventListener("change", updateAssessmentConfirmation);
  assessmentConfirmButton.addEventListener("click", async () => {
    if (assessmentConfirmButton.disabled) return;
    const confirmation = "页面确认：综合评估未发现异常候选，同意批量接受普通候选。";
    assessmentConfirmButton.disabled = true;
    assessmentConfirmStatus.className = "status-warn";
    assessmentConfirmStatus.textContent = "正在生成并保存批量复核记录…";
    try {
      const response = await fetch("/api/review-confirm", {
        method: "POST",
        headers: {"Content-Type": "application/json", "X-Review-Token": reviewSessionToken},
        body: JSON.stringify({reviewed_by: assessmentReviewer.value.trim(), confirmation, assessment_sha256: assessmentSha256})
      });
      const result = await response.json();
      if (!response.ok) throw new Error(result.error || `HTTP ${response.status}`);
      assessmentConfirmStatus.className = "status-good";
      assessmentConfirmStatus.textContent = `普通候选批次已保存：${result.saved_path}`;
      showToast("普通候选批次已确认并保存；尚未应用正式数据。", "accepted");
    } catch (error) {
      assessmentConfirmStatus.className = "status-error";
      assessmentConfirmStatus.textContent = `确认失败：${error && error.message ? error.message : "未知错误"}`;
      assessmentConfirmButton.disabled = false;
      showToast("异常组确认失败，正式状态没有推进。", "rejected");
    }
  });
  updateAssessmentConfirmation();
}
function updateSummary() {
  const anomalyPending = anomalySelects.filter(select => !select.value).length;
  const normalPending = normalSelects.filter(select => !select.value).length;
  const rejected = decisionSelects.filter(select => select.value === "rejected").length;
  const accepted = decisionSelects.filter(select => ["accepted", "corrected", "reassigned"].includes(select.value)).length;
  const corrected = decisionSelects.filter(select => select.value === "corrected").length;
  const reassigned = decisionSelects.filter(select => select.value === "reassigned").length;
  const uncertainAccepted = anomalySelects.filter(select =>
    ["accepted", "corrected", "reassigned"].includes(select.value) && ["ambiguous_shared_colour", "model_assisted_exclusive_assignment"].includes(select.dataset.evidenceStatus)
  ).length;
  summary.textContent = `异常 ${anomalySelects.length} 项（待决策 ${anomalyPending}）· 普通 ${normalSelects.length} 项（待决策 ${normalPending}）· 接受 ${accepted} · 拒绝 ${rejected} · 校正 ${corrected} · 重归属 ${reassigned}`;
  uncertainGate.hidden = uncertainAccepted === 0;
  uncertainConfirmLabel.textContent = `我已人工核对 ${uncertainAccepted} 个模型辅助或同色歧义接受项。`;
  updateExportReadiness();
}
function updateExportReadiness() {
  const anomalyPending = anomalySelects.filter(select => !select.value).length;
  const normalPending = normalSelects.filter(select => !select.value).length;
  const uncertainAccepted = anomalySelects.filter(select =>
    ["accepted", "corrected", "reassigned"].includes(select.value) && ["ambiguous_shared_colour", "model_assisted_exclusive_assignment"].includes(select.dataset.evidenceStatus)
  ).length;
  let message = "可以生成复核 JSON。";
  let ready = true;
  if (formalReviewStatus !== "not_run") {
    message = "当前项目已有正式复核记录；为保护证据链，本页只读。";
    ready = false;
  } else if (!reviewerInput.value.trim()) {
    message = "必须填写复核人或可追溯复核记录。";
    ready = false;
  } else if (anomalyPending > 0) {
    message = `仍有 ${anomalyPending} 个异常候选待独立决策。`;
    ready = false;
  } else if (normalPending > 0) {
    message = `普通候选批量区仍有 ${normalPending} 项待决策。`;
    ready = false;
  } else if (uncertainAccepted > 0 && !uncertainConfirm.checked) {
    message = `必须确认已人工核对 ${uncertainAccepted} 个模型辅助或同色歧义接受项。`;
    ready = false;
  } else if (!exportScopeConfirm.checked) {
    message = "必须确认本次操作只生成复核 JSON，不会自动推进正式状态。";
    ready = false;
  }
  exportButton.disabled = !ready;
  exportButton.setAttribute("aria-disabled", String(!ready));
  exportReadiness.className = `readiness ${ready ? "status-good" : "status-warn"}`;
  exportReadiness.textContent = message;
}
function invalidatePreparedReview() {
  reviewRevision += 1;
  if (!preparedReviewJson) return;
  preparedReviewJson = "";
  exportResult.hidden = true;
  generatedStatus.className = "status-warn";
  generatedStatus.textContent = "内容已变更，请重新生成";
  saveStatus.className = "status-warn";
  saveStatus.textContent = "未保存";
}
function resetApprovalGates() {
  uncertainConfirm.checked = false;
  exportScopeConfirm.checked = false;
}
function syncEditors(select, announce = false) {
  const candidateId = select.dataset.decision;
  const row = select.closest("tr");
  const target = row.querySelector("[data-target-series]");
  const correctedX = row.querySelector("[data-corrected-x]");
  const correctedY = row.querySelector("[data-corrected-y]");
  const state = row.querySelector("[data-decision-state]");
  const value = select.value || "pending";
  const readonly = select.dataset.readonly === "true";
  if (readonly) {
    target.disabled = true;
    correctedX.disabled = true;
    correctedY.disabled = true;
    return;
  }
  target.disabled = select.value !== "reassigned";
  correctedX.disabled = !["corrected", "reassigned"].includes(select.value);
  correctedY.disabled = !["corrected", "reassigned"].includes(select.value);
  row.className = `decision-${value}`;
  state.className = `decision-state state-${value}`;
  state.textContent = decisionLabels[select.value];
  if (announce) showToast(`候选 ${candidateId} 已设置为：${decisionLabels[select.value]}`, value);
}
function setAllDecisions(selects, value, scopeLabel) {
  invalidatePreparedReview();
  resetApprovalGates();
  selects.forEach(select => { select.value = value; syncEditors(select, false); });
  updateSummary();
  showToast(`已将${scopeLabel} ${selects.length} 项设置为：${decisionLabels[value]}`, value);
}
document.getElementById("accept-all-anomalies").addEventListener("click", () => setAllDecisions(anomalySelects, "accepted", "异常候选"));
document.getElementById("reject-all-anomalies").addEventListener("click", () => setAllDecisions(anomalySelects, "rejected", "异常候选"));
document.getElementById("accept-all-normal").addEventListener("click", () => setAllDecisions(normalSelects, "accepted", "普通候选"));
document.getElementById("reject-all-normal").addEventListener("click", () => setAllDecisions(normalSelects, "rejected", "普通候选"));
decisionSelects.forEach(select => select.addEventListener("change", () => {
  invalidatePreparedReview();
  resetApprovalGates();
  syncEditors(select, true);
  updateSummary();
}));
decisionSelects.forEach(select => syncEditors(select, false));
reviewerInput.addEventListener("input", () => { invalidatePreparedReview(); updateExportReadiness(); });
uncertainConfirm.addEventListener("change", updateExportReadiness);
exportScopeConfirm.addEventListener("change", updateExportReadiness);
document.addEventListener("input", event => {
  if (event.target.matches("[data-reason], [data-corrected-x], [data-corrected-y]")) invalidatePreparedReview();
});
document.getElementById("review-apply-agent-prompt").textContent = reviewApplyAgentPrompt;
document.getElementById("pipeline-agent-prompt").textContent = pipelineAgentPrompt;
if (!pipelineAgentPrompt) document.getElementById("pipeline-option").hidden = true;
if (!directProjectSave) {
  saveReviewButton.disabled = true;
  saveReviewButton.textContent = "需由 Agent 启动本地复核会话";
  saveModeNote.className = "compat-note status-warn";
}
async function copyText(text) {
  try {
    if (!navigator.clipboard) throw new Error("clipboard unavailable");
    await navigator.clipboard.writeText(text);
    return true;
  } catch (_) {
    const area = document.createElement("textarea");
    area.value = text;
    area.style.position = "fixed";
    area.style.opacity = "0";
    document.body.appendChild(area);
    area.select();
    const copied = document.execCommand("copy");
    area.remove();
    return copied;
  }
}
document.querySelectorAll("[data-copy-target]").forEach(button => {
  button.addEventListener("click", async () => {
    const copied = await copyText(document.getElementById(button.dataset.copyTarget).textContent);
    showToast(copied ? "Agent 任务语句已复制，请发送给 Codex 等 Agent。" : "复制失败，请手动选择任务语句。", copied ? "accepted" : "rejected");
  });
});
copyReviewButton.addEventListener("click", async () => {
  if (!preparedReviewJson) { showToast("请先生成复核文件内容。", "rejected"); return; }
  const copied = await copyText(preparedReviewJson);
  if (copied) {
    saveStatus.className = "status-warn";
    saveStatus.textContent = "已复制到剪贴板；尚未确认保存为文件";
    showToast("复核 JSON 已复制，仅供排障；项目文件仍未保存。", "accepted");
  } else {
    saveStatus.className = "status-error";
    saveStatus.textContent = "复制失败";
    showToast("复核 JSON 复制失败。", "rejected");
  }
});
saveReviewButton.addEventListener("click", async () => {
  if (!preparedReviewJson) { showToast("请先生成复核文件内容。", "rejected"); return; }
  if (!directProjectSave) {
    saveStatus.className = "status-error";
    saveStatus.textContent = "兼容模式不能写入项目目录";
    showToast("请让 Agent 启动本地复核会话后再保存。", "rejected");
    return;
  }
  try {
    saveReviewButton.disabled = true;
    saveStatus.className = "status-warn";
    saveStatus.textContent = "正在写入当前项目目录";
    const response = await fetch("/api/review-decisions", {
      method: "POST",
      headers: {"Content-Type": "application/json", "X-Review-Token": reviewSessionToken},
      body: preparedReviewJson
    });
    const result = await response.json();
    if (!response.ok) throw new Error(result.error || `HTTP ${response.status}`);
    saveStatus.className = "status-good";
    saveStatus.textContent = `保存成功：${result.saved_path}`;
    exportDetail.textContent = `复核文件已固定保存到：${result.saved_path}；SHA-256：${result.sha256}。现在可以复制 Agent 任务语句继续。`;
    showToast("复核文件已保存到当前项目目录。", "accepted");
  } catch (error) {
    saveStatus.className = "status-error";
    saveStatus.textContent = `保存失败：${error && error.message ? error.message : "未知错误"}`;
    showToast("保存失败，项目状态没有推进。", "rejected");
  } finally {
    saveReviewButton.disabled = false;
  }
});
updateSummary();
exportButton.addEventListener("click", async () => {
  if (exportButton.disabled) return;
  const reviewedBy = reviewerInput.value.trim();
  const decisions = [];
  const generationRevision = reviewRevision;
  exportButton.disabled = true;
  exportResult.hidden = false;
  generatedStatus.className = "status-warn";
  generatedStatus.textContent = `正在生成 0/${decisionSelects.length}`;
  saveStatus.className = "status-warn";
  saveStatus.textContent = "尚未保存";
  exportDetail.textContent = "正在分批校验并生成复核 JSON，请勿关闭页面。";
  for (let index = 0; index < decisionSelects.length; index += 1) {
    const select = decisionSelects[index];
    if (index > 0 && index % 200 === 0) {
      generatedStatus.textContent = `正在生成 ${index}/${decisionSelects.length}`;
      await new Promise(resolve => requestAnimationFrame(resolve));
      if (reviewRevision !== generationRevision) {
        generatedStatus.className = "status-error";
        generatedStatus.textContent = "生成已停止";
        exportDetail.textContent = "生成期间复核内容发生变化，请重新确认门控条件后再次生成。";
        showToast("复核内容已变化，本次生成已停止。", "rejected");
        updateExportReadiness();
        return;
      }
    }
    const candidateId = select.dataset.decision;
    const row = select.closest("tr");
    const reason = row.querySelector("[data-reason]").value.trim();
    const targetSeries = row.querySelector("[data-target-series]").value;
    const correctedX = row.querySelector("[data-corrected-x]").value.trim();
    const correctedY = row.querySelector("[data-corrected-y]").value.trim();
    if (["corrected", "reassigned"].includes(select.value) && !reason) {
      preparedReviewJson = "";
      exportResult.hidden = false;
      generatedStatus.className = "status-error";
      generatedStatus.textContent = "生成失败";
      exportDetail.textContent = `${candidateId} 的校正或重新归属必须填写理由。`;
      row.scrollIntoView({behavior: "smooth", block: "center"});
      row.querySelector("[data-reason]").focus();
      showToast("生成失败：校正或重新归属缺少理由。", "rejected");
      updateExportReadiness();
      return;
    }
    if (select.value === "corrected" && !correctedX && !correctedY) {
      preparedReviewJson = "";
      exportResult.hidden = false;
      generatedStatus.className = "status-error";
      generatedStatus.textContent = "生成失败";
      exportDetail.textContent = `${candidateId} 选择了校正坐标，但没有填写校正值。`;
      row.scrollIntoView({behavior: "smooth", block: "center"});
      row.querySelector("[data-corrected-x]").focus();
      showToast("生成失败：缺少校正坐标。", "rejected");
      updateExportReadiness();
      return;
    }
    if (select.value === "reassigned" && targetSeries === select.dataset.currentSeries) {
      preparedReviewJson = "";
      exportResult.hidden = false;
      generatedStatus.className = "status-error";
      generatedStatus.textContent = "生成失败";
      exportDetail.textContent = `${candidateId} 选择了重新归属，但目标系列没有变化。`;
      row.scrollIntoView({behavior: "smooth", block: "center"});
      row.querySelector("[data-target-series]").focus();
      showToast("生成失败：目标系列没有变化。", "rejected");
      updateExportReadiness();
      return;
    }
    const item = {candidate_id: candidateId, decision: select.value, reason};
    if (select.value === "reassigned") item.target_series = targetSeries;
    if (correctedX) item.corrected_x = Number(correctedX);
    if (correctedY) item.corrected_y = Number(correctedY);
    decisions.push(item);
  }
  const output = {...base, reviewed_by: reviewedBy, reviewed_at: new Date().toISOString(), decisions};
  preparedReviewJson = JSON.stringify(output, null, 2) + "\\n";
  const byteCount = new TextEncoder().encode(preparedReviewJson).length;
  generatedStatus.className = "status-good";
  generatedStatus.textContent = "生成成功";
  saveStatus.className = "status-warn";
  saveStatus.textContent = "尚未保存";
  exportDetail.textContent = `已生成 ${decisions.length} 项决定，共 ${byteCount.toLocaleString()} 字节。请保存到当前项目目录；看到实际保存路径后，再复制任务语句交给 Codex 等 Agent。`;
  exportResult.hidden = false;
  exportResult.scrollIntoView({behavior: "smooth", block: "nearest"});
  showToast("复核 JSON 生成成功；尚未保存，请选择下一步。", "accepted");
  updateExportReadiness();
});
</script>
</body>
</html>
"""
    html = (
        html.replace("__ANOMALY_TABLE_ROWS__", "\n".join(anomaly_table_rows))
        .replace("__NORMAL_TABLE_ROWS__", "\n".join(normal_table_rows))
        .replace("__ANOMALY_COUNT__", str(len(anomaly_table_rows)))
        .replace("__NORMAL_COUNT__", str(len(normal_table_rows)))
        .replace(
            "__ANOMALY_EMPTY__",
            "" if anomaly_table_rows else "<p class='empty-state'>综合评估未发现异常候选。</p>",
        )
        .replace("__REVIEWED_BY__", reviewed_by_existing)
        .replace(
            "__FORM_DISABLED__",
            "disabled" if formal_review_status in {"accepted", "partial", "rejected"} else "",
        )
        .replace("__ASSESSMENT_SUMMARY__", assessment_summary)
        .replace("__ASSESSMENT_SHA256__", json.dumps(assessment_hash))
        .replace("__ASSESSMENT_ANOMALY_COUNT__", str(assessment_anomalies))
        .replace("__FORMAL_REVIEW_STATUS__", json.dumps(formal_review_status))
        .replace("__PAYLOAD__", payload)
        .replace("__REVIEW_APPLY_PROMPT__", json.dumps(review_apply_prompt, ensure_ascii=False))
        .replace("__PIPELINE_PROMPT__", json.dumps(pipeline_prompt, ensure_ascii=False))
    )
    html_path = project_dir / "review.html"
    html_path.write_text(html, encoding="utf-8")
    manifest = load_manifest(project_dir)
    manifest["artifacts"]["review_html"] = artifact_entry(html_path)
    manifest["artifacts"]["review_template"] = artifact_entry(template_path)
    write_json(project_dir / "manifest.json", manifest)
    return {
        "status": "pass",
        "review_status": manifest.get("review_status", "not_run"),
        "候选数量": len(rows),
        "复核页面": str(html_path),
        "复核模板": str(template_path),
        "下一步": (
            "当前已有正式复核记录；页面分别展示异常候选决定与普通候选批次。"
            if formal_review_status in {"accepted", "partial", "rejected"}
            else str(
                acceptance.get(
                    "next_instruction",
                    "请按综合评估判定处理后再继续。",
                )
            )
        ),
    }


def validate_review_payload_for_save(
    project_dir: Path, payload: dict[str, Any]
) -> dict[str, int]:
    """Validate a complete review document without advancing formal workflow state."""
    project_dir = project_dir.expanduser().resolve()
    candidates_path = project_dir / "candidates.csv"
    if not candidates_path.is_file():
        raise FigureError("缺少 candidates.csv，请先运行 extract")
    if payload.get("schema") != REVIEW_SCHEMA:
        raise FigureError(f"复核文件 schema 必须是 {REVIEW_SCHEMA}")
    if payload.get("candidate_sha256") != sha256_file(candidates_path):
        raise FigureError("复核文件绑定的候选数据哈希与当前 candidates.csv 不一致")
    if not str(payload.get("reviewed_by", "")).strip():
        raise FigureError("复核文件必须填写 reviewed_by")

    rows = read_tabular_rows(candidates_path)
    row_by_id = {str(row["candidate_id"]): row for row in rows}
    project_path = project_dir / "project.json"
    project = read_json(project_path) if project_path.is_file() else {}
    valid_series = {
        str(item["id"])
        for item in project.get("chart", {}).get("series", [])
        if isinstance(item, dict) and item.get("id")
    } or {str(row.get("series", "")) for row in rows}
    decisions = payload.get("decisions")
    if not isinstance(decisions, list):
        raise FigureError("复核文件的 decisions 必须是数组")

    decision_by_id: dict[str, dict[str, Any]] = {}
    action_counts = {
        "accepted": 0,
        "rejected": 0,
        "corrected": 0,
        "reassigned": 0,
    }
    for item in decisions:
        if not isinstance(item, dict):
            raise FigureError("每条复核决定必须是对象")
        candidate_id = str(item.get("candidate_id", ""))
        decision = item.get("decision")
        if candidate_id not in row_by_id:
            raise FigureError(f"复核文件包含未知候选编号：{candidate_id}")
        if candidate_id in decision_by_id:
            raise FigureError(f"复核文件重复包含候选编号：{candidate_id}")
        if decision not in action_counts:
            detail = "仍为 pending（待决策）" if decision == "pending" else f"值为 {decision!r}"
            raise FigureError(
                f"{candidate_id} 的 decision {detail}；必须完成 accepted、rejected、corrected 或 reassigned 决策"
            )
        reason = str(item.get("reason", "")).strip()
        corrected_values: dict[str, float] = {}
        for key in ("corrected_x", "corrected_y"):
            if item.get(key) is None or item.get(key) == "":
                continue
            try:
                corrected = float(item[key])
            except (TypeError, ValueError) as exc:
                raise FigureError(f"{candidate_id} 的 {key} 必须是有限数值") from exc
            if not math.isfinite(corrected):
                raise FigureError(f"{candidate_id} 的 {key} 必须是有限数值")
            corrected_values[key] = corrected
        if decision == "corrected" and not corrected_values:
            raise FigureError(f"{candidate_id} 选择 corrected 时至少需要 corrected_x 或 corrected_y")
        if decision == "reassigned":
            target_series = str(item.get("target_series", ""))
            if target_series not in valid_series:
                raise FigureError(f"{candidate_id} 的 target_series 不属于项目系列")
            if target_series == str(row_by_id[candidate_id].get("series", "")):
                raise FigureError(f"{candidate_id} 的 target_series 必须不同于原系列")
        if decision in {"corrected", "reassigned"} and not reason:
            raise FigureError(f"{candidate_id} 的校正或重新归属必须填写 reason")
        if decision in {"accepted", "rejected"} and corrected_values:
            raise FigureError(f"{candidate_id} 包含校正值时 decision 必须是 corrected 或 reassigned")
        decision_by_id[candidate_id] = item
        action_counts[str(decision)] += 1
    missing = sorted(set(row_by_id) - set(decision_by_id))
    if missing:
        raise FigureError(f"复核文件未覆盖全部候选值，缺少 {len(missing)} 项")
    return action_counts


def save_review_decisions_command(
    project_dir: Path, payload: dict[str, Any]
) -> dict[str, Any]:
    """Save to the project's fixed review path without applying the decisions."""
    project_dir = project_dir.expanduser().resolve()
    counts = validate_review_payload_for_save(project_dir, payload)
    target = project_dir / "review-decisions.json"
    manifest = load_manifest(project_dir)
    formal_review_status = manifest.get("review_status", "not_run")
    if formal_review_status in {"accepted", "partial", "rejected"} or (
        project_dir / "data.csv"
    ).is_file():
        raise FigureError(
            "当前项目已有正式应用的复核结果或 data.csv；为保护证据链，不能从复核页面覆盖。"
            "请让 Agent 创建新的项目修订后再复核"
        )
    write_json(target, payload)
    return {
        "status": "pass",
        "save_status": "saved",
        "saved_path": str(target),
        "sha256": sha256_file(target),
        "review_status": formal_review_status,
        "applied": False,
        "counts": counts,
    }


def review_serve_command(project_dir: Path, port: int = 0) -> dict[str, Any]:
    """Serve one token-protected loopback review session with a fixed save target."""
    project_dir = project_dir.expanduser().resolve()
    if not 0 <= port <= 65535:
        raise FigureError("端口必须位于 0 到 65535 之间；0 表示自动选择")
    html_path = project_dir / "review.html"
    overlay_path = project_dir / "overlay.png"
    candidates_path = project_dir / "candidates.csv"
    if not candidates_path.is_file() or not overlay_path.is_file():
        raise FigureError("缺少 candidates.csv 或 overlay.png，请先运行 extract")
    review_command(project_dir)

    token = secrets.token_urlsafe(32)
    maximum_payload_bytes = 16 * 1024 * 1024

    class ReviewHandler(BaseHTTPRequestHandler):
        server_version = "MoreSciFigureReview/0.3.1"

        def send_bytes(self, status: int, body: bytes, content_type: str) -> None:
            self.send_response(status)
            self.send_header("Content-Type", content_type)
            self.send_header("Content-Length", str(len(body)))
            self.send_header("Cache-Control", "no-store")
            self.send_header("X-Content-Type-Options", "nosniff")
            self.send_header("Content-Security-Policy", "default-src 'self'; img-src 'self'; style-src 'unsafe-inline'; script-src 'unsafe-inline'; connect-src 'self'; base-uri 'none'; frame-ancestors 'none'")
            self.end_headers()
            self.wfile.write(body)

        def send_json(self, status: int, payload: dict[str, Any]) -> None:
            body = (json.dumps(payload, ensure_ascii=False) + "\n").encode("utf-8")
            self.send_bytes(status, body, "application/json; charset=utf-8")

        def query_token_is_valid(self) -> bool:
            supplied = parse_qs(urlparse(self.path).query).get("token", [""])[0]
            return bool(supplied) and secrets.compare_digest(supplied, token)

        def header_token_is_valid(self) -> bool:
            supplied = self.headers.get("X-Review-Token", "")
            return bool(supplied) and secrets.compare_digest(supplied, token)

        def do_GET(self) -> None:  # noqa: N802 - stdlib handler contract
            parsed = urlparse(self.path)
            if parsed.path in {"/", "/review.html"}:
                if not self.query_token_is_valid():
                    self.send_json(403, {"status": "failed", "error": "复核会话令牌无效"})
                    return
                page = html_path.read_text(encoding="utf-8").replace(
                    'src="overlay.png"', f'src="overlay.png?token={token}"'
                )
                self.send_bytes(200, page.encode("utf-8"), "text/html; charset=utf-8")
                return
            if parsed.path == "/overlay.png":
                if not self.query_token_is_valid():
                    self.send_json(403, {"status": "failed", "error": "复核会话令牌无效"})
                    return
                self.send_bytes(200, overlay_path.read_bytes(), "image/png")
                return
            self.send_json(404, {"status": "failed", "error": "资源不存在"})

        def do_POST(self) -> None:  # noqa: N802 - stdlib handler contract
            endpoint = urlparse(self.path).path
            if endpoint not in {"/api/review-decisions", "/api/review-confirm"}:
                self.send_json(404, {"status": "failed", "error": "接口不存在"})
                return
            if not self.header_token_is_valid():
                self.send_json(403, {"status": "failed", "error": "复核会话令牌无效"})
                return
            try:
                length = int(self.headers.get("Content-Length", "0"))
            except ValueError:
                self.send_json(400, {"status": "failed", "error": "Content-Length 无效"})
                return
            if length <= 0 or length > maximum_payload_bytes:
                self.send_json(413, {"status": "failed", "error": "复核文件为空或超过 16 MiB"})
                return
            try:
                payload = json.loads(self.rfile.read(length).decode("utf-8"))
                if not isinstance(payload, dict):
                    raise FigureError("复核文件必须是 JSON 对象")
                if endpoint == "/api/review-confirm":
                    assessment_path = project_dir / "review-assessment.json"
                    if not assessment_path.is_file():
                        raise FigureError("缺少 review-assessment.json，请先重新生成综合评估")
                    supplied_assessment_hash = str(payload.get("assessment_sha256", ""))
                    current_assessment_hash = sha256_file(assessment_path)
                    if not secrets.compare_digest(
                        supplied_assessment_hash, current_assessment_hash
                    ):
                        raise FigureError("综合评估哈希已变化，请刷新页面后重新确认")
                    result = review_confirm_command(
                        project_dir,
                        str(payload.get("reviewed_by", "")),
                        str(payload.get("confirmation", "")),
                    )
                else:
                    result = save_review_decisions_command(project_dir, payload)
            except (UnicodeDecodeError, json.JSONDecodeError) as exc:
                self.send_json(400, {"status": "failed", "error": f"复核 JSON 无效：{exc}"})
                return
            except (FigureError, OSError) as exc:
                self.send_json(400, {"status": "failed", "error": str(exc)})
                return
            self.send_json(200, result)

        def log_message(self, format: str, *args: Any) -> None:
            return

    try:
        server = ThreadingHTTPServer(("127.0.0.1", port), ReviewHandler)
    except OSError as exc:
        raise FigureError(f"无法启动本地复核会话：{exc}") from exc
    actual_port = int(server.server_address[1])
    url = f"http://127.0.0.1:{actual_port}/review.html?token={token}"
    serving = {
        "status": "serving",
        "url": url,
        "project_dir": str(project_dir),
        "fixed_save_path": str(project_dir / "review-decisions.json"),
        "说明": "仅监听本机回环地址；用户无需且不能在页面选择保存目录。按 Ctrl-C 停止。",
    }
    print(json.dumps(serving, ensure_ascii=False, indent=2, sort_keys=True), flush=True)
    try:
        server.serve_forever()
    except KeyboardInterrupt:
        pass
    finally:
        server.server_close()
    return {"status": "stopped", "fixed_save_path": serving["fixed_save_path"]}


def build_formal_data_rows(
    project: dict[str, Any], reviewed_observations: list[dict[str, Any]]
) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    """Build style-independent formal data from reviewed pixel observations.

    Pixel visibility gaps remain in ``observations.csv`` and in each formal row's
    ``evidence_segment_break`` field. They become formal curve breaks only when the
    reviewed project explicitly declares ``curve_topology=segmented``.
    """
    chart = project.get("chart", {}) if isinstance(project, dict) else {}
    chart_type = str(chart.get("type", "line")) if isinstance(chart, dict) else "line"
    if chart_type not in {"line", "polar_line"}:
        rows = []
        for index, observation in enumerate(reviewed_observations):
            row = dict(observation)
            row["data_id"] = f"data-{index + 1:06d}"
            row["data_provenance"] = "reviewed_visible_observation"
            rows.append(row)
        return rows, {
            "schema": "more-sci-figure.formal-data-report.v1",
            "chart_type": chart_type,
            "observation_rows": len(reviewed_observations),
            "formal_rows": len(rows),
            "series": [],
            "说明": "非曲线图保持复核观测坐标，不应用曲线连续性规则。",
        }

    series_specs = {
        str(item["id"]): item
        for item in chart.get("series", [])
        if isinstance(item, dict) and item.get("id")
    }
    grouped: dict[str, list[dict[str, Any]]] = {}
    for row in reviewed_observations:
        grouped.setdefault(str(row.get("series", "series")), []).append(row)
    declared_order = [series_id for series_id in series_specs if series_id in grouped]
    series_order = declared_order + [
        series_id for series_id in grouped if series_id not in declared_order
    ]
    formal_rows: list[dict[str, Any]] = []
    series_reports: list[dict[str, Any]] = []
    chart_x_map = AxisMap(chart["x_axis"]) if chart_type == "line" else None
    chart_y_map = AxisMap(chart["y_axis"]) if chart_type == "line" else None

    for series_id in series_order:
        observations = grouped[series_id]
        entry = series_specs.get(series_id, {})
        topology = str(entry.get("curve_topology", "continuous"))
        data_mode = str(entry.get("curve_data_mode", "observations"))
        if topology not in {"continuous", "segmented"}:
            raise FigureError(
                f"系列 {series_id} 的 curve_topology 只支持 continuous 或 segmented"
            )
        if data_mode not in {"observations", "guide_constrained"}:
            raise FigureError(
                f"系列 {series_id} 的 curve_data_mode 只支持 observations 或 guide_constrained"
            )
        ordered = sorted(
            observations,
            key=(
                (lambda row: numeric(row, "x"))
                if chart_type == "polar_line"
                else (
                    lambda row: numeric(row, "pixel_x")
                    if row.get("pixel_x") not in {None, ""}
                    else numeric(row, "x")
                )
            ),
        )
        evidence_segments = max(
            1,
            sum(1 for row in ordered if truthy(row.get("segment_break", False))),
        )
        normalized_visual_breaks = 0

        if data_mode == "guide_constrained":
            if chart_type != "line":
                raise FigureError(
                    f"系列 {series_id} 的 guide_constrained 正式数据目前只支持笛卡尔曲线"
                )
            if topology != "continuous":
                raise FigureError(
                    f"系列 {series_id} 的 guide_constrained 正式数据要求 curve_topology=continuous"
                )
            if len(parse_guide_points(entry)) < 2:
                raise FigureError(
                    f"系列 {series_id} 使用 guide_constrained 正式数据时至少需要两个 guide_points_px"
                )
            curve_x, curve_y, source_count, retained_count, derived_count = (
                guide_constrained_display_geometry(
                    entry,
                    ordered,
                    chart,
                    maximum_residual_px=float(
                        entry.get("curve_data_max_residual_px", 5.0)
                    ),
                    residual_smoothing_window=int(
                        entry.get("curve_data_residual_smoothing_window", 21)
                    ),
                )
            )
            start_x, _ = guide_x_bounds(entry, tuple(int(round(v)) for v in chart["plot_box"]))
            reviewer = str(ordered[0].get("reviewed_by", "")) if ordered else ""
            for index, (x_value, y_value) in enumerate(zip(curve_x, curve_y, strict=True)):
                pixel_x = float(start_x + index)
                pixel_y = float(chart_y_map.pixel(float(y_value))) if chart_y_map else ""
                formal_rows.append(
                    {
                        "series": series_id,
                        "x": float(x_value),
                        "y": float(y_value),
                        "x_uncertainty": (
                            chart_x_map.uncertainty(pixel_x, 0.5) if chart_x_map else ""
                        ),
                        "y_uncertainty": (
                            chart_y_map.uncertainty(pixel_y, 0.5) if chart_y_map else ""
                        ),
                        "pixel_x": pixel_x,
                        "pixel_y": pixel_y,
                        "segment_break": index == 0,
                        "evidence_segment_break": "",
                        "curve_topology": topology,
                        "curve_segment": 0,
                        "curve_order": index,
                        "data_id": f"curve-{series_id}-{index + 1:06d}",
                        "data_provenance": "derived_guide_constrained_curve_data",
                        "topology_provenance": "confirmed_continuous_curve",
                        "source_candidate_count": source_count,
                        "retained_residual_candidate_count": retained_count,
                        "derived_coordinate_count": derived_count,
                        "reviewed_by": reviewer,
                        "review_decision": "derived_from_reviewed_observations",
                    }
                )
            normalized_visual_breaks = max(0, evidence_segments - 1)
            formal_segment_count = 1
        else:
            segment_index = 0
            for index, observation in enumerate(ordered):
                row = dict(observation)
                evidence_break = truthy(row.get("segment_break", False))
                if topology == "continuous":
                    formal_break = index == 0
                    if index > 0 and evidence_break:
                        normalized_visual_breaks += 1
                else:
                    formal_break = index == 0 or evidence_break
                if index > 0 and formal_break:
                    segment_index += 1
                row["evidence_segment_break"] = evidence_break
                row["segment_break"] = formal_break
                row["curve_topology"] = topology
                row["curve_segment"] = segment_index
                row["curve_order"] = index
                row["data_id"] = f"curve-{series_id}-{index + 1:06d}"
                row["data_provenance"] = "reviewed_visible_observation"
                row["topology_provenance"] = (
                    "confirmed_continuous_curve"
                    if topology == "continuous"
                    else "confirmed_segmented_curve"
                )
                formal_rows.append(row)
            formal_segment_count = segment_index + 1 if ordered else 0

        series_reports.append(
            {
                "id": series_id,
                "curve_topology": topology,
                "curve_data_mode": data_mode,
                "observation_rows": len(observations),
                "formal_rows": sum(
                    1 for row in formal_rows if str(row.get("series", "")) == series_id
                ),
                "evidence_segment_count": evidence_segments,
                "formal_segment_count": formal_segment_count,
                "normalized_visual_breaks": normalized_visual_breaks,
            }
        )

    return formal_rows, {
        "schema": "more-sci-figure.formal-data-report.v1",
        "chart_type": chart_type,
        "observation_rows": len(reviewed_observations),
        "formal_rows": len(formal_rows),
        "series": series_reports,
        "说明": (
            "observations.csv 保存可见像素与证据断点；data.csv 保存经确认的曲线拓扑，"
            "视觉空档不再自动成为正式数据断点。"
        ),
    }


def review_apply_command(project_dir: Path, decisions_path: Path) -> dict[str, Any]:
    project_dir = project_dir.expanduser().resolve()
    decisions_path = decisions_path.expanduser().resolve()
    candidates_path = project_dir / "candidates.csv"
    if not candidates_path.is_file():
        raise FigureError("缺少 candidates.csv，请先运行 extract")
    payload = read_json(decisions_path)
    if payload.get("schema") != REVIEW_SCHEMA:
        raise FigureError(f"复核文件 schema 必须是 {REVIEW_SCHEMA}")
    if payload.get("candidate_sha256") != sha256_file(candidates_path):
        raise FigureError("复核文件绑定的候选数据哈希与当前 candidates.csv 不一致")
    reviewed_by = str(payload.get("reviewed_by", "")).strip()
    if not reviewed_by:
        raise FigureError("复核文件必须填写 reviewed_by")
    rows = read_tabular_rows(candidates_path)
    row_by_id = {str(row["candidate_id"]): row for row in rows}
    project_path = project_dir / "project.json"
    project = read_json(project_path) if project_path.is_file() else {}
    valid_series = {
        str(item["id"])
        for item in project.get("chart", {}).get("series", [])
        if isinstance(item, dict) and item.get("id")
    } or {str(row.get("series", "")) for row in rows}
    decisions = payload.get("decisions")
    if not isinstance(decisions, list):
        raise FigureError("复核文件的 decisions 必须是数组")
    decision_by_id: dict[str, dict[str, Any]] = {}
    for item in decisions:
        if not isinstance(item, dict):
            raise FigureError("每条复核决定必须是对象")
        candidate_id = str(item.get("candidate_id", ""))
        decision = item.get("decision")
        if candidate_id not in row_by_id:
            raise FigureError(f"复核文件包含未知候选编号：{candidate_id}")
        if candidate_id in decision_by_id:
            raise FigureError(f"复核文件重复包含候选编号：{candidate_id}")
        if decision not in {"accepted", "rejected", "corrected", "reassigned"}:
            detail = "仍为 pending（待决策）" if decision == "pending" else f"值为 {decision!r}"
            raise FigureError(
                f"{candidate_id} 的 decision {detail}；必须完成 accepted、rejected、corrected 或 reassigned 决策"
            )
        reason = str(item.get("reason", "")).strip()
        corrected_values: dict[str, float] = {}
        for key in ("corrected_x", "corrected_y"):
            if item.get(key) is None or item.get(key) == "":
                continue
            try:
                corrected = float(item[key])
            except (TypeError, ValueError) as exc:
                raise FigureError(f"{candidate_id} 的 {key} 必须是有限数值") from exc
            if not math.isfinite(corrected):
                raise FigureError(f"{candidate_id} 的 {key} 必须是有限数值")
            corrected_values[key] = corrected
        if decision == "corrected" and not corrected_values:
            raise FigureError(f"{candidate_id} 选择 corrected 时至少需要 corrected_x 或 corrected_y")
        if decision == "reassigned":
            target_series = str(item.get("target_series", ""))
            if target_series not in valid_series:
                raise FigureError(f"{candidate_id} 的 target_series 不属于项目系列")
            if target_series == str(row_by_id[candidate_id].get("series", "")):
                raise FigureError(f"{candidate_id} 的 target_series 必须不同于原系列")
        if decision in {"corrected", "reassigned"} and not reason:
            raise FigureError(f"{candidate_id} 的校正或重新归属必须填写 reason")
        if decision in {"accepted", "rejected"} and corrected_values:
            raise FigureError(f"{candidate_id} 包含校正值时 decision 必须是 corrected 或 reassigned")
        decision_by_id[candidate_id] = item
    missing = sorted(set(row_by_id) - set(decision_by_id))
    if missing:
        raise FigureError(f"复核文件未覆盖全部候选值，缺少 {len(missing)} 项")

    accepted_rows: list[dict[str, Any]] = []
    corrected_count = 0
    reassigned_count = 0
    for candidate_id, row in row_by_id.items():
        decision = decision_by_id[candidate_id]
        action = str(decision["decision"])
        if action != "rejected":
            accepted = dict(row)
            if action == "reassigned":
                accepted["original_series"] = accepted.get("series", "")
                accepted["series"] = str(decision["target_series"])
                reassigned_count += 1
            x_key = "x" if "x" in accepted else "x_value" if "x_value" in accepted else "category_index"
            y_key = "y" if "y" in accepted else "value"
            if decision.get("corrected_x") not in {None, ""}:
                accepted[f"original_{x_key}"] = accepted.get(x_key, "")
                accepted[x_key] = float(decision["corrected_x"])
            if decision.get("corrected_y") not in {None, ""}:
                accepted[f"original_{y_key}"] = accepted.get(y_key, "")
                accepted[y_key] = float(decision["corrected_y"])
            if action == "corrected" or decision.get("corrected_x") not in {None, ""} or decision.get("corrected_y") not in {None, ""}:
                corrected_count += 1
            accepted["reviewed_by"] = reviewed_by
            accepted["review_decision"] = action
            accepted["review_reason"] = str(decision.get("reason", ""))
            accepted_rows.append(accepted)
    accepted_count = len(accepted_rows)
    rejected_count = len(rows) - accepted_count
    if accepted_count == 0:
        review_status = "rejected"
    elif rejected_count == 0:
        review_status = "accepted"
    else:
        review_status = "partial"

    normalized = {
        "schema": REVIEW_SCHEMA,
        "candidate_sha256": sha256_file(candidates_path),
        "reviewed_by": reviewed_by,
        "reviewed_at": payload.get("reviewed_at") or datetime.now(timezone.utc).isoformat(),
        "applied_at": datetime.now(timezone.utc).isoformat(),
        "decisions": decisions,
        "summary": {
            "total": len(rows),
            "accepted": accepted_count,
            "rejected": rejected_count,
            "corrected": corrected_count,
            "reassigned": reassigned_count,
            "review_status": review_status,
        },
    }
    for key in (
        "project_id",
        "source_sha256",
        "review_method",
        "assessment_sha256",
        "assessment_score",
        "assessment_risk_level",
        "conversation_confirmation",
        "anomaly_acknowledgement",
        "anomaly_review",
    ):
        if payload.get(key) is not None:
            normalized[key] = payload[key]
    applied_path = project_dir / "review-decisions.json"
    write_json(applied_path, normalized)
    observations_path = project_dir / "observations.csv"
    write_csv(observations_path, accepted_rows)
    formal_rows, formal_data_report = build_formal_data_rows(project, accepted_rows)
    data_path = project_dir / "data.csv"
    write_csv(data_path, formal_rows)
    formal_data_report.update(
        {
            "project_spec": artifact_entry(project_path),
            "candidates": artifact_entry(candidates_path),
            "review_decisions": artifact_entry(applied_path),
            "observations": artifact_entry(observations_path),
            "data": artifact_entry(data_path),
        }
    )
    formal_data_report_path = project_dir / "formal-data-report.json"
    write_json(formal_data_report_path, formal_data_report)

    report_path = project_dir / "extraction-report.json"
    report = read_json(report_path)
    report["review_status"] = review_status
    report["numeric_output_authorized"] = accepted_count > 0
    report["review_summary"] = normalized["summary"]
    report["formal_data"] = {
        "observations": artifact_entry(observations_path),
        "data": artifact_entry(data_path),
        "report": artifact_entry(formal_data_report_path),
        "rows": len(formal_rows),
    }
    write_json(report_path, report)

    manifest = load_manifest(project_dir)
    manifest["review_status"] = review_status
    manifest["artifacts"]["review_decisions"] = artifact_entry(applied_path)
    manifest["artifacts"]["observations"] = artifact_entry(observations_path)
    manifest["artifacts"]["data"] = artifact_entry(data_path)
    manifest["artifacts"]["formal_data_report"] = artifact_entry(
        formal_data_report_path
    )
    manifest["artifacts"]["extraction_report"] = artifact_entry(report_path)
    write_json(project_dir / "manifest.json", manifest)
    return {
        "status": "pass" if accepted_count else "failed",
        "review_status": review_status,
        "accepted": accepted_count,
        "rejected": rejected_count,
        "corrected": corrected_count,
        "reassigned": reassigned_count,
        "observations": str(observations_path),
        "data": str(data_path),
        "formal_data_report": str(formal_data_report_path),
    }


def read_tabular_rows(path: Path) -> list[dict[str, Any]]:
    suffix = path.suffix.lower()
    if suffix in {".csv", ".tsv"}:
        with path.open("r", encoding="utf-8-sig", newline="") as handle:
            return list(csv.DictReader(handle, delimiter="\t" if suffix == ".tsv" else ","))
    if suffix == ".json":
        payload = read_json(path)
        rows = payload.get("rows") if isinstance(payload, dict) else payload
        if not isinstance(rows, list) or any(not isinstance(row, dict) for row in rows):
            raise FigureError("JSON 数据必须是对象列表，或包含 rows 列表的对象")
        return rows
    if suffix in {".xlsx", ".xlsm"}:
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise FigureError("读取 Excel 需要安装 openpyxl") from exc
        workbook = load_workbook(path, read_only=True, data_only=True)
        sheet = workbook.active
        values = list(sheet.iter_rows(values_only=True))
        workbook.close()
        if not values:
            return []
        headers = [str(value) for value in values[0]]
        return [dict(zip(headers, row, strict=False)) for row in values[1:]]
    raise FigureError(f"不支持的数据格式：{suffix}")


def numeric(row: dict[str, Any], key: str) -> float:
    try:
        value = float(row[key])
        if not math.isfinite(value):
            raise ValueError
        return value
    except (KeyError, TypeError, ValueError) as exc:
        raise FigureError(f"列 {key!r} 必须包含有限数值") from exc


def truthy(value: Any) -> bool:
    return str(value).strip().lower() in {"1", "true", "yes", "y"}


def series_colors(spec: dict[str, Any]) -> dict[str, str]:
    return {str(item["id"]): str(item["color"]) for item in spec.get("chart", {}).get("series", [])}


def smooth_curve_points(
    x: np.ndarray,
    y: np.ndarray,
    *,
    samples_per_interval: int = 4,
    smoothing_window: int = 7,
) -> tuple[np.ndarray, np.ndarray]:
    """生成不越界的形状保持显示几何；只用于重绘，不回写观测数据。"""
    x_array = np.asarray(x, dtype=float)
    y_array = np.asarray(y, dtype=float)
    if x_array.size < 2 or x_array.size != y_array.size or np.any(np.diff(x_array) <= 0):
        return x_array, y_array
    window = max(1, int(smoothing_window))
    if window > 1:
        if window % 2 == 0:
            window += 1
        window = min(window, x_array.size if x_array.size % 2 == 1 else x_array.size - 1)
    if window > 1:
        radius = window // 2
        padded = np.pad(y_array, (radius, radius), mode="reflect")
        filtered = np.convolve(padded, np.full(window, 1.0 / window), mode="valid")
        filtered[0], filtered[-1] = y_array[0], y_array[-1]
    else:
        filtered = y_array.copy()
    widths = np.diff(x_array)
    deltas = np.diff(filtered) / widths
    slopes = np.zeros_like(filtered)
    if filtered.size == 2:
        slopes[:] = deltas[0]
    else:
        for index in range(1, filtered.size - 1):
            if deltas[index - 1] * deltas[index] <= 0:
                slopes[index] = 0.0
            else:
                left_weight = 2.0 * widths[index] + widths[index - 1]
                right_weight = widths[index] + 2.0 * widths[index - 1]
                slopes[index] = (left_weight + right_weight) / (
                    left_weight / deltas[index - 1] + right_weight / deltas[index]
                )
        slopes[0], slopes[-1] = deltas[0], deltas[-1]
    samples = max(1, int(samples_per_interval))
    dense_x_parts: list[np.ndarray] = []
    dense_y_parts: list[np.ndarray] = []
    local = np.linspace(0.0, 1.0, samples, endpoint=False)
    for index, width in enumerate(widths):
        t, t2, t3 = local, local * local, local * local * local
        dense_x_parts.append(x_array[index] + width * t)
        dense_y_parts.append(
            (2 * t3 - 3 * t2 + 1) * filtered[index]
            + (t3 - 2 * t2 + t) * width * slopes[index]
            + (-2 * t3 + 3 * t2) * filtered[index + 1]
            + (t3 - t2) * width * slopes[index + 1]
        )
    return (
        np.concatenate([*dense_x_parts, x_array[-1:]]),
        np.concatenate([*dense_y_parts, filtered[-1:]]),
    )


def display_segments(
    group_rows: list[dict[str, Any]],
    x_key: str,
    y_key: str,
    *,
    mode: str,
    smoothing_window: int,
    samples_per_interval: int,
    outlier_window: int = 1,
    max_outlier_pixel_residual: float | None = None,
    knot_stride: int = 1,
    sort_by_data_x: bool = False,
) -> list[tuple[np.ndarray, np.ndarray, int, int, int]]:
    ordered = sorted(
        group_rows,
        key=(
            lambda row: numeric(row, "curve_order")
            if row.get("curve_order") not in {None, ""}
            else numeric(row, x_key)
            if sort_by_data_x
            else numeric(row, "pixel_x")
            if row.get("pixel_x") not in {None, ""}
            else numeric(row, x_key)
        ),
    )
    raw_segments: list[list[dict[str, Any]]] = []
    current: list[dict[str, Any]] = []
    for row in ordered:
        if current and truthy(row.get("segment_break", False)):
            raw_segments.append(current)
            current = []
        current.append(row)
    if current:
        raw_segments.append(current)
    result: list[tuple[np.ndarray, np.ndarray, int, int, int]] = []
    for segment in raw_segments:
        source_count = len(segment)
        filtered_segment = list(segment)
        window = max(1, int(outlier_window))
        if (
            max_outlier_pixel_residual is not None
            and window > 1
            and len(filtered_segment) >= 3
            and all(row.get("pixel_y") not in {None, ""} for row in filtered_segment)
        ):
            if window % 2 == 0:
                window += 1
            window = min(
                window,
                len(filtered_segment)
                if len(filtered_segment) % 2 == 1
                else len(filtered_segment) - 1,
            )
            pixel_y = np.asarray([numeric(row, "pixel_y") for row in filtered_segment])
            radius = window // 2
            padded = np.pad(pixel_y, (radius, radius), mode="edge")
            baseline = np.asarray(
                [np.median(padded[index : index + window]) for index in range(len(pixel_y))]
            )
            keep = np.abs(pixel_y - baseline) <= float(max_outlier_pixel_residual)
            if int(np.count_nonzero(keep)) >= 2:
                filtered_segment = [
                    row for row, retained in zip(filtered_segment, keep, strict=True) if retained
                ]
        stride = max(1, int(knot_stride))
        if stride > 1 and len(filtered_segment) > 2:
            indices = list(range(0, len(filtered_segment), stride))
            if indices[-1] != len(filtered_segment) - 1:
                indices.append(len(filtered_segment) - 1)
            filtered_segment = [filtered_segment[index] for index in indices]
        retained_count = len(filtered_segment)
        xs = np.asarray([numeric(row, x_key) for row in filtered_segment], dtype=float)
        ys = np.asarray([numeric(row, y_key) for row in filtered_segment], dtype=float)
        unique_x, indices = np.unique(xs, return_index=True)
        xs, ys = unique_x, ys[indices]
        if mode == "shape_preserving" and xs.size >= 2:
            dense_x, dense_y = smooth_curve_points(
                xs,
                ys,
                samples_per_interval=samples_per_interval,
                smoothing_window=smoothing_window,
            )
        else:
            dense_x, dense_y = xs, ys
        result.append(
            (
                dense_x,
                dense_y,
                source_count,
                retained_count,
                max(0, len(dense_x) - retained_count),
            )
        )
    return result


def guide_constrained_display_geometry(
    entry: dict[str, Any],
    group_rows: list[dict[str, Any]],
    chart: dict[str, Any],
    *,
    maximum_residual_px: float,
    residual_smoothing_window: int,
) -> tuple[np.ndarray, np.ndarray, int, int, int]:
    """用已声明引导路径生成派生几何，并仅以可见像素残差作平滑校正。"""
    plot_box = tuple(int(round(value)) for value in chart["plot_box"])
    start_x, end_x = guide_x_bounds(entry, plot_box)
    pixel_x = np.arange(start_x, end_x + 1, dtype=float)
    guide_y = np.asarray([float(guide_y_at(entry, value)) for value in pixel_x])
    supported = [
        row
        for row in group_rows
        if row.get("pixel_x") not in {None, ""} and row.get("pixel_y") not in {None, ""}
    ]
    residual_x: list[float] = []
    residual_y: list[float] = []
    for row in supported:
        x_value = numeric(row, "pixel_x")
        expected = guide_y_at(entry, x_value)
        if expected is None:
            continue
        residual = numeric(row, "pixel_y") - expected
        if abs(residual) <= maximum_residual_px:
            residual_x.append(x_value)
            residual_y.append(residual)
    retained = len(residual_x)
    if retained:
        x_array = np.asarray(residual_x, dtype=float)
        y_array = np.asarray(residual_y, dtype=float)
        unique_x = np.unique(x_array)
        unique_residual = np.asarray(
            [float(np.median(y_array[x_array == value])) for value in unique_x]
        )
        correction = np.interp(pixel_x, unique_x, unique_residual)
        window = max(1, int(residual_smoothing_window))
        if window > 1:
            if window % 2 == 0:
                window += 1
            window = min(window, len(correction) if len(correction) % 2 == 1 else len(correction) - 1)
            if window > 1:
                radius = window // 2
                padded = np.pad(correction, (radius, radius), mode="edge")
                correction = np.convolve(padded, np.full(window, 1.0 / window), mode="valid")
        display_pixel_y = guide_y + correction
    else:
        display_pixel_y = guide_y
    x_map = AxisMap(chart["x_axis"])
    y_map = AxisMap(chart["y_axis"])
    display_x = np.asarray([x_map.value(value) for value in pixel_x])
    display_y = np.asarray([y_map.value(value) for value in display_pixel_y])
    source_count = len(group_rows)
    return display_x, display_y, source_count, retained, len(display_x)


def render_command(
    spec_path: Path,
    data_path: Path,
    out_dir: Path,
    *,
    artifact_basename: str = "render",
    update_manifest: bool = True,
    input_status: str = "accepted_or_supplied",
) -> dict[str, Any]:
    cache_root = Path(os.environ.get("TMPDIR", "/tmp")) / "more-sci-figure-mpl"
    cache_root.mkdir(parents=True, exist_ok=True)
    os.environ.setdefault("MPLCONFIGDIR", str(cache_root))
    import matplotlib

    matplotlib.use("Agg")
    matplotlib.rcParams["font.sans-serif"] = [
        "PingFang SC",
        "Hiragino Sans GB",
        "Microsoft YaHei",
        "Noto Sans CJK SC",
        "Arial Unicode MS",
        "DejaVu Sans",
    ]
    matplotlib.rcParams["axes.unicode_minus"] = False
    import matplotlib.pyplot as plt

    spec_path = spec_path.expanduser().resolve()
    data_path = data_path.expanduser().resolve()
    out_dir.mkdir(parents=True, exist_ok=True)
    spec = read_json(spec_path)
    errors = validate_spec(spec, spec_path, extraction=False)
    if errors:
        raise FigureError("；".join(errors))
    rows = read_tabular_rows(data_path)
    if not rows:
        raise FigureError("没有可用于重绘的数据行")
    if update_manifest and any(
        row.get("candidate_id")
        and row.get("review_decision") not in {"accepted", "corrected", "reassigned"}
        and str(row.get("status", "")).lower() in {"visible", "visible_candidate", "candidate"}
        for row in rows
    ):
        raise FigureError("正式 render 拒绝未复核候选；请先 review-apply，或使用 preview")
    render = spec.get("render", {})
    plot_type = render.get("plot_type") or spec.get("chart", {}).get("type")
    if plot_type not in SUPPORTED_CHARTS:
        raise FigureError(f"不支持的重绘类型：{plot_type}")
    formal_curve_data = (
        plot_type in {"line", "polar_line"}
        and any(row.get("data_id") not in {None, ""} for row in rows)
    )
    if formal_curve_data and any(
        row.get("curve_topology") not in {"continuous", "segmented"}
        or row.get("data_provenance") in {None, ""}
        for row in rows
    ):
        raise FigureError(
            "正式曲线 data.csv 必须包含 curve_topology 与 data_provenance；"
            "请重新运行 review-apply 构建样式无关曲线数据"
        )
    polar_plot = plot_type == "polar_line"
    x_key = str(
        render.get(
            "x",
            "x" if plot_type in {"line", "polar_line", "scatter"} else "category_index",
        )
    )
    y_key = str(
        render.get("y", "y" if plot_type in {"line", "polar_line", "scatter"} else "value")
    )
    group_key = render.get("group")
    groups: dict[str, list[dict[str, Any]]] = {}
    for row in rows:
        group = str(row.get(str(group_key), "series")) if group_key else "series"
        groups.setdefault(group, []).append(row)
    colors = series_colors(spec)
    canvas = render.get("canvas_px")
    output_dpi = int(render.get("dpi", 160))
    if isinstance(canvas, list) and len(canvas) == 2:
        canvas_width, canvas_height = int(canvas[0]), int(canvas[1])
        fig = plt.figure(figsize=(canvas_width / output_dpi, canvas_height / output_dpi), dpi=output_dpi)
        axes_box = render.get("axes_box_px")
        if isinstance(axes_box, list) and len(axes_box) == 4:
            left_px, top_px, right_px, bottom_px = [float(value) for value in axes_box]
            axis = fig.add_axes(
                [
                    left_px / canvas_width,
                    (canvas_height - bottom_px) / canvas_height,
                    (right_px - left_px) / canvas_width,
                    (bottom_px - top_px) / canvas_height,
                ],
                projection="polar" if polar_plot else None,
            )
        else:
            axis = fig.add_subplot(111, projection="polar" if polar_plot else None)
    else:
        canvas_width = canvas_height = None
        fig, axis = plt.subplots(
            figsize=(6.4, 6.0) if polar_plot else (7.2, 4.8),
            dpi=output_dpi,
            constrained_layout=True,
            subplot_kw={"projection": "polar"} if polar_plot else None,
        )
    fallback = ["#0072B2", "#D55E00", "#009E73", "#CC79A7", "#E69F00", "#56B4E9"]
    rendered_segments: dict[str, int] = {}
    geometry_sources: dict[str, str] = {}
    display_geometry_rows: list[dict[str, Any]] = []
    display_policy = render.get("display_geometry", {}) if isinstance(render.get("display_geometry", {}), dict) else {}
    display_mode = str(display_policy.get("mode", "none"))
    smoothing_window = int(display_policy.get("smoothing_window", 7))
    samples_per_interval = int(display_policy.get("samples_per_interval", 4))
    outlier_window = int(display_policy.get("outlier_window", 1))
    maximum_outlier_residual = display_policy.get("max_outlier_pixel_residual")
    knot_stride = int(display_policy.get("knot_stride", 1))
    styles = render.get("series_styles", {}) if isinstance(render.get("series_styles", {}), dict) else {}
    chart = spec.get("chart", {})
    series_specs = {
        str(item["id"]): item
        for item in chart.get("series", [])
        if isinstance(item, dict) and item.get("id")
    }
    if plot_type in {"line", "polar_line", "scatter"}:
        for index, (group, group_rows) in enumerate(groups.items()):
            style = styles.get(group, {}) if isinstance(styles.get(group, {}), dict) else {}
            color = str(style.get("color", colors.get(group, fallback[index % len(fallback)])))
            label = str(style.get("label", group))
            if plot_type in {"line", "polar_line"}:
                geometry_source = str(style.get("geometry_source", "observations"))
                if formal_curve_data and geometry_source == "guide_constrained":
                    raise FigureError(
                        f"系列 {group} 不得在 render 阶段使用 guide_constrained；"
                        "请在 chart.series.curve_data_mode 中构建正式曲线数据"
                    )
                if (
                    plot_type == "line"
                    and geometry_source == "guide_constrained"
                    and group in series_specs
                ):
                    if len(parse_guide_points(series_specs[group])) < 2:
                        raise FigureError(
                            f"系列 {group} 使用 guide_constrained 时至少需要两个 guide_points_px"
                        )
                    segments = [
                        guide_constrained_display_geometry(
                            series_specs[group],
                            group_rows,
                            chart,
                            maximum_residual_px=float(
                                style.get("guide_constraint_max_residual_px", 5.0)
                            ),
                            residual_smoothing_window=int(
                                style.get("guide_residual_smoothing_window", 21)
                            ),
                        )
                    ]
                    geometry_provenance = "derived_guide_constrained_geometry"
                else:
                    segments = display_segments(
                        group_rows,
                        x_key,
                        y_key,
                        mode=display_mode,
                        smoothing_window=smoothing_window,
                        samples_per_interval=samples_per_interval,
                        outlier_window=outlier_window,
                        max_outlier_pixel_residual=(
                            float(maximum_outlier_residual)
                            if maximum_outlier_residual is not None
                            else None
                        ),
                        knot_stride=knot_stride,
                        sort_by_data_x=polar_plot,
                    )
                    group_data_provenance = {
                        str(row.get("data_provenance", "")) for row in group_rows
                    }
                    if display_mode == "shape_preserving":
                        geometry_provenance = "derived_display_geometry"
                    elif "derived_guide_constrained_curve_data" in group_data_provenance:
                        geometry_provenance = "formal_derived_curve_data_geometry"
                    elif formal_curve_data:
                        geometry_provenance = "formal_curve_data_geometry"
                    else:
                        geometry_provenance = "accepted_observation_geometry"
                geometry_sources[group] = geometry_provenance
                rendered_segments[group] = len(segments)
                for segment_index, (
                    segment_x,
                    segment_y,
                    observed_count,
                    retained_count,
                    derived_count,
                ) in enumerate(segments):
                    if (
                        polar_plot
                        and len(segment_x) > 1
                        and float(segment_x[-1]) - float(segment_x[0]) >= 300.0
                    ):
                        segment_x = np.append(segment_x, float(segment_x[0]) + 360.0)
                        segment_y = np.append(segment_y, float(segment_y[0]))
                    plot_x = np.deg2rad(segment_x) if polar_plot else segment_x
                    axis.plot(
                        plot_x,
                        segment_y,
                        color=color,
                        linewidth=float(style.get("linewidth", 1.8)),
                        linestyle=str(style.get("line_style", "-")),
                        label=label if segment_index == 0 else None,
                    )
                    for x_value, y_value in zip(segment_x, segment_y):
                        display_geometry_rows.append(
                            {
                                "series": group,
                                "segment": segment_index,
                                "x": float(x_value),
                                "y": float(y_value),
                                "provenance": geometry_provenance,
                                "observed_anchor_count": observed_count,
                                "display_anchor_count": retained_count,
                                "discarded_or_thinned_count": observed_count - retained_count,
                                "derived_point_count": derived_count,
                            }
                        )
                marker = style.get("marker")
                if marker:
                    marker_every = max(1, int(style.get("marker_every", max(1, len(group_rows) // 16))))
                    marker_rows = sorted(group_rows, key=lambda row: numeric(row, x_key))[::marker_every]
                    marker_x = [numeric(row, x_key) for row in marker_rows]
                    axis.plot(
                        np.deg2rad(marker_x) if polar_plot else marker_x,
                        [numeric(row, y_key) for row in marker_rows],
                        linestyle="none",
                        marker=str(marker),
                        markersize=float(style.get("marker_size", 5.0)),
                        markerfacecolor=str(style.get("marker_facecolor", "white")),
                        markeredgecolor=color,
                        markeredgewidth=float(style.get("marker_edgewidth", 1.0)),
                    )
            else:
                xs = [numeric(row, x_key) for row in group_rows]
                ys = [numeric(row, y_key) for row in group_rows]
                axis.scatter(xs, ys, color=color, s=24, label=label)
    else:
        category_values = sorted({numeric(row, x_key) for row in rows})
        group_count = len(groups)
        width = 0.8 / max(1, group_count)
        for index, (group, group_rows) in enumerate(groups.items()):
            mapping = {numeric(row, x_key): numeric(row, y_key) for row in group_rows}
            positions = np.asarray(category_values) - 0.4 + width / 2 + index * width
            heights = [mapping.get(category, np.nan) for category in category_values]
            axis.bar(
                positions,
                heights,
                width=width,
                color=colors.get(group, fallback[index % len(fallback)]),
                label=group,
            )
        axis.set_xticks(category_values)
    x_scale = str(render.get("x_scale") or chart.get("x_axis", {}).get("scale", "linear"))
    y_scale = str(
        render.get("y_scale")
        or (
            chart.get("polar", {}).get("radius_axis", {}).get("scale", "linear")
            if polar_plot
            else chart.get("y_axis", {}).get("scale", "linear")
        )
    )
    if polar_plot:
        polar = chart.get("polar", {})
        angle_axis = polar.get("angle_axis", {})
        axis.set_theta_offset(math.radians(-float(angle_axis.get("zero_bearing_deg", 0.0))))
        axis.set_theta_direction(
            -1 if angle_axis.get("direction", "clockwise") == "clockwise" else 1
        )
        x_scale = "circular_degree"
        axis.set_yscale("log" if y_scale == "log10" else "linear")
    else:
        axis.set_xscale("log" if x_scale == "log10" else "linear")
        axis.set_yscale("log" if y_scale == "log10" else "linear")
        if isinstance(render.get("x_limits"), list) and len(render["x_limits"]) == 2:
            axis.set_xlim(float(render["x_limits"][0]), float(render["x_limits"][1]))
    if isinstance(render.get("y_limits"), list) and len(render["y_limits"]) == 2:
        axis.set_ylim(float(render["y_limits"][0]), float(render["y_limits"][1]))
    if isinstance(render.get("x_ticks"), list):
        x_ticks = [float(value) for value in render["x_ticks"]]
        axis.set_xticks(np.deg2rad(x_ticks) if polar_plot else x_ticks)
    if isinstance(render.get("y_ticks"), list):
        axis.set_yticks([float(value) for value in render["y_ticks"]])
    if not polar_plot:
        axis.set_xlabel(str(render.get("x_label", "")), fontsize=float(render.get("label_fontsize", 12)))
        axis.set_ylabel(str(render.get("y_label", "")), fontsize=float(render.get("label_fontsize", 12)))
    elif render.get("radial_label_position_deg") is not None:
        axis.set_rlabel_position(float(render["radial_label_position_deg"]))
    axis.set_title(str(render.get("title", "")))
    if not polar_plot and not bool(render.get("boxed_axes", False)):
        axis.spines["top"].set_visible(False)
        axis.spines["right"].set_visible(False)
    grid = render.get("grid", "y")
    if grid in {"x", "y", "both"}:
        axis.grid(axis=grid, color="#d9d9d9", linewidth=0.6, alpha=0.6)
    else:
        axis.grid(False)
    axis.tick_params(
        direction=str(render.get("tick_direction", "out")),
        labelsize=float(render.get("tick_fontsize", 10)),
    )
    if len(groups) > 1 or (group_key and next(iter(groups)) != "series"):
        legend = render.get("legend", {}) if isinstance(render.get("legend", {}), dict) else {}
        axis.legend(
            frameon=bool(legend.get("frameon", False)),
            ncol=int(legend.get("ncol", 1)),
            loc=str(legend.get("loc", "best")),
            bbox_to_anchor=tuple(legend["bbox_to_anchor"]) if isinstance(legend.get("bbox_to_anchor"), list) else None,
            fontsize=float(legend.get("fontsize", 10)),
        )
    if not update_manifest:
        fig.text(
            0.995,
            0.006,
            "CANDIDATE PREVIEW · NOT REVIEWED",
            ha="right",
            va="bottom",
            fontsize=6,
            color="#8b1a1a",
            alpha=0.8,
        )
    outputs: dict[str, dict[str, Any]] = {}
    for suffix in ("png", "svg", "pdf"):
        path = out_dir / f"{artifact_basename}.{suffix}"
        fig.savefig(path, facecolor="white", dpi=output_dpi)
        outputs[suffix] = artifact_entry(path)
    display_geometry_path: Path | None = None
    if display_geometry_rows:
        geometry_name = (
            "display-geometry.csv"
            if artifact_basename == "render"
            else f"{artifact_basename}-display-geometry.csv"
        )
        display_geometry_path = out_dir / geometry_name
        write_csv(display_geometry_path, display_geometry_rows)
    plt.close(fig)
    report_name = (
        "render-report.json"
        if artifact_basename == "render"
        else f"{artifact_basename}-report.json"
    )
    report = {
        "schema": "more-sci-figure.render-report.v1",
        "status": "pass",
        "formal_render": update_manifest,
        "input_status": input_status,
        "plot_type": plot_type,
        "rows": len(rows),
        "mapping": {"x": x_key, "y": y_key, "group": group_key},
        "axis_scales": {"x": x_scale, "y": y_scale},
        "rendered_segments": rendered_segments,
        "geometry_sources": geometry_sources,
        "formal_curve_data": formal_curve_data,
        "data_provenance": sorted(
            {
                str(row.get("data_provenance", ""))
                for row in rows
                if row.get("data_provenance") not in {None, ""}
            }
        ),
        "display_geometry": {
            "mode": display_mode,
            "rows": len(display_geometry_rows),
            "path": str(display_geometry_path) if display_geometry_path else None,
            "provenance": "derived display geometry never overwrites data.csv",
        },
        "project_spec": artifact_entry(spec_path),
        "canvas_px": [canvas_width, canvas_height] if canvas_width and canvas_height else None,
        "outputs": outputs,
    }
    write_json(out_dir / report_name, report)
    if not update_manifest:
        return report
    root = out_dir.parent if out_dir.name == "render" else out_dir
    manifest = load_manifest(root)
    if manifest.get("extraction_status") == "not_run":
        canonical_data = root / "data.csv"
        write_csv(canonical_data, rows)
        project_copy = root / "project.json"
        if project_copy.resolve() != spec_path:
            write_json(project_copy, spec)
        manifest["extraction_status"] = "not_applicable"
        manifest["review_status"] = "not_applicable"
        manifest["data_origin"] = "supplied"
        manifest["artifacts"]["supplied_data"] = artifact_entry(data_path)
        manifest["artifacts"]["data"] = artifact_entry(canonical_data)
        manifest["project_spec"] = artifact_entry(project_copy if project_copy.exists() else spec_path)
        manifest["artifacts"]["project_spec"] = artifact_entry(
            project_copy if project_copy.exists() else spec_path
        )
    manifest["render_status"] = "pass"
    manifest["tool_version"] = VERSION
    previous_project_spec = manifest.get("artifacts", {}).get("project_spec")
    if previous_project_spec and previous_project_spec.get("sha256") != sha256_file(spec_path):
        manifest["artifacts"]["extraction_project_spec"] = previous_project_spec
    manifest["project_spec"] = artifact_entry(spec_path)
    manifest["artifacts"]["project_spec"] = artifact_entry(spec_path)
    manifest["artifacts"]["render_report"] = artifact_entry(out_dir / report_name)
    if display_geometry_path is not None:
        manifest["artifacts"]["display_geometry"] = artifact_entry(display_geometry_path)
    for suffix, entry in outputs.items():
        manifest["artifacts"][f"render_{suffix}"] = entry
    write_json(root / "manifest.json", manifest)
    return report


def preview_command(spec_path: Path, candidates_path: Path, out_dir: Path) -> dict[str, Any]:
    """渲染带水印的候选预览；不生成 data.csv，也不推进正式交付状态。"""
    rows = read_tabular_rows(candidates_path.expanduser().resolve())
    if not rows or any("candidate_id" not in row for row in rows):
        raise FigureError("preview 只接受 extract 产生且包含 candidate_id 的候选表")
    return render_command(
        spec_path,
        candidates_path,
        out_dir,
        artifact_basename="candidate-preview",
        update_manifest=False,
        input_status="unreviewed_candidates",
    )


def validate_command(project_dir: Path, reference: Path | None = None) -> dict[str, Any]:
    project_dir = project_dir.expanduser().resolve()
    manifest_path = project_dir / "manifest.json"
    errors: list[str] = []
    warnings: list[str] = []
    if not manifest_path.exists():
        raise FigureError(f"找不到 manifest.json：{manifest_path}")
    manifest = read_json(manifest_path)
    recorded_project_path = str(
        manifest.get("artifacts", {}).get("project_spec", {}).get("path", "")
    )
    project_path = project_dir / "project.json"
    if recorded_project_path:
        recorded_candidate = Path(recorded_project_path).expanduser()
        if recorded_candidate.is_absolute() and recorded_candidate.is_file():
            project_path = recorded_candidate.resolve()
        elif recorded_candidate.is_file():
            project_path = recorded_candidate.resolve()
        elif (project_dir / recorded_candidate).is_file():
            project_path = (project_dir / recorded_candidate).resolve()
    required = {
        "project_spec": project_path,
        "data": project_dir / "data.csv",
        "render_report": project_dir / "render" / "render-report.json",
        "render_png": project_dir / "render" / "render.png",
        "render_svg": project_dir / "render" / "render.svg",
        "render_pdf": project_dir / "render" / "render.pdf",
    }
    if manifest.get("extraction_status") != "not_applicable":
        required.update(
            {
                "candidates": project_dir / "candidates.csv",
                "observations": project_dir / "observations.csv",
                "formal_data_report": project_dir / "formal-data-report.json",
                "overlay": project_dir / "overlay.png",
                "extraction_report": project_dir / "extraction-report.json",
                "review_decisions": project_dir / "review-decisions.json",
            }
        )
    project_path = required["project_spec"]
    project: dict[str, Any] = {}
    if project_path.is_file():
        project = read_json(project_path)
        project_errors = validate_spec(project, project_path, extraction=False)
        errors.extend(f"项目规格验证失败：{error}" for error in project_errors)
    artifact_checks: dict[str, Any] = {}
    for name, path in required.items():
        exists = path.is_file()
        artifact_checks[name] = {"path": str(path), "exists": exists}
        if not exists:
            errors.append(f"缺少交付物：{name}")
            continue
        recorded = manifest.get("artifacts", {}).get(name, {}).get("sha256")
        current = sha256_file(path)
        artifact_checks[name]["sha256"] = current
        if not recorded:
            errors.append(f"清单未记录交付物哈希：{name}")
        elif recorded != current:
            errors.append(f"交付物哈希不一致：{name}")
    visual: dict[str, Any] = {"status": "not_run"}
    if reference is not None:
        reference = reference.expanduser().resolve()
        render_path = required["render_png"]
        if reference.is_file() and render_path.is_file():
            with Image.open(reference).convert("RGB") as ref_image, Image.open(render_path).convert("RGB") as rendered:
                if ref_image.size != rendered.size:
                    visual = {
                        "status": "canvas_mismatch",
                        "reference_size": list(ref_image.size),
                        "render_size": list(rendered.size),
                    }
                    warnings.append("参考图与重绘图画布不同；未通过缩放进行比较")
                else:
                    ref_array = np.asarray(ref_image, dtype=np.float32)
                    render_array = np.asarray(rendered, dtype=np.float32)
                    absolute = np.abs(ref_array - render_array)
                    mae = float(np.mean(absolute) / 255.0)
                    plot_mae: float | None = None
                    if project_path.is_file():
                        box = project.get("chart", {}).get("plot_box")
                        if (
                            isinstance(box, list)
                            and len(box) == 4
                            and all(isinstance(value, (int, float)) for value in box)
                        ):
                            left, top, right, bottom = [int(round(value)) for value in box]
                            if 0 <= left < right < ref_image.width and 0 <= top < bottom < ref_image.height:
                                plot_mae = float(
                                    np.mean(absolute[top : bottom + 1, left : right + 1]) / 255.0
                                )
                    heatmap = np.mean(absolute, axis=2)
                    maximum = float(np.max(heatmap))
                    normalized = (
                        np.clip(heatmap / maximum * 255.0, 0, 255).astype(np.uint8)
                        if maximum > 0
                        else np.zeros_like(heatmap, dtype=np.uint8)
                    )
                    heatmap_path = project_dir / "residual-heatmap.png"
                    Image.fromarray(normalized, mode="L").save(heatmap_path)
                    visual = {
                        "status": "measured",
                        "normalized_mae": mae,
                        "plot_box_normalized_mae": plot_mae,
                        "canvas": list(ref_image.size),
                        "residual_heatmap": str(heatmap_path),
                        "说明": "残差指标用于定位差异，不单独决定科学验收结论。",
                    }
        else:
            warnings.append("参考图或重绘 PNG 缺失")
    extraction = manifest.get("extraction_status", "not_run")
    review_status = manifest.get("review_status", "not_run")
    render_status = manifest.get("render_status", "not_run")
    if errors:
        delivery = "failed"
    elif extraction == "not_applicable" and render_status == "pass":
        delivery = "pass"
    elif extraction == "pass" and review_status == "accepted" and render_status == "pass":
        delivery = "pass"
    elif (
        extraction in {"pass", "partial"}
        and review_status in {"accepted", "partial"}
        and render_status == "pass"
    ):
        delivery = "partial"
    else:
        delivery = "failed"

    def bounded_delivery_score(value: float) -> float:
        return round(max(0.0, min(100.0, float(value))), 1)

    verified_artifacts = sum(
        bool(check.get("exists"))
        and bool(check.get("sha256"))
        and manifest.get("artifacts", {}).get(name, {}).get("sha256")
        == check.get("sha256")
        for name, check in artifact_checks.items()
    )
    artifact_integrity_score = bounded_delivery_score(
        verified_artifacts / max(1, len(artifact_checks)) * 100.0
    )
    render_formats = ("render_png", "render_svg", "render_pdf")
    complete_formats = sum(
        bool(artifact_checks.get(name, {}).get("exists"))
        and Path(str(artifact_checks[name]["path"])).stat().st_size > 0
        for name in render_formats
    )
    format_completeness_score = bounded_delivery_score(
        complete_formats / len(render_formats) * 100.0
    )

    render_report_path = required["render_report"]
    render_report = read_json(render_report_path) if render_report_path.is_file() else {}
    mapping = render_report.get("mapping", {}) if isinstance(render_report, dict) else {}
    traceability_checks = {
        "formal_data_present": required["data"].is_file(),
        "project_spec_present": project_path.is_file(),
        "render_report_pass": render_report.get("status") == "pass",
        "positive_rendered_rows": isinstance(render_report.get("rows"), int)
        and int(render_report.get("rows", 0)) > 0,
        "declared_xy_mapping": bool(mapping.get("x")) and bool(mapping.get("y")),
    }
    data_geometry_traceability_score = bounded_delivery_score(
        sum(traceability_checks.values()) / len(traceability_checks) * 100.0
    )

    declared_render = project.get("render", {}) if isinstance(project, dict) else {}
    expected_plot_type = str(declared_render.get("plot_type", ""))
    expected_x_scale = str(
        declared_render.get("x_scale")
        or project.get("chart", {}).get("x_axis", {}).get("scale", "linear")
    )
    expected_y_scale = str(
        declared_render.get("y_scale")
        or (
            project.get("chart", {})
            .get("polar", {})
            .get("radius_axis", {})
            .get("scale", "linear")
            if expected_plot_type == "polar_line"
            else project.get("chart", {}).get("y_axis", {}).get("scale", "linear")
        )
    )
    reported_scales = (
        render_report.get("axis_scales", {}) if isinstance(render_report, dict) else {}
    )
    expected_canvas = declared_render.get("canvas_px")
    reported_canvas = render_report.get("canvas_px") if isinstance(render_report, dict) else None
    spec_checks = {
        "plot_type_matches": bool(expected_plot_type)
        and render_report.get("plot_type") == expected_plot_type,
        "x_mapping_matches": bool(declared_render.get("x"))
        and mapping.get("x") == declared_render.get("x"),
        "y_mapping_matches": bool(declared_render.get("y"))
        and mapping.get("y") == declared_render.get("y"),
        "x_scale_matches": reported_scales.get("x")
        == ("circular_degree" if expected_plot_type == "polar_line" else expected_x_scale),
        "y_scale_matches": reported_scales.get("y") == expected_y_scale,
        "canvas_matches_when_declared": expected_canvas is None
        or expected_canvas == []
        or reported_canvas == expected_canvas,
    }
    render_spec_compliance_score = bounded_delivery_score(
        sum(spec_checks.values()) / len(spec_checks) * 100.0
    )
    delivery_dimensions = {
        "artifact_integrity": {
            "label": "交付物与哈希完整性",
            "weight": 30,
            "score": artifact_integrity_score,
            "metrics": {
                "verified_artifacts": verified_artifacts,
                "required_artifacts": len(artifact_checks),
            },
        },
        "format_completeness": {
            "label": "PNG/SVG/PDF 格式完整性",
            "weight": 20,
            "score": format_completeness_score,
            "metrics": {
                name: artifact_checks.get(name, {}).get("exists", False)
                for name in render_formats
            },
        },
        "data_geometry_traceability": {
            "label": "数据到图形可追溯性",
            "weight": 25,
            "score": data_geometry_traceability_score,
            "metrics": traceability_checks,
        },
        "render_spec_compliance": {
            "label": "重绘规格符合度",
            "weight": 25,
            "score": render_spec_compliance_score,
            "metrics": spec_checks,
        },
    }
    delivery_score = round(
        sum(
            float(item["score"]) * float(item["weight"])
            for item in delivery_dimensions.values()
        )
        / sum(float(item["weight"]) for item in delivery_dimensions.values()),
        1,
    )
    delivery_minimum_dimension_score = min(
        float(item["score"]) for item in delivery_dimensions.values()
    )
    profile_name = str(
        project.get("assessment", {}).get("acceptance_profile", "engineering")
        if isinstance(project.get("assessment", {}), dict)
        else "engineering"
    )
    profile = ACCEPTANCE_PROFILES.get(profile_name, ACCEPTANCE_PROFILES["engineering"])
    delivery_hard_gates = {
        "artifact_hashes_valid": artifact_integrity_score == 100.0,
        "required_formats_complete": format_completeness_score == 100.0,
        "stage_status_allows_delivery": delivery in {"pass", "partial"},
    }
    delivery_hard_gates_pass = all(delivery_hard_gates.values())
    delivery_threshold_pass = delivery_score >= float(profile["overall_threshold"])
    delivery_dimension_floor_pass = delivery_minimum_dimension_score >= float(
        profile["minimum_dimension_score"]
    )
    if not delivery_hard_gates_pass:
        delivery_decision = "blocked"
        delivery_next_instruction = "停止交付：先修复缺失文件、哈希或阶段状态，再重新验证。"
    elif not delivery_threshold_pass or not delivery_dimension_floor_pass:
        delivery_decision = "not_qualified"
        delivery_next_instruction = "暂不接受重绘：优先修复最低分维度，重新生成后再次 validate。"
    elif delivery == "partial":
        delivery_decision = "conditional_acceptance"
        delivery_next_instruction = "交付仅部分通过：请用户确认已声明的提取或复核限制后再决定是否使用。"
    else:
        delivery_decision = "accepted"
        delivery_next_instruction = "重绘技术交付达到所选用途阈值；可接受当前版本，或按需进行视觉审美复核。"
    delivery_assessment = {
        "method": "weighted_render_delivery_dimensions_with_non_compensable_hard_gates",
        "profile": profile_name,
        "profile_label": profile["label"],
        "overall_score": delivery_score,
        "minimum_dimension_score": round(delivery_minimum_dimension_score, 1),
        "thresholds": {
            "overall_score": profile["overall_threshold"],
            "minimum_dimension_score": profile["minimum_dimension_score"],
        },
        "dimensions": delivery_dimensions,
        "worst_dimension": min(
            delivery_dimensions,
            key=lambda key: float(delivery_dimensions[key]["score"]),
        ),
        "hard_gates": delivery_hard_gates,
        "hard_gates_pass": delivery_hard_gates_pass,
        "score_threshold_pass": delivery_threshold_pass,
        "dimension_floor_pass": delivery_dimension_floor_pass,
        "decision": delivery_decision,
        "next_instruction": delivery_next_instruction,
        "visual_reference_diagnostic": visual,
        "qualification_note": (
            "此分数评价文件、哈希、格式、数据映射和规格执行；参考图像素残差仅用于定位差异，"
            "不等同于科研数据准确率或视觉审美评分。"
        ),
    }
    report = {
        "schema": "more-sci-figure.validation-report.v1",
        "status": "pass" if not errors else "failed",
        "delivery_status": delivery,
        "stage_statuses": {
            "extraction_status": extraction,
            "review_status": review_status,
            "render_status": render_status,
        },
        "artifact_checks": artifact_checks,
        "visual_comparison": visual,
        "delivery_assessment": delivery_assessment,
        "errors": errors,
        "warnings": warnings,
    }
    report_path = project_dir / "validation-report.json"
    write_json(report_path, report)
    manifest["delivery_status"] = delivery
    if visual.get("residual_heatmap"):
        manifest["artifacts"]["residual_heatmap"] = artifact_entry(
            Path(str(visual["residual_heatmap"]))
        )
    manifest["artifacts"]["validation_report"] = artifact_entry(report_path)
    write_json(manifest_path, manifest)
    return report


def pipeline_command(
    spec_path: Path,
    out_dir: Path,
    review_decisions: Path | None = None,
) -> dict[str, Any]:
    extraction = extract_command(spec_path, out_dir)
    if review_decisions is None:
        assessment = read_json(out_dir / "review-assessment.json")
        recommendation = str(assessment.get("recommended_action", "stop"))
        acceptance = assessment.get("acceptance", {})
        review_waiting_status = {
            "batch_confirm": "awaiting_confirmation",
            "apply_review": "awaiting_review_apply",
            "review_anomaly_groups": "awaiting_anomaly_review",
            "re_extract": "re_extract_required",
            "stop": "blocked",
        }.get(recommendation, "attention_required")
        return {
            "extraction": extraction["status"],
            "review": review_waiting_status,
            "render": "not_run",
            "delivery": "not_run",
            "用途等级": acceptance.get("profile_label", acceptance.get("profile")),
            "综合评分": assessment.get("overall_score"),
            "最低维度分": assessment.get("minimum_dimension_score"),
            "判定": acceptance.get("decision"),
            "硬门通过": acceptance.get("hard_gates_pass"),
            "责任分工": acceptance.get("responsibility"),
            "风险等级": assessment.get("risk_level"),
            "异常组": assessment.get("anomaly_groups"),
            "建议动作": recommendation,
            "综合评估": str(out_dir / "review-assessment.json"),
            "下一步": acceptance.get("next_instruction", "请检查综合评估后再继续。"),
        }
    review = review_apply_command(out_dir, review_decisions)
    if review["review_status"] == "rejected":
        return {
            "extraction": extraction["status"],
            "review": "rejected",
            "render": "not_run",
            "delivery": "failed",
        }
    rendering = render_command(out_dir / "project.json", out_dir / "data.csv", out_dir / "render")
    validation = validate_command(out_dir)
    return {
        "extraction": extraction["status"],
        "review": review["review_status"],
        "render": rendering["status"],
        "delivery": validation["delivery_status"],
    }


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="可审计的科研图表数据提取、人工复核、重绘与验证。")
    parser.add_argument(
        "--version",
        action="version",
        version=f"%(prog)s {VERSION}",
        help="显示版本并退出",
    )
    subparsers = parser.add_subparsers(dest="command", required=True)

    inspect_parser = subparsers.add_parser("inspect", help="检查来源并生成项目模板")
    inspect_parser.add_argument("--input", required=True, type=Path, help="图片、PDF 或数据文件")
    inspect_parser.add_argument("--chart-type", choices=sorted(SUPPORTED_CHARTS), help="图表类型")
    inspect_parser.add_argument("--page", type=int, default=1, help="PDF 页码，从 1 开始")
    inspect_parser.add_argument("--dpi", type=int, default=144, help="PDF 整页测量栅格 DPI")
    inspect_parser.add_argument(
        "--pdf-image-index",
        type=int,
        help="直接锁定该页从 0 开始的嵌入栅格对象，避免整页重采样",
    )
    inspect_parser.add_argument("--out-dir", required=True, type=Path, help="输出目录")

    spec_review_parser = subparsers.add_parser(
        "spec-review", help="生成提取前规格叠图并等待用户确认"
    )
    spec_review_parser.add_argument("--spec", required=True, type=Path, help="项目规格 project.json")
    spec_review_parser.add_argument("--out-dir", required=True, type=Path, help="证据目录")

    spec_confirm_parser = subparsers.add_parser(
        "spec-confirm", help="记录用户对原图、绘图区、锚点、系列和排除规则的确认"
    )
    spec_confirm_parser.add_argument("--spec", required=True, type=Path, help="已展示的项目规格")
    spec_confirm_parser.add_argument("--project-dir", required=True, type=Path, help="证据目录")
    spec_confirm_parser.add_argument("--confirmed-by", required=True, help="确认人或可追溯身份")
    spec_confirm_parser.add_argument("--confirmation", required=True, help="用户的原始确认语句")

    extract_parser = subparsers.add_parser("extract", help="提取可见候选值并生成复核页面")
    extract_parser.add_argument("--spec", required=True, type=Path, help="项目规格 project.json")
    extract_parser.add_argument("--out-dir", required=True, type=Path, help="证据目录")

    review_parser = subparsers.add_parser("review", help="重新生成本地人工复核页面")
    review_parser.add_argument("--project-dir", required=True, type=Path, help="证据目录")

    assess_parser = subparsers.add_parser(
        "review-assess", help="综合评估全部候选并输出评分、系列摘要和异常组"
    )
    assess_parser.add_argument("--project-dir", required=True, type=Path, help="证据目录")

    confirm_parser = subparsers.add_parser(
        "review-confirm", help="把用户对话确认转换为覆盖全部候选的批量复核记录"
    )
    confirm_parser.add_argument("--project-dir", required=True, type=Path, help="证据目录")
    confirm_parser.add_argument("--reviewed-by", required=True, help="复核人或可追溯身份")
    confirm_parser.add_argument("--confirmation", required=True, help="用户的原始确认语句")
    confirm_parser.add_argument(
        "--accept-anomalies",
        action="store_true",
        help="仅在用户明确授权复核人查看证据并接受全部异常候选时使用",
    )

    serve_parser = subparsers.add_parser(
        "review-serve", help="启动仅本机复核会话并固定保存到当前项目目录"
    )
    serve_parser.add_argument("--project-dir", required=True, type=Path, help="证据目录")
    serve_parser.add_argument(
        "--port", type=int, default=0, help="本机端口；0 表示自动选择可用端口"
    )

    apply_parser = subparsers.add_parser("review-apply", help="应用人工复核决定并生成正式 data.csv")
    apply_parser.add_argument("--project-dir", required=True, type=Path, help="证据目录")
    apply_parser.add_argument("--decisions", required=True, type=Path, help="复核决定 JSON")

    render_parser = subparsers.add_parser("render", help="把声明数据重绘为 PNG、SVG 和 PDF")
    render_parser.add_argument("--spec", required=True, type=Path, help="项目规格")
    render_parser.add_argument("--data", required=True, type=Path, help="正式数据或外部数据")
    render_parser.add_argument("--out-dir", required=True, type=Path, help="重绘输出目录")

    preview_parser = subparsers.add_parser(
        "preview", help="把未复核候选值绘制为带水印预览，不推进正式状态"
    )
    preview_parser.add_argument("--spec", required=True, type=Path, help="项目规格")
    preview_parser.add_argument(
        "--candidates", required=True, type=Path, help="extract 产生的 candidates.csv"
    )
    preview_parser.add_argument("--out-dir", required=True, type=Path, help="候选预览输出目录")

    validate_parser = subparsers.add_parser("validate", help="验证交付物、哈希、状态和可选图像残差")
    validate_parser.add_argument("--project-dir", required=True, type=Path, help="证据目录")
    validate_parser.add_argument("--reference", type=Path, help="可选参考图片")

    pipeline_parser = subparsers.add_parser("pipeline", help="按门控顺序执行提取、复核、重绘和验证")
    pipeline_parser.add_argument("--spec", required=True, type=Path, help="项目规格")
    pipeline_parser.add_argument("--out-dir", required=True, type=Path, help="证据目录")
    pipeline_parser.add_argument(
        "--review-decisions",
        type=Path,
        help="人工复核决定；缺省时管线输出综合评分并停在 awaiting_confirmation",
    )
    for current in (
        parser,
        inspect_parser,
        spec_review_parser,
        spec_confirm_parser,
        extract_parser,
        review_parser,
        assess_parser,
        confirm_parser,
        serve_parser,
        apply_parser,
        render_parser,
        preview_parser,
        validate_parser,
        pipeline_parser,
    ):
        current._positionals.title = "位置参数"
        current._optionals.title = "选项"
        for action in current._actions:
            if isinstance(action, argparse._HelpAction):
                action.help = "显示帮助并退出"
    return parser


def main(argv: list[str] | None = None) -> int:
    args = build_parser().parse_args(argv)
    try:
        if args.command == "inspect":
            inspect_command(
                args.input,
                args.chart_type,
                args.out_dir,
                args.page,
                dpi=args.dpi,
                pdf_image_index=args.pdf_image_index,
            )
            return 0
        if args.command == "extract":
            result = extract_command(args.spec, args.out_dir)
        elif args.command == "spec-review":
            result = spec_review_command(args.spec, args.out_dir)
        elif args.command == "spec-confirm":
            result = spec_confirm_command(
                args.spec, args.project_dir, args.confirmed_by, args.confirmation
            )
        elif args.command == "review":
            result = review_command(args.project_dir)
        elif args.command == "review-assess":
            result = review_assess_command(args.project_dir)
        elif args.command == "review-confirm":
            result = review_confirm_command(
                args.project_dir,
                args.reviewed_by,
                args.confirmation,
                accept_anomalies=args.accept_anomalies,
            )
        elif args.command == "review-serve":
            result = review_serve_command(args.project_dir, args.port)
        elif args.command == "review-apply":
            result = review_apply_command(args.project_dir, args.decisions)
        elif args.command == "render":
            result = render_command(args.spec, args.data, args.out_dir)
        elif args.command == "preview":
            result = preview_command(args.spec, args.candidates, args.out_dir)
        elif args.command == "validate":
            result = validate_command(args.project_dir, args.reference)
        else:
            result = pipeline_command(args.spec, args.out_dir, args.review_decisions)
        print(json.dumps(result, ensure_ascii=False, indent=2, sort_keys=True))
        return 0
    except FigureError as exc:
        print(json.dumps({"status": "failed", "error": str(exc)}, ensure_ascii=False), file=sys.stderr)
        return 2


if __name__ == "__main__":
    raise SystemExit(main())
